from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from datetime import datetime
from io import BytesIO
from typing import Any
import base64
import json
import time
import urllib.error
import urllib.request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from sqlalchemy.orm import Session
from backend.db.database import get_db
from backend.db import models, schemas
from backend.services.entitlement_service import get_client_entitlements
from backend.services.rbac import require_roles

router = APIRouter(
    prefix="/client-config",
    tags=["Client Config"],
    dependencies=[Depends(require_roles("super_admin"))],
)


def _sync_key_from_connection(row: models.ClientConnection) -> str | None:
    config_json = row.config_json or {}
    sync_object = str(config_json.get("sync_object") or "").strip().upper()
    if sync_object in {"UOM", "ADDRESS"}:
        return sync_object
    message_type = str(getattr(row, "message_type", "") or "").strip().upper()
    if "UOM" in message_type:
        return "UOM"
    if "ADDRESS" in message_type:
        return "ADDRESS"
    connection_name = str(getattr(row, "connection_name", "") or "").strip().upper()
    if "UOM" in connection_name:
        return "UOM"
    if "ADDRESS" in connection_name:
        return "ADDRESS"
    return None


def _sync_key_from_erp(row: models.ClientERPConfig) -> str | None:
    message_type = str(getattr(row, "message_type", "") or "").strip().upper()
    if "UOM" in message_type:
        return "UOM"
    if "ADDRESS" in message_type:
        return "ADDRESS"
    return None


def _sync_status_for_key(db: Session, client_id: str, sync_key: str) -> str:
    has_active_connection = (
        db.query(models.ClientConnection)
        .filter(models.ClientConnection.client_id == client_id)
        .filter(models.ClientConnection.is_active.is_(True))
        .filter(
            (models.ClientConnection.config_json["sync_object"].astext.ilike(sync_key))
            | (models.ClientConnection.connection_name.ilike(f"%{sync_key}%"))
        )
        .first()
        is not None
    )
    has_active_erp = (
        db.query(models.ClientERPConfig)
        .filter(models.ClientERPConfig.client_id == client_id)
        .filter(models.ClientERPConfig.is_active.is_(True))
        .filter(models.ClientERPConfig.message_type.ilike(f"%{sync_key}%"))
        .first()
        is not None
    )
    if has_active_connection and has_active_erp:
        return "READY"
    if has_active_connection or has_active_erp:
        return "CONFIGURED"
    return "NOT CONFIGURED"


def _record_sync_event(
    db: Session,
    *,
    client_id: str,
    sync_key: str,
    event_type: str,
    status: str,
    message: str | None = None,
    endpoint_url: str | None = None,
    source_system: str | None = None,
    target_system: str | None = None,
    records_synced: int = 0,
    duration_ms: int | None = None,
    last_synced_at: datetime | None = None,
    details_json: dict[str, Any] | None = None,
) -> None:
    try:
        row = models.ClientSyncEvent(
            client_id=client_id,
            sync_key=sync_key,
            event_type=event_type,
            status=status,
            message=message,
            endpoint_url=endpoint_url,
            source_system=source_system,
            target_system=target_system,
            records_synced=records_synced,
            duration_ms=duration_ms,
            last_synced_at=last_synced_at or datetime.utcnow(),
            details_json=details_json or {},
        )
        db.add(row)
        db.commit()
    except Exception:
        db.rollback()


def _sync_key_from_connection(row: models.ClientConnection) -> str | None:
    config_json = row.config_json or {}
    sync_object = str(config_json.get("sync_object") or "").strip().upper()
    if sync_object in {"UOM", "ADDRESS"}:
        return sync_object
    message_type = str(getattr(row, "message_type", "") or "").strip().upper()
    if "UOM" in message_type:
        return "UOM"
    if "ADDRESS" in message_type:
        return "ADDRESS"
    connection_name = str(getattr(row, "connection_name", "") or "").strip().upper()
    if "UOM" in connection_name:
        return "UOM"
    if "ADDRESS" in connection_name:
        return "ADDRESS"
    return None


def _sync_key_from_erp(row: models.ClientERPConfig) -> str | None:
    message_type = str(getattr(row, "message_type", "") or "").strip().upper()
    if "UOM" in message_type:
        return "UOM"
    if "ADDRESS" in message_type:
        return "ADDRESS"
    return None


def _sync_status_for_key(db: Session, client_id: str, sync_key: str) -> str:
    has_active_connection = (
        db.query(models.ClientConnection)
        .filter(models.ClientConnection.client_id == client_id)
        .filter(models.ClientConnection.is_active.is_(True))
        .filter(
            (models.ClientConnection.config_json["sync_object"].astext.ilike(sync_key))
            | (models.ClientConnection.connection_name.ilike(f"%{sync_key}%"))
        )
        .first()
        is not None
    )
    has_active_erp = (
        db.query(models.ClientERPConfig)
        .filter(models.ClientERPConfig.client_id == client_id)
        .filter(models.ClientERPConfig.is_active.is_(True))
        .filter(models.ClientERPConfig.message_type.ilike(f"%{sync_key}%"))
        .first()
        is not None
    )
    if has_active_connection and has_active_erp:
        return "READY"
    if has_active_connection or has_active_erp:
        return "CONFIGURED"
    return "NOT CONFIGURED"


def _record_sync_event(
    db: Session,
    *,
    client_id: str,
    sync_key: str,
    event_type: str,
    status: str,
    message: str | None = None,
    endpoint_url: str | None = None,
    source_system: str | None = None,
    target_system: str | None = None,
    records_synced: int = 0,
    duration_ms: int | None = None,
    last_synced_at: datetime | None = None,
    details_json: dict[str, Any] | None = None,
) -> None:
    try:
        row = models.ClientSyncEvent(
            client_id=client_id,
            sync_key=sync_key,
            event_type=event_type,
            status=status,
            message=message,
            endpoint_url=endpoint_url,
            source_system=source_system,
            target_system=target_system,
            records_synced=records_synced,
            duration_ms=duration_ms,
            last_synced_at=last_synced_at or datetime.utcnow(),
            details_json=details_json or {},
        )
        db.add(row)
        db.commit()
    except Exception:
        db.rollback()


def _build_sync_url(config_json: dict[str, Any], sync_key: str) -> str | None:
    key = sync_key.upper()
    endpoint = (
        config_json.get("endpoint_url")
        or config_json.get("webhook_url")
        or config_json.get("endpoint")
        or config_json.get(f"{key.lower()}_endpoint_url")
        or config_json.get(f"{key.lower()}_webhook_url")
    )
    if not endpoint:
        return None
    path = (
        config_json.get("resource_path")
        or config_json.get("path")
        or config_json.get(f"{key.lower()}_path")
        or ""
    )
    endpoint = str(endpoint).rstrip("/")
    path = str(path).strip()
    if not path:
        return endpoint
    if not path.startswith("/"):
        path = f"/{path}"
    return f"{endpoint}{path}"


def _build_sync_headers(config_json: dict[str, Any]) -> dict[str, str]:
    auth_type = str(config_json.get("auth_type") or "").strip().upper()
    headers = {"Accept": "application/json, text/plain, */*"}
    token = config_json.get("token") or config_json.get("bearer_token") or config_json.get("access_token") or config_json.get("api_token")
    username = config_json.get("username")
    password = config_json.get("password_token") or config_json.get("password")
    if auth_type in {"BEARER", "TOKEN"} and token:
        headers["Authorization"] = f"Bearer {token}"
    elif auth_type == "BASIC" and username and password:
        raw = f"{username}:{password}".encode("utf-8")
        headers["Authorization"] = "Basic " + base64.b64encode(raw).decode("ascii")
    elif auth_type in {"API_KEY", "X_API_KEY"} and token:
        headers["X-API-Key"] = str(token)
    elif token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def _parse_records_synced(response_bytes: bytes, content_type: str) -> int:
    if "json" not in content_type.lower():
        return 0
    try:
        payload = json.loads(response_bytes.decode("utf-8"))
    except Exception:
        return 0
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict):
        for key in ("records_synced", "rows_synced", "row_count", "count", "synced"):
            value = payload.get(key)
            if isinstance(value, int):
                return value
            if isinstance(value, str) and value.isdigit():
                return int(value)
        for key in ("records", "items", "data", "results"):
            value = payload.get(key)
            if isinstance(value, list):
                return len(value)
        return 1
    return 0


STOREFRONT_TEMPLATE_HEADERS = [
    "sku",
    "name",
    "description",
    "details",
    "category",
    "brand",
    "unit_price",
    "currency",
    "uom",
    "stock_status",
    "lead_time",
    "min_order_qty",
    "moq_uom",
    "payment_terms",
    "discount_mode",
    "discount_value",
    "tax_mode",
    "tax_value",
    "freight_mode",
    "freight_value",
    "octroi_mode",
    "octroi_value",
    "shipping_mode",
    "shipping_value",
    "image_url",
    "video_url",
    "media_urls",
    "specifications",
]

STOREFRONT_TEMPLATE_SAMPLE_ROW = {
    "sku": "SKU-1001",
    "name": "Industrial Product",
    "description": "Short buyer-facing summary",
    "details": "Longer storefront description for product pages and approvals.",
    "category": "Industrial Supplies",
    "brand": "Supplier Brand",
    "unit_price": 125.50,
    "currency": "USD",
    "uom": "EA",
    "stock_status": "Available",
    "lead_time": "2-3 business days",
    "min_order_qty": 10,
    "moq_uom": "EA",
    "payment_terms": "Net 30",
    "discount_mode": "PERCENT",
    "discount_value": 5,
    "tax_mode": "PERCENT",
    "tax_value": 18,
    "freight_mode": "AMOUNT",
    "freight_value": 12,
    "octroi_mode": "NONE",
    "octroi_value": "",
    "shipping_mode": "AMOUNT",
    "shipping_value": 8,
    "image_url": "",
    "video_url": "",
    "media_urls": "",
    "specifications": "Color: Blue; Size: Medium; Material: Polymer",
}

STOREFRONT_TEMPLATE_GUIDANCE = [
    (
        "sku",
        "Yes",
        "Unique product code used in the storefront and uploads.",
        "Uppercase code or your internal SKU format.",
        "SKU-1001",
    ),
    (
        "name",
        "Yes",
        "Buyer-facing product name.",
        "Short title.",
        "Industrial Product",
    ),
    (
        "description",
        "No",
        "Short summary shown in cards and previews.",
        "Plain text.",
        "Short buyer-facing summary",
    ),
    (
        "details",
        "No",
        "Detailed product description for the storefront page.",
        "Plain text.",
        "Longer storefront description for product pages and approvals.",
    ),
    (
        "unit_price",
        "No",
        "Numeric selling price.",
        "Number only.",
        "125.50",
    ),
    (
        "currency",
        "No",
        "3-letter ISO currency code.",
        "ISO code.",
        "USD",
    ),
    (
        "discount_mode",
        "No",
        "How to apply discount to the base unit price.",
        "Use NONE, PERCENT, or AMOUNT.",
        "PERCENT",
    ),
    (
        "discount_value",
        "No",
        "Discount numeric value.",
        "If mode=PERCENT, enter 5 for 5 percent. If mode=AMOUNT, enter flat amount per line item.",
        "5",
    ),
    (
        "tax_mode",
        "No",
        "How tax should be applied in buyer calculations.",
        "Use NONE, PERCENT, or AMOUNT.",
        "PERCENT",
    ),
    (
        "tax_value",
        "No",
        "Tax numeric value.",
        "If mode=PERCENT, enter 18 for 18 percent. If mode=AMOUNT, enter flat amount per line item.",
        "18",
    ),
    (
        "freight_mode",
        "No",
        "How freight should be applied.",
        "Use NONE, PERCENT, or AMOUNT.",
        "AMOUNT",
    ),
    (
        "freight_value",
        "No",
        "Freight numeric value.",
        "If mode=AMOUNT, enter flat freight per line item.",
        "12",
    ),
    (
        "octroi_mode",
        "No",
        "How octroi or local duty should be applied.",
        "Use NONE, PERCENT, or AMOUNT.",
        "NONE",
    ),
    (
        "octroi_value",
        "No",
        "Octroi numeric value.",
        "Leave blank when mode=NONE.",
        "",
    ),
    (
        "shipping_mode",
        "No",
        "How shipping should be applied.",
        "Use NONE, PERCENT, or AMOUNT.",
        "AMOUNT",
    ),
    (
        "shipping_value",
        "No",
        "Shipping numeric value.",
        "If mode=AMOUNT, enter flat shipping per line item.",
        "8",
    ),
    (
        "image_url",
        "No",
        "Primary product image URL.",
        "Use an Ordanex upload URL like /files/<file_id>/download or a public HTTPS URL.",
        "/files/abc123/download",
    ),
    (
        "video_url",
        "No",
        "Primary product video URL.",
        "Use an Ordanex upload URL or a public HTTPS MP4/WebM URL.",
        "https://media.yourdomain.com/catalog/sku-1001-demo.mp4",
    ),
    (
        "media_urls",
        "No",
        "Additional gallery media URLs.",
        "Comma-separated list of Ordanex file URLs or public HTTPS URLs.",
        "/files/abc123/download, https://media.yourdomain.com/catalog/sku-1001-alt.jpg",
    ),
    (
        "specifications",
        "No",
        "Attribute list shown to buyers.",
        "Separate entries with semicolons using Key: Value format.",
        "Color: Blue; Size: Medium",
    ),
]


def _build_storefront_catalog_template() -> bytes:
    workbook = Workbook()
    catalog_sheet = workbook.active
    catalog_sheet.title = "Catalog Upload"
    guidance_sheet = workbook.create_sheet("Field Guidance")

    header_fill = PatternFill(fill_type="solid", fgColor="D9E7FF")
    header_font = Font(bold=True, color="153E8A")
    wrapped_alignment = Alignment(wrap_text=True, vertical="top")

    for column_index, header in enumerate(STOREFRONT_TEMPLATE_HEADERS, start=1):
        cell = catalog_sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrapped_alignment

        sample_cell = catalog_sheet.cell(
            row=2,
            column=column_index,
            value=STOREFRONT_TEMPLATE_SAMPLE_ROW.get(header, ""),
        )
        sample_cell.alignment = wrapped_alignment

    catalog_sheet.freeze_panes = "A2"
    catalog_sheet.auto_filter.ref = f"A1:AB202"
    catalog_sheet.sheet_view.showGridLines = True

    catalog_widths = {
        "A": 18,
        "B": 26,
        "C": 28,
        "D": 42,
        "E": 20,
        "F": 18,
        "G": 14,
        "H": 12,
        "I": 12,
        "J": 18,
        "K": 18,
        "L": 14,
        "M": 12,
        "N": 18,
        "O": 16,
        "P": 16,
        "Q": 16,
        "R": 16,
        "S": 16,
        "T": 16,
        "U": 16,
        "V": 16,
        "W": 16,
        "X": 16,
        "Y": 42,
        "Z": 42,
        "AA": 52,
        "AB": 34,
    }
    for column_letter, width in catalog_widths.items():
        catalog_sheet.column_dimensions[column_letter].width = width

    guidance_sheet.append(
        [
            "Header",
            "Required",
            "What to enter",
            "Accepted format",
            "Example",
        ]
    )
    for cell in guidance_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrapped_alignment
        cell.protection = Protection(locked=True)

    guidance_rows = [
        (
            "Media workflow",
            "",
            "Upload product images or videos inside the Ordanex Product media manager first, then copy the returned file URL into image_url, video_url, or media_urls. Ordanex stores the file behind the scenes and exposes a reusable URL like /files/<file_id>/download.",
            "Use Ordanex file URLs or public HTTPS URLs.",
            "/files/abc123/download",
        ),
        (
            "Header handling",
            "",
            "Keep the first row unchanged and enter product data from row 2 onward. The catalog sheet stays editable so business users can paste or update products without workbook protection prompts.",
            "Use the provided headers exactly as shown.",
            "Replace or extend the sample row.",
        ),
    ]
    for row in guidance_rows + STOREFRONT_TEMPLATE_GUIDANCE:
        guidance_sheet.append(row)

    for row in guidance_sheet.iter_rows():
        for cell in row:
            cell.alignment = wrapped_alignment

    for column_letter, width in {"A": 24, "B": 12, "C": 52, "D": 56, "E": 40}.items():
        guidance_sheet.column_dimensions[column_letter].width = width

    guidance_sheet.freeze_panes = "A2"
    guidance_sheet.protection.sheet = True

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()



@router.get("/clients", response_model=list[schemas.ClientRead])
def get_clients(db: Session = Depends(get_db)):
    return db.query(models.Client).order_by(models.Client.client_name.asc()).all()

@router.post("/clients", response_model=schemas.ClientRead)
def create_client(payload: schemas.ClientCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Client).filter(models.Client.client_id == payload.client_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Client already exists")
    row = models.Client(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row

@router.put("/clients/{client_id}", response_model=schemas.ClientRead)
def update_client(client_id: str, payload: schemas.ClientUpdate, db: Session = Depends(get_db)):
    row = db.query(models.Client).filter(models.Client.client_id == client_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Client not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(row, field, value)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _client_profile_config_map(db: Session, client_id: str) -> dict[str, models.ClientConfig]:
    rows = (
        db.query(models.ClientConfig)
        .filter(models.ClientConfig.client_id == client_id)
        .filter(models.ClientConfig.config_type == "CLIENT_PROFILE")
        .filter(
            models.ClientConfig.config_key.in_(
                ["LEGAL_TAX", "BILLING_INVOICING", "BANKING_REMITTANCE"]
            )
        )
        .all()
    )
    return {str(row.config_key or "").upper(): row for row in rows}


def _profile_section_payload(row: models.ClientConfig | None) -> dict[str, Any]:
    return row.config_value_json if row and isinstance(row.config_value_json, dict) else {}


@router.get("/client-profile-details/{client_id}", response_model=schemas.ClientProfileDetailsRead)
def get_client_profile_details(client_id: str, db: Session = Depends(get_db)):
    client = db.query(models.Client).filter(models.Client.client_id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    rows = _client_profile_config_map(db, client_id)
    return {
        "client_id": client_id,
        "legal_tax": _profile_section_payload(rows.get("LEGAL_TAX")),
        "billing_invoicing": _profile_section_payload(rows.get("BILLING_INVOICING")),
        "banking_remittance": _profile_section_payload(rows.get("BANKING_REMITTANCE")),
    }


@router.put("/client-profile-details/{client_id}", response_model=schemas.ClientProfileDetailsRead)
def upsert_client_profile_details(
    client_id: str,
    payload: schemas.ClientProfileDetailsUpdate,
    db: Session = Depends(get_db),
):
    client = db.query(models.Client).filter(models.Client.client_id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    rows = _client_profile_config_map(db, client_id)
    sections = {
        "LEGAL_TAX": payload.legal_tax.model_dump(),
        "BILLING_INVOICING": payload.billing_invoicing.model_dump(),
        "BANKING_REMITTANCE": payload.banking_remittance.model_dump(),
    }

    for key, value in sections.items():
        row = rows.get(key)
        if row is None:
            row = models.ClientConfig(
                client_id=client_id,
                config_type="CLIENT_PROFILE",
                config_key=key,
                config_value_json={},
                is_active=True,
            )
            db.add(row)
        row.config_value_json = value
        row.is_active = True

    db.commit()
    rows = _client_profile_config_map(db, client_id)
    return {
        "client_id": client_id,
        "legal_tax": _profile_section_payload(rows.get("LEGAL_TAX")),
        "billing_invoicing": _profile_section_payload(rows.get("BILLING_INVOICING")),
        "banking_remittance": _profile_section_payload(rows.get("BANKING_REMITTANCE")),
    }


def _normalize_storefront_environment(value: str | None) -> str:
    normalized = str(value or "").strip().upper()
    if normalized in {"STAGING", "STAGE", "STG"}:
        return "STAGING"
    return "PROD"


def _buyer_storefront_access_keys(environment: str | None) -> list[str]:
    env = _normalize_storefront_environment(environment).lower()
    return [
        f"buyer_storefront_{env}",
        f"buyer-storefront-{env}",
        f"buyerstorefront_{env}",
        "buyer_storefront",
        "buyer-storefront",
        "buyerstorefront",
    ]


def _buyer_storefront_settings_keys(environment: str | None) -> list[str]:
    env = _normalize_storefront_environment(environment)
    return [f"SETTINGS_{env}", "SETTINGS"]


def _pick_config_row(rows: list[models.ClientConfig], keys: list[str]):
    row_map = {
        str(getattr(row, "config_key", "") or "").strip().lower(): row
        for row in rows
    }
    for key in keys:
        row = row_map.get(key.strip().lower())
        if row is not None:
            return row
    return None


def _buyer_storefront_access_row(db: Session, client_id: str, environment: str | None = None):
    rows = (
        db.query(models.ClientConfig)
        .filter(models.ClientConfig.client_id == client_id)
        .filter(models.ClientConfig.is_active.is_(True))
        .filter(models.ClientConfig.config_type.in_(["FEATURES", "FEATURE_FLAG", "BUYER_PORTAL"]))
        .order_by(models.ClientConfig.updated_at.desc(), models.ClientConfig.created_at.desc())
        .all()
    )
    return _pick_config_row(rows, _buyer_storefront_access_keys(environment))


def _ensure_buyer_storefront_access_row(db: Session, client_id: str, environment: str | None = None):
    row = _buyer_storefront_access_row(db, client_id, environment)
    expected_key = _buyer_storefront_access_keys(environment)[0]
    if row and str(getattr(row, "config_key", "") or "").strip().lower() == expected_key:
        return row
    row = models.ClientConfig(
        client_id=client_id,
        config_type="BUYER_PORTAL",
        config_key=expected_key,
        config_value_json={},
        is_active=True,
    )
    db.add(row)
    return row


def _buyer_storefront_settings_row(db: Session, client_id: str, environment: str | None = None):
    rows = (
        db.query(models.ClientConfig)
        .filter(models.ClientConfig.client_id == client_id)
        .filter(models.ClientConfig.is_active.is_(True))
        .filter(models.ClientConfig.config_type == "BUYER_PORTAL")
        .order_by(models.ClientConfig.updated_at.desc(), models.ClientConfig.created_at.desc())
        .all()
    )
    return _pick_config_row(rows, _buyer_storefront_settings_keys(environment))


def _ensure_buyer_storefront_settings_row(db: Session, client_id: str, environment: str | None = None):
    row = _buyer_storefront_settings_row(db, client_id, environment)
    expected_key = _buyer_storefront_settings_keys(environment)[0]
    if row and str(getattr(row, "config_key", "") or "").strip().upper() == expected_key:
        return row
    row = models.ClientConfig(
        client_id=client_id,
        config_type="BUYER_PORTAL",
        config_key=expected_key,
        config_value_json={},
        is_active=True,
    )
    db.add(row)
    return row


def _storefront_explicit_enabled(row: models.ClientConfig | None) -> bool | None:
    config = row.config_value_json if row and isinstance(row.config_value_json, dict) else {}
    if "enabled" in config:
        return bool(config.get("enabled"))
    if "disabled" in config:
        return not bool(config.get("disabled"))
    return None


@router.get("/buyer-storefront/{client_id}")
def get_buyer_storefront_access(
    client_id: str,
    environment: str | None = Query(None),
    db: Session = Depends(get_db),
):
    row = _buyer_storefront_access_row(db, client_id, environment)
    entitlements = get_client_entitlements(db, client_id)
    config = row.config_value_json if row and isinstance(row.config_value_json, dict) else {}
    explicit_enabled = _storefront_explicit_enabled(row)
    entitled = bool(entitlements.get("buyer_storefront"))
    enabled = entitled if explicit_enabled is None else (entitled and explicit_enabled)

    return {
        "client_id": client_id,
        "environment": _normalize_storefront_environment(environment),
        "subscription_type": entitlements.get("subscription_type"),
        "enabled": enabled,
        "source": entitlements.get("buyer_storefront_source"),
        "explicit_enabled": explicit_enabled,
        "config": config,
    }


@router.put("/buyer-storefront/{client_id}")
def update_buyer_storefront_access(
    client_id: str,
    payload: dict[str, Any],
    environment: str | None = Query(None),
    db: Session = Depends(get_db),
):
    if "enabled" not in payload:
        raise HTTPException(status_code=400, detail="'enabled' is required")

    enabled = bool(payload.get("enabled"))
    row = _ensure_buyer_storefront_access_row(db, client_id, environment)

    config = dict(row.config_value_json or {})
    config["enabled"] = enabled
    config["disabled"] = not enabled
    config["updated_at"] = datetime.utcnow().isoformat()
    if payload.get("note"):
        config["note"] = str(payload.get("note"))
    row.config_value_json = config
    row.is_active = True
    db.commit()
    db.refresh(row)

    entitlements = get_client_entitlements(db, client_id)
    entitled = bool(entitlements.get("buyer_storefront"))
    return {
        "client_id": client_id,
        "environment": _normalize_storefront_environment(environment),
        "subscription_type": entitlements.get("subscription_type"),
        "enabled": entitled and enabled,
        "source": entitlements.get("buyer_storefront_source"),
        "explicit_enabled": enabled,
        "config": row.config_value_json or {},
    }


@router.get("/buyer-storefront-settings/{client_id}")
def get_buyer_storefront_settings(
    client_id: str,
    environment: str | None = Query(None),
    db: Session = Depends(get_db),
):
    row = _buyer_storefront_settings_row(db, client_id, environment)
    settings = row.config_value_json if row and isinstance(row.config_value_json, dict) else {}
    return {
        "client_id": client_id,
        "environment": _normalize_storefront_environment(environment),
        "settings": settings,
    }


@router.put("/buyer-storefront-settings/{client_id}")
def update_buyer_storefront_settings(
    client_id: str,
    payload: dict[str, Any],
    environment: str | None = Query(None),
    db: Session = Depends(get_db),
):
    row = _ensure_buyer_storefront_settings_row(db, client_id, environment)

    current = row.config_value_json if isinstance(row.config_value_json, dict) else {}
    branding = payload.get("branding") if isinstance(payload.get("branding"), dict) else {}
    catalog = payload.get("catalog") if isinstance(payload.get("catalog"), dict) else {}
    commerce = payload.get("commerce") if isinstance(payload.get("commerce"), dict) else {}
    payments = payload.get("payments") if isinstance(payload.get("payments"), dict) else {}
    experience = payload.get("experience") if isinstance(payload.get("experience"), dict) else {}
    pricing = payload.get("pricing") if isinstance(payload.get("pricing"), dict) else {}
    access = payload.get("access") if isinstance(payload.get("access"), dict) else {}
    current = {
        **current,
        "branding": {**(current.get("branding") if isinstance(current.get("branding"), dict) else {}), **branding},
        "catalog": {**(current.get("catalog") if isinstance(current.get("catalog"), dict) else {}), **catalog},
        "commerce": {**(current.get("commerce") if isinstance(current.get("commerce"), dict) else {}), **commerce},
        "payments": {**(current.get("payments") if isinstance(current.get("payments"), dict) else {}), **payments},
        "experience": {**(current.get("experience") if isinstance(current.get("experience"), dict) else {}), **experience},
        "pricing": {**(current.get("pricing") if isinstance(current.get("pricing"), dict) else {}), **pricing},
    }
    if access:
        approved_buyers = access.get("approved_buyers") if isinstance(access.get("approved_buyers"), list) else []
        normalized_buyers = []
        seen_emails = set()
        for item in approved_buyers:
            email = ""
            if isinstance(item, dict):
                email = str(item.get("email") or item.get("buyer_email") or "").strip().lower()
            else:
                email = str(item or "").strip().lower()
            if not email or email in seen_emails:
                continue
            seen_emails.add(email)
            normalized_buyers.append({"email": email})
        current["access"] = {
            **(current.get("access") if isinstance(current.get("access"), dict) else {}),
            "approval_mode": str(access.get("approval_mode") or "EMAIL_APPROVAL").strip().upper(),
            "approved_buyers": normalized_buyers,
        }
    if isinstance(payload.get("banner_text"), str):
        current.setdefault("branding", {})["banner_text"] = payload.get("banner_text")
    row.config_value_json = current
    row.is_active = True
    db.commit()
    db.refresh(row)
    return {
        "client_id": client_id,
        "environment": _normalize_storefront_environment(environment),
        "settings": row.config_value_json if isinstance(row.config_value_json, dict) else {},
    }


@router.post("/buyer-storefront-settings/{client_id}/publish")
def publish_buyer_storefront_settings(
    client_id: str,
    from_environment: str = Query("STAGING"),
    to_environment: str = Query("PROD"),
    db: Session = Depends(get_db),
):
    source_environment = _normalize_storefront_environment(from_environment)
    target_environment = _normalize_storefront_environment(to_environment)
    if source_environment == target_environment:
        raise HTTPException(status_code=400, detail="Source and target storefront environments must be different.")

    source_settings = _buyer_storefront_settings_row(db, client_id, source_environment)
    if source_settings is None or not isinstance(source_settings.config_value_json, dict):
        raise HTTPException(status_code=404, detail="No storefront settings found for the source environment.")

    target_settings = _ensure_buyer_storefront_settings_row(db, client_id, target_environment)
    target_settings.config_value_json = json.loads(json.dumps(source_settings.config_value_json))
    target_settings.is_active = True

    source_access = _buyer_storefront_access_row(db, client_id, source_environment)
    if source_access is not None and isinstance(source_access.config_value_json, dict):
        target_access = _ensure_buyer_storefront_access_row(db, client_id, target_environment)
        target_access.config_value_json = json.loads(json.dumps(source_access.config_value_json))
        target_access.is_active = True

    db.commit()
    db.refresh(target_settings)

    return {
        "client_id": client_id,
        "source_environment": source_environment,
        "target_environment": target_environment,
        "settings": target_settings.config_value_json if isinstance(target_settings.config_value_json, dict) else {},
    }


DEFAULT_COMMERCIAL_PRIORITIES = [
    ("ERP_BUYER_CONTRACT", "Buyer-specific ERP contract"),
    ("ERP_SHIP_TO_RULE", "Ship-to jurisdiction ERP rule"),
    ("ERP_PRODUCT_MAPPING", "Product ERP mapping"),
    ("ORDANEX_OVERRIDE", "Ordanex storefront override"),
    ("CLIENT_DEFAULT", "Ordanex client default"),
    ("ZERO_FALLBACK", "Zero fallback"),
]


def _commercial_environment(environment: str | None) -> str:
    return _normalize_storefront_environment(environment)


def _commercial_setting_row(db: Session, client_id: str, environment: str | None = None):
    env = _commercial_environment(environment)
    return (
        db.query(models.ClientCommercialSetting)
        .filter(models.ClientCommercialSetting.client_id == client_id)
        .filter(models.ClientCommercialSetting.environment == env)
        .filter(models.ClientCommercialSetting.is_active.is_(True))
        .order_by(models.ClientCommercialSetting.updated_at.desc(), models.ClientCommercialSetting.created_at.desc())
        .first()
    )


def _ensure_commercial_setting_row(db: Session, client_id: str, environment: str | None = None):
    row = _commercial_setting_row(db, client_id, environment)
    if row:
        return row
    row = models.ClientCommercialSetting(
        client_id=client_id,
        environment=_commercial_environment(environment),
        source_mode="ORDANEX_MASTER",
        erp_sync_enabled=False,
        erp_sync_frequency="DAILY",
        currency_mode="CLIENT_DEFAULT",
        fallback_policy="ZERO_FALLBACK",
        is_active=True,
    )
    db.add(row)
    db.flush()
    return row


def _serialize_priority_rows(rows: list[models.ClientCommercialPriority]) -> list[dict[str, Any]]:
    payload = [
        {
            "priority_id": str(row.priority_id),
            "sequence_no": row.sequence_no,
            "priority_code": row.priority_code,
            "priority_label": row.priority_label,
            "source_system": row.source_system,
            "is_active": row.is_active,
        }
        for row in rows
    ]
    if payload:
        return payload
    return [
        {
            "priority_id": None,
            "sequence_no": index,
            "priority_code": code,
            "priority_label": label,
            "source_system": "SYSTEM",
            "is_active": True,
        }
        for index, (code, label) in enumerate(DEFAULT_COMMERCIAL_PRIORITIES, start=1)
    ]


def _serialize_charge_code_rows(rows: list[models.ClientChargeCode]) -> list[dict[str, Any]]:
    return [
        {
            "charge_code_id": str(row.charge_code_id),
            "charge_code": row.charge_code,
            "charge_type": row.charge_type,
            "description": row.description,
            "mode": row.mode,
            "default_value": float(row.default_value) if row.default_value is not None else None,
            "currency": row.currency,
            "source_system": row.source_system,
            "is_active": row.is_active,
        }
        for row in rows
    ]


def _serialize_charge_rule_rows(rows: list[models.ClientChargeRule]) -> list[dict[str, Any]]:
    return [
        {
            "charge_rule_id": str(row.charge_rule_id),
            "rule_name": row.rule_name,
            "priority": row.priority,
            "country": row.country,
            "state": row.state,
            "postal_code": row.postal_code,
            "buyer_group": row.buyer_group,
            "buyer_email": row.buyer_email,
            "sku": row.sku,
            "category": row.category,
            "ship_to_code": row.ship_to_code,
            "sold_to_code": row.sold_to_code,
            "charge_code": row.charge_code,
            "override_mode": row.override_mode,
            "override_value": float(row.override_value) if row.override_value is not None else None,
            "source_system": row.source_system,
            "is_active": row.is_active,
        }
        for row in rows
    ]


def _serialize_buyer_term_rows(rows: list[models.ClientBuyerCommercialTerm]) -> list[dict[str, Any]]:
    return [
        {
            "buyer_term_id": str(row.buyer_term_id),
            "buyer_email": row.buyer_email,
            "buyer_name": row.buyer_name,
            "payment_terms": row.payment_terms,
            "discount_code": row.discount_code,
            "credit_rules": row.credit_rules,
            "tax_exemption_code": row.tax_exemption_code,
            "source_system": row.source_system,
            "is_active": row.is_active,
        }
        for row in rows
    ]


def _serialize_product_map_rows(rows: list[models.ClientProductCommercialMap]) -> list[dict[str, Any]]:
    return [
        {
            "product_commercial_map_id": str(row.product_commercial_map_id),
            "sku": row.sku,
            "default_tax_code": row.default_tax_code,
            "default_freight_code": row.default_freight_code,
            "default_shipping_code": row.default_shipping_code,
            "default_octroi_code": row.default_octroi_code,
            "default_discount_code": row.default_discount_code,
            "source_system": row.source_system,
            "is_active": row.is_active,
        }
        for row in rows
    ]


def _load_commercial_settings_payload(db: Session, client_id: str, environment: str | None = None) -> dict[str, Any]:
    env = _commercial_environment(environment)
    row = _commercial_setting_row(db, client_id, env)
    priorities = (
        db.query(models.ClientCommercialPriority)
        .filter(models.ClientCommercialPriority.client_id == client_id)
        .filter(models.ClientCommercialPriority.environment == env)
        .order_by(models.ClientCommercialPriority.sequence_no.asc(), models.ClientCommercialPriority.created_at.asc())
        .all()
    )
    charge_codes = (
        db.query(models.ClientChargeCode)
        .filter(models.ClientChargeCode.client_id == client_id)
        .filter(models.ClientChargeCode.environment == env)
        .order_by(models.ClientChargeCode.charge_type.asc(), models.ClientChargeCode.charge_code.asc())
        .all()
    )
    rules = (
        db.query(models.ClientChargeRule)
        .filter(models.ClientChargeRule.client_id == client_id)
        .filter(models.ClientChargeRule.environment == env)
        .order_by(models.ClientChargeRule.priority.asc(), models.ClientChargeRule.rule_name.asc())
        .all()
    )
    buyer_terms = (
        db.query(models.ClientBuyerCommercialTerm)
        .filter(models.ClientBuyerCommercialTerm.client_id == client_id)
        .filter(models.ClientBuyerCommercialTerm.environment == env)
        .order_by(models.ClientBuyerCommercialTerm.buyer_email.asc())
        .all()
    )
    product_mapping = (
        db.query(models.ClientProductCommercialMap)
        .filter(models.ClientProductCommercialMap.client_id == client_id)
        .filter(models.ClientProductCommercialMap.environment == env)
        .order_by(models.ClientProductCommercialMap.sku.asc())
        .all()
    )
    return {
        "client_id": client_id,
        "environment": env,
        "source_mode": str(getattr(row, "source_mode", "ORDANEX_MASTER") or "ORDANEX_MASTER").strip().upper(),
        "erp_sync_enabled": bool(getattr(row, "erp_sync_enabled", False)),
        "erp_sync_frequency": str(getattr(row, "erp_sync_frequency", "DAILY") or "DAILY").strip().upper(),
        "erp_last_sync_at": getattr(row, "erp_last_sync_at", None),
        "currency_mode": str(getattr(row, "currency_mode", "CLIENT_DEFAULT") or "CLIENT_DEFAULT").strip().upper(),
        "fallback_policy": str(getattr(row, "fallback_policy", "ZERO_FALLBACK") or "ZERO_FALLBACK").strip().upper(),
        "checkout_priority": _serialize_priority_rows(priorities),
        "charge_codes": _serialize_charge_code_rows(charge_codes),
        "jurisdiction_rules": _serialize_charge_rule_rows(rules),
        "buyer_terms": _serialize_buyer_term_rows(buyer_terms),
        "product_mapping": _serialize_product_map_rows(product_mapping),
    }


@router.get("/client-commercial-settings/{client_id}", response_model=schemas.ClientCommercialSettingsRead)
def get_client_commercial_settings(
    client_id: str,
    environment: str | None = Query(None),
    db: Session = Depends(get_db),
):
    return _load_commercial_settings_payload(db, client_id, environment)


@router.put("/client-commercial-settings/{client_id}", response_model=schemas.ClientCommercialSettingsRead)
def update_client_commercial_settings(
    client_id: str,
    payload: schemas.ClientCommercialSettingsUpdate,
    environment: str | None = Query(None),
    db: Session = Depends(get_db),
):
    env = _commercial_environment(environment)
    row = _ensure_commercial_setting_row(db, client_id, env)
    row.source_mode = str(payload.source_mode or "ORDANEX_MASTER").strip().upper()
    row.erp_sync_enabled = bool(payload.erp_sync_enabled)
    row.erp_sync_frequency = str(payload.erp_sync_frequency or "DAILY").strip().upper()
    row.erp_last_sync_at = payload.erp_last_sync_at
    row.currency_mode = str(payload.currency_mode or "CLIENT_DEFAULT").strip().upper()
    row.fallback_policy = str(payload.fallback_policy or "ZERO_FALLBACK").strip().upper()
    row.is_active = True
    db.add(row)

    db.query(models.ClientCommercialPriority).filter(models.ClientCommercialPriority.client_id == client_id).filter(models.ClientCommercialPriority.environment == env).delete(synchronize_session=False)
    db.query(models.ClientChargeCode).filter(models.ClientChargeCode.client_id == client_id).filter(models.ClientChargeCode.environment == env).delete(synchronize_session=False)
    db.query(models.ClientChargeRule).filter(models.ClientChargeRule.client_id == client_id).filter(models.ClientChargeRule.environment == env).delete(synchronize_session=False)
    db.query(models.ClientBuyerCommercialTerm).filter(models.ClientBuyerCommercialTerm.client_id == client_id).filter(models.ClientBuyerCommercialTerm.environment == env).delete(synchronize_session=False)
    db.query(models.ClientProductCommercialMap).filter(models.ClientProductCommercialMap.client_id == client_id).filter(models.ClientProductCommercialMap.environment == env).delete(synchronize_session=False)

    for entry in payload.checkout_priority:
        db.add(models.ClientCommercialPriority(
            client_id=client_id,
            environment=env,
            sequence_no=entry.sequence_no,
            priority_code=entry.priority_code,
            priority_label=entry.priority_label,
            source_system=entry.source_system,
            is_active=entry.is_active,
        ))
    for entry in payload.charge_codes:
        db.add(models.ClientChargeCode(
            client_id=client_id,
            environment=env,
            charge_code=entry.charge_code,
            charge_type=entry.charge_type,
            description=entry.description,
            mode=entry.mode,
            default_value=entry.default_value,
            currency=entry.currency,
            source_system=entry.source_system,
            is_active=entry.is_active,
        ))
    for entry in payload.jurisdiction_rules:
        db.add(models.ClientChargeRule(
            client_id=client_id,
            environment=env,
            rule_name=entry.rule_name,
            priority=entry.priority,
            country=entry.country,
            state=entry.state,
            postal_code=entry.postal_code,
            buyer_group=entry.buyer_group,
            buyer_email=entry.buyer_email,
            sku=entry.sku,
            category=entry.category,
            ship_to_code=entry.ship_to_code,
            sold_to_code=entry.sold_to_code,
            charge_code=entry.charge_code,
            override_mode=entry.override_mode,
            override_value=entry.override_value,
            source_system=entry.source_system,
            is_active=entry.is_active,
        ))
    for entry in payload.buyer_terms:
        db.add(models.ClientBuyerCommercialTerm(
            client_id=client_id,
            environment=env,
            buyer_email=str(entry.buyer_email or "").strip().lower(),
            buyer_name=entry.buyer_name,
            payment_terms=entry.payment_terms,
            discount_code=entry.discount_code,
            credit_rules=entry.credit_rules,
            tax_exemption_code=entry.tax_exemption_code,
            source_system=entry.source_system,
            is_active=entry.is_active,
        ))
    for entry in payload.product_mapping:
        db.add(models.ClientProductCommercialMap(
            client_id=client_id,
            environment=env,
            sku=entry.sku,
            default_tax_code=entry.default_tax_code,
            default_freight_code=entry.default_freight_code,
            default_shipping_code=entry.default_shipping_code,
            default_octroi_code=entry.default_octroi_code,
            default_discount_code=entry.default_discount_code,
            source_system=entry.source_system,
            is_active=entry.is_active,
        ))

    db.commit()
    return _load_commercial_settings_payload(db, client_id, env)


@router.post("/client-commercial-settings/{client_id}/publish")
def publish_client_commercial_settings(
    client_id: str,
    from_environment: str = Query("STAGING"),
    to_environment: str = Query("PROD"),
    db: Session = Depends(get_db),
):
    source_environment = _commercial_environment(from_environment)
    target_environment = _commercial_environment(to_environment)
    if source_environment == target_environment:
        raise HTTPException(status_code=400, detail="Source and target commercial environments must be different.")

    source_payload = _load_commercial_settings_payload(db, client_id, source_environment)
    if not source_payload:
        raise HTTPException(status_code=404, detail="No commercial settings found for the source environment.")

    update_client_commercial_settings(
        client_id,
        schemas.ClientCommercialSettingsUpdate(
            source_mode=source_payload["source_mode"],
            erp_sync_enabled=source_payload["erp_sync_enabled"],
            erp_sync_frequency=source_payload["erp_sync_frequency"],
            erp_last_sync_at=source_payload["erp_last_sync_at"],
            currency_mode=source_payload["currency_mode"],
            fallback_policy=source_payload["fallback_policy"],
            checkout_priority=source_payload["checkout_priority"],
            charge_codes=source_payload["charge_codes"],
            jurisdiction_rules=source_payload["jurisdiction_rules"],
            buyer_terms=source_payload["buyer_terms"],
            product_mapping=source_payload["product_mapping"],
        ),
        target_environment,
        db,
    )
    return {
        "client_id": client_id,
        "source_environment": source_environment,
        "target_environment": target_environment,
        "settings": _load_commercial_settings_payload(db, client_id, target_environment),
    }


def _normalize_catalog_items(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    if isinstance(value, dict):
        for key in ("items", "catalog", "catalog_items", "products", "entries"):
            items = value.get(key)
            if isinstance(items, list):
                return [item for item in items if isinstance(item, dict)]
    return []


def _catalog_source_payloads(db: Session, client_id: str) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []

    erp_rows = (
        db.query(models.ClientERPConfig)
        .filter(models.ClientERPConfig.client_id == client_id)
        .filter(models.ClientERPConfig.is_active.is_(True))
        .order_by(models.ClientERPConfig.updated_at.desc(), models.ClientERPConfig.created_at.desc())
        .all()
    )
    for row in erp_rows:
        config = row.config_json if isinstance(row.config_json, dict) else {}
        payloads.append(
            {
                "source_system": getattr(row, "erp_name", None) or getattr(row, "erp_system", None) or "ERP",
                "endpoint_url": config.get("endpoint_url") or config.get("webhook_url") or config.get("endpoint"),
                "items": _normalize_catalog_items(
                    config.get("catalog")
                    or config.get("catalog_items")
                    or config.get("items")
                    or config.get("product_catalog")
                    or config.get("products")
                ),
            }
        )

    connection_rows = (
        db.query(models.ClientConnection)
        .filter(models.ClientConnection.client_id == client_id)
        .filter(models.ClientConnection.is_active.is_(True))
        .order_by(models.ClientConnection.updated_at.desc(), models.ClientConnection.created_at.desc())
        .all()
    )
    for row in connection_rows:
        config = row.config_json if isinstance(row.config_json, dict) else {}
        payloads.append(
            {
                "source_system": getattr(row, "connection_name", None) or getattr(row, "message_type", None) or "CLIENT ERP",
                "endpoint_url": config.get("endpoint_url") or config.get("webhook_url") or config.get("endpoint"),
                "items": _normalize_catalog_items(
                    config.get("catalog")
                    or config.get("catalog_items")
                    or config.get("items")
                    or config.get("product_catalog")
                    or config.get("products")
                ),
            }
        )

    return payloads


@router.post("/buyer-storefront-catalog-sync/{client_id}")
def sync_buyer_storefront_catalog(
    client_id: str,
    environment: str | None = Query(None),
    db: Session = Depends(get_db),
):
    row = _buyer_storefront_settings_row(db, client_id, environment)
    if row is None:
        raise HTTPException(status_code=404, detail="Buyer storefront settings not found.")

    settings = row.config_value_json if isinstance(row.config_value_json, dict) else {}
    catalog = settings.get("catalog") if isinstance(settings.get("catalog"), dict) else {}
    source_mode = str(catalog.get("source_mode") or "ERP_SYNCED").strip().upper()
    if source_mode != "ERP_SYNCED":
        raise HTTPException(
            status_code=400,
            detail="Catalog sync is available for ERP-synced storefronts. Switch the storefront source mode first.",
        )

    payloads = _catalog_source_payloads(db, client_id)
    synced_items: list[dict[str, Any]] = []
    source_system = "ERP"
    endpoint_url = None
    for payload in payloads:
        if payload["items"]:
            synced_items = payload["items"]
            source_system = str(payload.get("source_system") or "ERP")
            endpoint_url = payload.get("endpoint_url")
            break

    if not synced_items:
        synced_items = []

    current = row.config_value_json if isinstance(row.config_value_json, dict) else {}
    current_catalog = current.get("catalog") if isinstance(current.get("catalog"), dict) else {}
    current_catalog = {
        **current_catalog,
        "items": synced_items,
        "last_synced_at": datetime.utcnow().isoformat(),
        "source_mode": "ERP_SYNCED",
        "source_label": "ERP-synced catalog",
        "sync_note": "Catalog refreshed from the ERP-side configuration sources.",
    }
    current["catalog"] = current_catalog
    row.config_value_json = current
    row.is_active = True
    db.commit()
    db.refresh(row)

    try:
        db.add(
            models.ClientSyncEvent(
                client_id=client_id,
                sync_key="CATALOG",
                event_type="CATALOG_SYNC",
                status="SUCCESS" if synced_items else "NO_SOURCE",
                message="Catalog refreshed from ERP-side configuration." if synced_items else "No ERP catalog source found; storefront catalog left empty.",
                endpoint_url=endpoint_url,
                source_system=source_system,
                target_system="BUYER_PORTAL",
                records_synced=len(synced_items),
                last_synced_at=datetime.utcnow(),
                details_json={"catalog_items": len(synced_items)},
            )
        )
        db.commit()
    except Exception:
        db.rollback()

    return {
        "client_id": client_id,
        "environment": _normalize_storefront_environment(environment),
        "records_synced": len(synced_items),
        "source_system": source_system,
        "settings": row.config_value_json if isinstance(row.config_value_json, dict) else {},
    }


@router.get("/verticals/{client_id}", response_model=list[schemas.VerticalRead])
def get_verticals(client_id: str, db: Session = Depends(get_db)):
    return (
        db.query(models.BusinessVertical)
        .filter(models.BusinessVertical.client_id == client_id)
        .order_by(models.BusinessVertical.vertical_name.asc())
        .all()
    )

@router.post("/verticals", response_model=schemas.VerticalRead)
def create_vertical(payload: schemas.VerticalCreate, db: Session = Depends(get_db)):
    row = models.BusinessVertical(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row

@router.put("/verticals/{vertical_id}", response_model=schemas.VerticalRead)
def update_vertical(vertical_id: str, payload: schemas.VerticalUpdate, db: Session = Depends(get_db)):
    row = db.query(models.BusinessVertical).filter(models.BusinessVertical.vertical_id == vertical_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Business vertical not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(row, field, value)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row

@router.get("/connections/{client_id}", response_model=list[schemas.ClientConnectionRead])
def get_connections(client_id: str, db: Session = Depends(get_db)):
    return (
        db.query(models.ClientConnection)
        .filter(models.ClientConnection.client_id == client_id)
        .order_by(models.ClientConnection.created_at.desc())
        .all()
    )

@router.post("/connections", response_model=schemas.ClientConnectionRead)
def create_connection(payload: schemas.ClientConnectionCreate, db: Session = Depends(get_db)):
    row = models.ClientConnection(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)

    sync_key = _sync_key_from_connection(row)
    if sync_key:
        _record_sync_event(
            db,
            client_id=row.client_id,
            sync_key=sync_key,
            event_type="CONFIG_UPDATED",
            status=_sync_status_for_key(db, row.client_id, sync_key),
            message=f"{sync_key} connection saved",
            endpoint_url=(row.config_json or {}).get("endpoint_url") or (row.config_json or {}).get("webhook_url"),
            source_system="CLIENT ERP",
            target_system=sync_key,
            details_json={"connection_id": str(row.connection_id), "connection_name": row.connection_name, "is_active": row.is_active},
        )
    return row

@router.put("/connections/{connection_id}", response_model=schemas.ClientConnectionRead)
def update_connection(connection_id: str, payload: schemas.ClientConnectionUpdate, db: Session = Depends(get_db)):
    row = db.query(models.ClientConnection).filter(models.ClientConnection.connection_id == connection_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Connection not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(row, field, value)
    db.add(row)
    db.commit()
    db.refresh(row)

    sync_key = _sync_key_from_connection(row)
    if sync_key:
        _record_sync_event(
            db,
            client_id=row.client_id,
            sync_key=sync_key,
            event_type="CONFIG_UPDATED",
            status=_sync_status_for_key(db, row.client_id, sync_key),
            message=f"{sync_key} connection updated",
            endpoint_url=(row.config_json or {}).get("endpoint_url") or (row.config_json or {}).get("webhook_url"),
            source_system="CLIENT ERP",
            target_system=sync_key,
            details_json={"connection_id": str(row.connection_id), "connection_name": row.connection_name, "is_active": row.is_active},
        )
    return row

@router.get("/erp/{client_id}", response_model=list[schemas.ClientERPConfigRead])
def get_erp_configs(client_id: str, db: Session = Depends(get_db)):
    return (
        db.query(models.ClientERPConfig)
        .filter(models.ClientERPConfig.client_id == client_id)
        .order_by(models.ClientERPConfig.created_at.desc())
        .all()
    )

@router.post("/erp", response_model=schemas.ClientERPConfigRead)
def create_erp_config(payload: schemas.ClientERPConfigCreate, db: Session = Depends(get_db)):
    row = models.ClientERPConfig(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)

    sync_key = _sync_key_from_erp(row)
    if sync_key:
        _record_sync_event(
            db,
            client_id=row.client_id,
            sync_key=sync_key,
            event_type="ERP_CONFIG_UPDATED",
            status=_sync_status_for_key(db, row.client_id, sync_key),
            message=f"{sync_key} ERP config saved",
            source_system=row.erp_name,
            target_system=sync_key,
            details_json={"erp_config_id": str(row.erp_config_id), "message_type": row.message_type, "is_active": row.is_active},
        )
    return row

@router.put("/erp/{erp_config_id}", response_model=schemas.ClientERPConfigRead)
def update_erp_config(erp_config_id: str, payload: schemas.ClientERPConfigUpdate, db: Session = Depends(get_db)):
    row = db.query(models.ClientERPConfig).filter(models.ClientERPConfig.erp_config_id == erp_config_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="ERP config not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(row, field, value)
    db.add(row)
    db.commit()
    db.refresh(row)

    sync_key = _sync_key_from_erp(row)
    if sync_key:
        _record_sync_event(
            db,
            client_id=row.client_id,
            sync_key=sync_key,
            event_type="ERP_CONFIG_UPDATED",
            status=_sync_status_for_key(db, row.client_id, sync_key),
            message=f"{sync_key} ERP config updated",
            source_system=row.erp_name,
            target_system=sync_key,
            details_json={"erp_config_id": str(row.erp_config_id), "message_type": row.message_type, "is_active": row.is_active},
        )
    return row



@router.post("/sync/{client_id}/{sync_key}")
def trigger_client_sync(
    client_id: str,
    sync_key: str,
    db: Session = Depends(get_db),
):
    normalized_key = sync_key.strip().upper()
    if normalized_key not in {"UOM", "ADDRESS"}:
        raise HTTPException(status_code=400, detail="Unsupported sync key")

    connection = (
        db.query(models.ClientConnection)
        .filter(models.ClientConnection.client_id == client_id)
        .filter(models.ClientConnection.is_active.is_(True))
        .filter(
            (models.ClientConnection.config_json["sync_object"].astext.ilike(normalized_key))
            | (models.ClientConnection.connection_name.ilike(f"%{normalized_key}%"))
        )
        .order_by(models.ClientConnection.created_at.desc())
        .first()
    )
    if not connection:
        _record_sync_event(
            db,
            client_id=client_id,
            sync_key=normalized_key,
            event_type="SYNC_TRIGGER",
            status="FAILED",
            message=f"No active {normalized_key} connection found",
        )
        raise HTTPException(status_code=404, detail=f"No active {normalized_key} connection found")

    config_json = connection.config_json or {}
    url = _build_sync_url(config_json, normalized_key)
    if not url:
        _record_sync_event(
            db,
            client_id=client_id,
            sync_key=normalized_key,
            event_type="SYNC_TRIGGER",
            status="FAILED",
            message=f"No endpoint configured for {normalized_key}",
            details_json={"connection_id": str(connection.connection_id)},
        )
        raise HTTPException(status_code=400, detail=f"No endpoint configured for {normalized_key}")

    method = str(config_json.get("http_method") or "POST").strip().upper()
    timeout_seconds = int(config_json.get("timeout_seconds") or 30)
    headers = _build_sync_headers(config_json)
    payload = {
        "client_id": client_id,
        "sync_key": normalized_key,
        "connection_id": str(connection.connection_id),
        "source_system": config_json.get("source_system") or config_json.get("erp_name") or "CLIENT ERP",
        "target_system": normalized_key,
    }
    data_bytes = json.dumps(payload).encode("utf-8") if method != "GET" else None
    if data_bytes is not None:
        headers.setdefault("Content-Type", "application/json")

    request = urllib.request.Request(url=url, data=data_bytes, headers=headers, method=method)
    started = time.perf_counter()
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            payload_bytes = response.read()
            content_type = response.headers.get("Content-Type") or ""
            elapsed_ms = int((time.perf_counter() - started) * 1000)
            records_synced = _parse_records_synced(payload_bytes, content_type)
            _record_sync_event(
                db,
                client_id=client_id,
                sync_key=normalized_key,
                event_type="SYNC_SUCCESS",
                status="SUCCESS",
                message=f"{normalized_key} sync completed successfully",
                endpoint_url=url,
                source_system=str(config_json.get("source_system") or config_json.get("erp_name") or "CLIENT ERP"),
                target_system=normalized_key,
                records_synced=records_synced,
                duration_ms=elapsed_ms,
                details_json={
                    "connection_id": str(connection.connection_id),
                    "http_method": method,
                    "response_status": getattr(response, "status", None),
                    "content_type": content_type,
                },
            )
            return {
                "success": True,
                "sync_key": normalized_key,
                "endpoint_url": url,
                "records_synced": records_synced,
                "duration_ms": elapsed_ms,
                "response_status": getattr(response, "status", None),
            }
    except urllib.error.HTTPError as exc:
        elapsed_ms = int((time.perf_counter() - started) * 1000)
        body = exc.read().decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
        _record_sync_event(
            db,
            client_id=client_id,
            sync_key=normalized_key,
            event_type="SYNC_FAILED",
            status="FAILED",
            message=f"{normalized_key} sync failed: HTTP {exc.code}",
            endpoint_url=url,
            source_system=str(config_json.get("source_system") or config_json.get("erp_name") or "CLIENT ERP"),
            target_system=normalized_key,
            duration_ms=elapsed_ms,
            details_json={
                "connection_id": str(connection.connection_id),
                "http_status": exc.code,
                "response_body": body[:2000],
            },
        )
        raise HTTPException(status_code=502, detail=f"{normalized_key} sync failed with HTTP {exc.code}")
    except Exception as exc:
        elapsed_ms = int((time.perf_counter() - started) * 1000)
        _record_sync_event(
            db,
            client_id=client_id,
            sync_key=normalized_key,
            event_type="SYNC_FAILED",
            status="FAILED",
            message=f"{normalized_key} sync failed: {exc}",
            endpoint_url=url,
            source_system=str(config_json.get("source_system") or config_json.get("erp_name") or "CLIENT ERP"),
            target_system=normalized_key,
            duration_ms=elapsed_ms,
            details_json={"connection_id": str(connection.connection_id), "error": str(exc)},
        )
        raise HTTPException(status_code=502, detail=f"{normalized_key} sync failed: {exc}")

@router.get("/sync-events/{client_id}", response_model=list[schemas.ClientSyncEventRead])
def get_sync_events(
    client_id: str,
    sync_key: str | None = None,
    limit: int = 20,
    db: Session = Depends(get_db),
):
    query = db.query(models.ClientSyncEvent).filter(models.ClientSyncEvent.client_id == client_id)
    if sync_key:
        query = query.filter(models.ClientSyncEvent.sync_key == sync_key.strip().upper())
    return query.order_by(models.ClientSyncEvent.created_at.desc()).limit(max(1, min(limit, 100))).all()


@router.post("/sync-events", response_model=schemas.ClientSyncEventRead)
def create_sync_event(payload: schemas.ClientSyncEventCreate, db: Session = Depends(get_db)):
    row = models.ClientSyncEvent(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/storefront/catalog-template")
def download_storefront_catalog_template():
    return Response(
        content=_build_storefront_catalog_template(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="ordanex-storefront-catalog-template.xlsx"'
        },
    )
