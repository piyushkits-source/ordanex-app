from __future__ import annotations

import io
from uuid import UUID
from pydantic import BaseModel

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from backend.db.database import get_db
from backend.db import models, schemas
from backend.services.address_master_excel_validation import validate_address_master_workbook
from backend.services.address_matching_engine import rank_address_candidates

router = APIRouter(prefix="/address-master", tags=["Address Master"])


@router.get("", response_model=list[schemas.AddressMasterRead])
def get_address_master(partner_id: UUID, db: Session = Depends(get_db)):
    return (
        db.query(models.AddressMaster)
        .filter(models.AddressMaster.partner_id == partner_id)
        .order_by(models.AddressMaster.created_at.desc())
        .all()
    )


@router.post("", response_model=schemas.AddressMasterRead)
def create_address_master(payload: schemas.AddressMasterCreate, db: Session = Depends(get_db)):
    row = models.AddressMaster(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/{address_id}")
def delete_address_master(address_id: UUID, db: Session = Depends(get_db)):
    row = db.query(models.AddressMaster).filter(models.AddressMaster.address_id == address_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Address record not found.")
    db.delete(row)
    db.commit()
    return {"status": "deleted"}


@router.get("/template/{partner_id}")
def download_address_master_template(partner_id: UUID):
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info["A1"] = "Address Master Template Instructions"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "1. Do not change header names"
    ws_info["A4"] = "2. Fill only data rows"
    ws_info["A5"] = "3. Upload back in same format"
    ws_info["A6"] = "4. Use dropdown values where available"

    ws = wb.create_sheet(title="Address_Master")

    headers = [
        "client_id",
        "partner_id",
        "direction",
        "partner_type",
        "role_code",
        "address_name",
        "address_line1",
        "address_line2",
        "city",
        "state",
        "postal_code",
        "country",
        "ship_to_code",
        "sold_to_code",
        "bill_to_code",
        "supplier_code",
        "warehouse_code",
        "delivery_location_code",
        "is_active",
        "notes",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        cell.protection = Protection(locked=True)

    for row in ws.iter_rows(min_row=2, max_row=500):
        for cell in row:
            cell.protection = Protection(locked=False)

    dv_direction = DataValidation(type="list", formula1='"INBOUND,OUTBOUND"', allow_blank=True)
    ws.add_data_validation(dv_direction)
    dv_direction.add("C2:C500")

    dv_partner_type = DataValidation(
        type="list",
        formula1='"CUSTOMER,SUPPLIER,LOGISTICS_PROVIDER,WAREHOUSE"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_partner_type)
    dv_partner_type.add("D2:D500")

    dv_role = DataValidation(
        type="list",
        formula1='"SHIP_TO,SOLD_TO,BILL_TO,SUPPLIER,WAREHOUSE,DELIVERY_LOCATION"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_role)
    dv_role.add("E2:E500")

    dv_active = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv_active)
    dv_active.add("S2:S500")

    ws.protection.sheet = True

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=address_master_template.xlsx"},
    )


@router.post("/upload/{partner_id}")
async def upload_address_master(
    partner_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    contents = await file.read()
    result = validate_address_master_workbook(contents)

    if not result.is_valid:
        return StreamingResponse(
            io.BytesIO(result.workbook_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=address_master_validation_errors.xlsx",
                "X-Upload-Status": "validation_failed",
            },
        )

    for row in result.parsed_rows:
        db_row = models.AddressMaster(
            client_id=row["client_id"],
            partner_id=row["partner_id"],
            direction=row["direction"],
            partner_type=row["partner_type"],
            role_code=row["role_code"],
            address_name=row["address_name"],
            address_line1=row["address_line1"],
            address_line2=row["address_line2"],
            city=row["city"],
            state=row["state"],
            postal_code=row["postal_code"],
            country=row["country"],
            ship_to_code=row["ship_to_code"],
            sold_to_code=row["sold_to_code"],
            bill_to_code=row["bill_to_code"],
            supplier_code=row["supplier_code"],
            warehouse_code=row["warehouse_code"],
            delivery_location_code=row["delivery_location_code"],
            is_active=row["is_active"],
            notes=row["notes"],
        )
        db.add(db_row)

    db.commit()

    return {
        "status": "success",
        "rows_processed": len(result.parsed_rows),
    }

class AddressMatchPreviewRequest(BaseModel):
    partner_id: UUID
    source_address_text: str
    direction: str | None = None
    partner_type: str | None = None
    role_code: str | None = None
    top_n: int = 5

@router.post("/match-preview")
def preview_address_match(payload: AddressMatchPreviewRequest, db: Session = Depends(get_db)):
    query = db.query(models.AddressMaster).filter(
        models.AddressMaster.partner_id == payload.partner_id,
        models.AddressMaster.is_active == True,  # noqa: E712
    )

    if payload.direction:
        query = query.filter(models.AddressMaster.direction == payload.direction)

    if payload.partner_type:
        query = query.filter(models.AddressMaster.partner_type == payload.partner_type)

    if payload.role_code:
        query = query.filter(models.AddressMaster.role_code == payload.role_code)

    rows = query.all()
    ranked = rank_address_candidates(
        source_text=payload.source_address_text,
        rows=rows,
        limit=payload.top_n,
    )

    return {
        "status": "success",
        "input_text": payload.source_address_text,
        "candidates": [
            {
                "address_id": item.address_id,
                "score": item.score,
                "reason": item.reason,
                "payload": item.payload,
            }
            for item in ranked
        ],
        "best_match": (
            {
                "address_id": ranked[0].address_id,
                "score": ranked[0].score,
                "reason": ranked[0].reason,
                "payload": ranked[0].payload,
            }
            if ranked
            else None
        ),
    }
