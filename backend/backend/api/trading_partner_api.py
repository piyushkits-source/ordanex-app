from uuid import UUID
from pathlib import Path
import io
import json
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from backend.services.excel_validation_engine import (
    validate_bulk_onboarding_workbook,
    validate_uom_workbook,
)
from backend.services.email_polling_service import email_polling_service

from backend.db.database import get_db
from backend.db import models, schemas
from backend.db import schemas_partner_patch
from backend.db.models_rules_uom_mapping import (
    TradingPartnerBusinessRule,
    TradingPartnerMappingProfile,
    TradingPartnerOnboardingAudit,
    TradingPartnerUomRule,
)
from backend.db.models_partner_patch import PartnerNotification
from backend.core.environment import current_environment, is_production, is_staging
from backend.services.onboarding_config_service import write_audit

router = APIRouter(prefix="/trading-partners", tags=["Trading Partners"])

BASE_DIR = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = BASE_DIR / "templates"


def _normalize_connection_payload(
    payload: schemas.TradingPartnerConnectionCreate | schemas.TradingPartnerConnectionUpdate,
    existing_row: models.TradingPartnerConnection | None = None,
) -> dict:
    data = payload.model_dump(exclude_unset=True)
    connection_type = str(
        data.get("connection_type")
        or getattr(existing_row, "connection_type", "")
        or ""
    ).upper()

    if connection_type != "EMAIL":
        return data

    merged_config = {
        **dict(getattr(existing_row, "config_json", {}) or {}),
        **dict(data.get("config_json") or {}),
    }
    normalized = email_polling_service.normalize_email_config(
        client_id=str(data.get("client_id") or getattr(existing_row, "client_id", "")),
        connection_key=data.get("connection_id") or getattr(existing_row, "connection_id", None),
        config=merged_config,
    )

    data["connection_type"] = "EMAIL"
    data["config_json"] = {
        "email_address": normalized.get("email_address") or normalized.get("username") or merged_config.get("email_address") or merged_config.get("email") or merged_config.get("username"),
        "username": normalized.get("username") or normalized.get("email_address") or merged_config.get("username") or merged_config.get("email"),
        "password_token": normalized.get("password_token") or normalized.get("password") or merged_config.get("password_token") or merged_config.get("password") or merged_config.get("app_password"),
        "imap_host": normalized.get("imap_host") or merged_config.get("imap_host") or merged_config.get("host") or merged_config.get("server"),
        "port": normalized.get("imap_port") or merged_config.get("port") or 993,
        "folder": normalized.get("folder") or normalized.get("mailbox") or merged_config.get("folder") or merged_config.get("mailbox") or "INBOX",
        "subject_filter": normalized.get("subject_filter") or normalized.get("subject_contains") or merged_config.get("subject_filter") or "",
        "allowed_senders": normalized.get("allowed_senders") or merged_config.get("allowed_senders") or [],
    }
    return data


def _validate_connection_payload(data: dict) -> None:
    if str(data.get("connection_type") or "").upper() != "EMAIL":
        return

    if not data.get("is_active", True):
        return

    cfg = dict(data.get("config_json") or {})
    missing = []
    for key in ["email_address", "imap_host", "username", "password_token"]:
        if not str(cfg.get(key) or "").strip():
            missing.append(key)

    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required EMAIL settings: {', '.join(missing)}",
        )


def _model_to_dict(model_obj):
    return {c.name: getattr(model_obj, c.name) for c in model_obj.__table__.columns}


def _ensure_default_partner_profile(db: Session, partner: models.TradingPartner):
    profile = (
        db.query(models.TradingPartnerProfile)
        .filter(models.TradingPartnerProfile.partner_id == partner.partner_id)
        .first()
    )
    if profile:
        return profile

    profile = models.TradingPartnerProfile(
        client_id=partner.client_id,
        partner_id=partner.partner_id,
        profile_name="Default Profile",
        profile_status="ACTIVE",
        duplicate_check_enabled=True,
        duplicate_check_scope="PO_NUMBER",
        split_rule="NONE",
        split_po_number_strategy="SAME_PO_NUMBER",
        split_po_separator="-",
        delivery_date_source="PO_DELIVERY_DATE",
        delivery_date_offset_type="NONE",
        delivery_date_offset_days=0,
        po_date_source="PO_DATE",
        split_quantity_basis="ORDER_QTY",
        split_rounding_mode="UP",
    )
    db.add(profile)
    db.flush()
    return profile


def _upsert_bulk_mapping_validation(
    db: Session,
    *,
    partner: models.TradingPartner,
    validation_json: dict | None,
    target_defaults: dict | None = None,
    layout_hints: dict | None = None,
):
    if not validation_json and not target_defaults and not layout_hints:
        return None, None

    existing = (
        db.query(TradingPartnerMappingProfile)
        .filter(TradingPartnerMappingProfile.partner_id == partner.partner_id)
        .filter(TradingPartnerMappingProfile.profile_name == "Bulk Validation Profile")
        .filter(TradingPartnerMappingProfile.document_type == "PO")
        .first()
    )
    before = _model_to_dict(existing) if existing else None

    if existing:
        existing.validation_json = validation_json
        existing.header_defaults_json = {**dict(existing.header_defaults_json or {}), **dict(target_defaults or {})}
        existing.layout_hint_json = {**dict(existing.layout_hint_json or {}), **dict(layout_hints or {})}
        existing.input_format = "UNKNOWN"
        existing.is_active = True
        existing.is_default = True
        existing.priority = 10
        existing.notes = "Bulk onboarding validation profile"
        db.add(existing)
        db.flush()
        return existing, before

    created = TradingPartnerMappingProfile(
        client_id=partner.client_id,
        partner_id=partner.partner_id,
        profile_name="Bulk Validation Profile",
        document_type="PO",
        input_format="UNKNOWN",
        field_mapping_json={},
        header_defaults_json=dict(target_defaults or {}),
        line_mapping_json={},
        validation_json=validation_json,
        layout_hint_json=dict(layout_hints or {}),
        version_no=1,
        priority=10,
        is_default=True,
        is_active=True,
        notes="Bulk onboarding validation profile",
    )
    db.add(created)
    db.flush()
    return created, None


def _upsert_bulk_business_validation_rule(
    db: Session,
    *,
    partner: models.TradingPartner,
    action_json: dict | None,
):
    if not action_json:
        return None, None

    existing = (
        db.query(TradingPartnerBusinessRule)
        .filter(TradingPartnerBusinessRule.partner_id == partner.partner_id)
        .filter(TradingPartnerBusinessRule.rule_name == "Bulk Validation Rule")
        .filter(TradingPartnerBusinessRule.rule_type == "VALIDATION")
        .filter(TradingPartnerBusinessRule.document_type == "PO")
        .filter(TradingPartnerBusinessRule.message_direction == "INBOUND")
        .first()
    )
    before = _model_to_dict(existing) if existing else None

    if existing:
        existing.condition_json = {}
        existing.action_json = action_json
        existing.priority = 10
        existing.stop_on_match = False
        existing.is_active = True
        existing.notes = "Bulk onboarding validation rule"
        db.add(existing)
        db.flush()
        return existing, before

    created = TradingPartnerBusinessRule(
        client_id=partner.client_id,
        partner_id=partner.partner_id,
        rule_name="Bulk Validation Rule",
        rule_type="VALIDATION",
        document_type="PO",
        message_direction="INBOUND",
        condition_json={},
        action_json=action_json,
        priority=10,
        stop_on_match=False,
        is_active=True,
        notes="Bulk onboarding validation rule",
    )
    db.add(created)
    db.flush()
    return created, None


PROMOTION_PACKAGE_VERSION = "3.0"
PROMOTION_TARGET_ENVIRONMENT = "production"
PROMOTION_DOMAIN_NAMES = [
    "client",
    "business_verticals",
    "client_connections",
    "client_erp_configs",
    "client_configs",
    "client_email_config",
    "partner",
    "profile",
    "connections",
    "mapping_profiles",
    "business_rules",
    "uom_rules",
    "address_master",
    "message_flows",
    "parser_profiles",
    "notifications",
]
PROMOTION_SYSTEM_FIELDS = {"created_at", "updated_at"}
FIELD_REQUIREMENT_TEMPLATE_COLUMNS = [
    "req_document_number",
    "req_document_date",
    "req_document_type",
    "req_order_type",
    "req_customer_name",
    "req_supplier_name",
    "req_currency_code",
    "req_ship_to_code",
    "req_ship_to_name",
    "req_bill_to_code",
    "req_bill_to_name",
    "req_transaction_id",
    "req_header_text_id",
    "req_line_text_id",
    "req_invoice_number",
    "req_invoice_date",
    "req_invoice_due_date",
    "req_invoice_total",
    "req_item_material_code",
    "req_item_description",
    "req_item_quantity",
    "req_item_customer_uom",
    "req_item_supplier_uom",
    "req_item_delivery_date",
    "req_item_unit_price",
    "req_item_amount",
    "req_item_plant_code",
    "req_item_ship_to_override",
]


def _clone_model_row(
    db: Session,
    model_cls,
    source_row,
    *,
    overrides: dict | None = None,
):
    data = {}
    primary_keys = {column.name for column in model_cls.__table__.primary_key.columns}
    for column in model_cls.__table__.columns:
        if column.name in primary_keys or column.name in PROMOTION_SYSTEM_FIELDS:
            continue
        data[column.name] = getattr(source_row, column.name)
    data.update(overrides or {})
    cloned = model_cls(**data)
    db.add(cloned)
    db.flush()
    return cloned

TARGET_PROFILE_TEMPLATE_COLUMNS = [
    "target_erp",
    "target_standard",
    "target_message_type",
    "target_message_version",
    "transaction_id_source",
    "invoice_profile_type",
    "invoice_number_source",
    "invoice_date_source",
    "invoice_total_source",
    "customization_required",
    "customization_notes",
]

TARGET_DEFAULT_TEMPLATE_COLUMNS = [
    "header_text_id",
    "line_text_id",
]


def _promotion_counts(
    *,
    client,
    business_verticals,
    client_connections,
    client_erp_configs,
    client_configs,
    client_email_config,
    partner,
    profile,
    connections,
    mapping_profiles,
    business_rules,
    uom_rules,
    address_master,
    message_flows,
    parser_profiles,
    notifications,
) -> dict:
    return {
        "client": 1 if client else 0,
        "business_verticals": len(business_verticals),
        "client_connections": len(client_connections),
        "client_erp_configs": len(client_erp_configs),
        "client_configs": len(client_configs),
        "client_email_config": 1 if client_email_config else 0,
        "partner": 1 if partner else 0,
        "profile": 1 if profile else 0,
        "connections": len(connections),
        "mapping_profiles": len(mapping_profiles),
        "business_rules": len(business_rules),
        "uom_rules": len(uom_rules),
        "address_master": len(address_master),
        "message_flows": len(message_flows),
        "parser_profiles": len(parser_profiles),
        "notifications": len(notifications),
    }


def _promotion_package(
    client,
    business_verticals,
    client_connections,
    client_erp_configs,
    client_configs,
    client_email_config,
    partner,
    profile,
    connections,
    mapping_profiles,
    business_rules,
    uom_rules,
    address_master,
    message_flows,
    parser_profiles,
    notifications,
) -> dict:
    counts = _promotion_counts(
        client=client,
        business_verticals=business_verticals,
        client_connections=client_connections,
        client_erp_configs=client_erp_configs,
        client_configs=client_configs,
        client_email_config=client_email_config,
        partner=partner,
        profile=profile,
        connections=connections,
        mapping_profiles=mapping_profiles,
        business_rules=business_rules,
        uom_rules=uom_rules,
        address_master=address_master,
        message_flows=message_flows,
        parser_profiles=parser_profiles,
        notifications=notifications,
    )
    return {
        "meta": {
            "package_version": PROMOTION_PACKAGE_VERSION,
            "package_kind": "TRADING_PARTNER_PROMOTION",
            "promotion_mode": "PACKAGE_EXPORT_IMPORT",
            "source_environment": current_environment(),
            "target_environment": PROMOTION_TARGET_ENVIRONMENT,
            "exported_at": datetime.utcnow().isoformat(),
            "client_id": getattr(client, "client_id", getattr(partner, "client_id", None)),
            "partner_id": str(getattr(partner, "partner_id", "") or "") or None,
            "partner_code": getattr(partner, "partner_code", None),
            "partner_name": getattr(partner, "partner_name", None),
            "config_domains": list(PROMOTION_DOMAIN_NAMES),
            "domain_counts": counts,
        },
        "client": _model_to_dict(client) if client else None,
        "business_verticals": [_model_to_dict(row) for row in business_verticals],
        "client_connections": [_model_to_dict(row) for row in client_connections],
        "client_erp_configs": [_model_to_dict(row) for row in client_erp_configs],
        "client_configs": [_model_to_dict(row) for row in client_configs],
        "client_email_config": _model_to_dict(client_email_config) if client_email_config else None,
        "partner": _model_to_dict(partner) if partner else None,
        "profile": _model_to_dict(profile) if profile else None,
        "connections": [_model_to_dict(row) for row in connections],
        "mapping_profiles": [_model_to_dict(row) for row in mapping_profiles],
        "business_rules": [_model_to_dict(row) for row in business_rules],
        "uom_rules": [_model_to_dict(row) for row in uom_rules],
        "address_master": [_model_to_dict(row) for row in address_master],
        "message_flows": [_model_to_dict(row) for row in message_flows],
        "parser_profiles": [_model_to_dict(row) for row in parser_profiles],
        "notifications": [_model_to_dict(row) for row in notifications],
    }


def _promotion_importable_payload(model_cls, payload: dict | None, *, client_id: str, partner_id=None, preserve_primary_keys: bool = True):
    if not payload:
        return None
    allowed = {column.name for column in model_cls.__table__.columns}
    primary_keys = {column.name for column in model_cls.__table__.primary_key.columns}
    blocked = set(PROMOTION_SYSTEM_FIELDS)
    if not preserve_primary_keys:
        blocked |= primary_keys
    data = {key: value for key, value in dict(payload).items() if key in allowed and key not in blocked}
    if "client_id" in allowed:
        data["client_id"] = client_id
    if partner_id is not None and "partner_id" in allowed:
        data["partner_id"] = partner_id
    return data


def _replace_scoped_domain_rows(
    db: Session,
    model_cls,
    *,
    scope_field: str,
    scope_value,
    rows: list[dict] | None,
    client_id: str,
    partner_id=None,
):
    if not hasattr(model_cls, scope_field):
        return []
    db.query(model_cls).filter(getattr(model_cls, scope_field) == scope_value).delete(synchronize_session=False)
    db.flush()
    created_rows = []
    for row_payload in list(rows or []):
        data = _promotion_importable_payload(
            model_cls,
            row_payload,
            client_id=client_id,
            partner_id=partner_id,
        )
        if data is None:
            continue
        row = model_cls(**data)
        db.add(row)
        created_rows.append(row)
    db.flush()
    return created_rows


def _replace_client_domain_rows(db: Session, model_cls, *, client, rows: list[dict] | None):
    return _replace_scoped_domain_rows(
        db,
        model_cls,
        scope_field="client_id",
        scope_value=client.client_id,
        rows=rows,
        client_id=client.client_id,
    )


def _replace_partner_domain_rows(db: Session, model_cls, *, partner, rows: list[dict] | None):
    return _replace_scoped_domain_rows(
        db,
        model_cls,
        scope_field="partner_id",
        scope_value=partner.partner_id,
        rows=rows,
        client_id=partner.client_id,
        partner_id=partner.partner_id,
    )


@router.get("", response_model=list[schemas.TradingPartnerRead])
def get_trading_partners(
    client_id: str,
    vertical_id: UUID | None = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.TradingPartner).filter(models.TradingPartner.client_id == client_id)
    if vertical_id:
        query = query.filter(models.TradingPartner.vertical_id == vertical_id)
    return query.order_by(models.TradingPartner.partner_name.asc()).all()


@router.post("", response_model=schemas.TradingPartnerRead)
def create_trading_partner(payload: schemas.TradingPartnerCreate, db: Session = Depends(get_db)):
    existing = (
        db.query(models.TradingPartner)
        .filter(
            models.TradingPartner.client_id == payload.client_id,
            models.TradingPartner.partner_code == payload.partner_code,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Trading partner already exists.")

    row = models.TradingPartner(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/{partner_id}", response_model=schemas.TradingPartnerRead)
def update_trading_partner(partner_id: UUID, payload: schemas.TradingPartnerUpdate, db: Session = Depends(get_db)):
    row = db.query(models.TradingPartner).filter(models.TradingPartner.partner_id == partner_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(row, field, value)

    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.post("/{partner_id}/transfer", response_model=schemas.TradingPartnerTransferResponse)
def transfer_trading_partner_setup(
    partner_id: UUID,
    payload: schemas.TradingPartnerTransferRequest,
    db: Session = Depends(get_db),
):
    source_partner = (
        db.query(models.TradingPartner)
        .filter(models.TradingPartner.partner_id == partner_id)
        .first()
    )
    if not source_partner:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    target_client = (
        db.query(models.Client)
        .filter(models.Client.client_id == payload.target_client_id)
        .first()
    )
    if not target_client:
        raise HTTPException(status_code=404, detail="Target client not found.")

    target_partner_code = str(payload.target_partner_code or source_partner.partner_code or "").strip()
    target_partner_name = str(payload.target_partner_name or source_partner.partner_name or "").strip()
    if not target_partner_code:
        raise HTTPException(status_code=400, detail="Target partner code is required.")
    if not target_partner_name:
        raise HTTPException(status_code=400, detail="Target partner name is required.")

    if payload.target_vertical_id is not None:
        target_vertical = (
            db.query(models.BusinessVertical)
            .filter(models.BusinessVertical.vertical_id == payload.target_vertical_id)
            .first()
        )
        if not target_vertical or target_vertical.client_id != payload.target_client_id:
            raise HTTPException(
                status_code=400,
                detail="Target business vertical must belong to the selected target client.",
            )

    effective_vertical_id = (
        payload.target_vertical_id
        if payload.target_vertical_id is not None
        else source_partner.vertical_id
        if payload.target_client_id == source_partner.client_id
        else None
    )

    existing = (
        db.query(models.TradingPartner)
        .filter(
            models.TradingPartner.client_id == payload.target_client_id,
            models.TradingPartner.partner_code == target_partner_code,
        )
        .first()
    )
    if existing:
        raise HTTPException(
            status_code=400,
            detail="A trading partner with this code already exists for the target client.",
        )

    cloned_partner = _clone_model_row(
        db,
        models.TradingPartner,
        source_partner,
        overrides={
            "client_id": payload.target_client_id,
            "vertical_id": effective_vertical_id,
            "partner_code": target_partner_code,
            "partner_name": target_partner_name,
            "notes": (
                f"{(source_partner.notes or '').strip()} | Cloned from {source_partner.client_id}/{source_partner.partner_code}"
            ).strip(" |")
            or None,
        },
    )

    copied_items: dict[str, int] = {}
    connection_id_map: dict[UUID, UUID] = {}
    mapping_profile_id_map: dict[UUID, UUID] = {}
    rule_profile_id_map: dict[UUID, UUID] = {}
    uom_rule_id_map: dict[UUID, UUID] = {}
    address_id_map: dict[UUID, UUID] = {}
    parser_profile_id_map: dict[UUID, UUID] = {}

    profiles = (
        db.query(models.TradingPartnerProfile)
        .filter(models.TradingPartnerProfile.partner_id == source_partner.partner_id)
        .all()
    )
    for row in profiles:
        _clone_model_row(
            db,
            models.TradingPartnerProfile,
            row,
            overrides={"client_id": payload.target_client_id, "partner_id": cloned_partner.partner_id},
        )
    copied_items["profiles"] = len(profiles)

    connections = (
        db.query(models.TradingPartnerConnection)
        .filter(models.TradingPartnerConnection.partner_id == source_partner.partner_id)
        .all()
    )
    for row in connections:
        cloned = _clone_model_row(
            db,
            models.TradingPartnerConnection,
            row,
            overrides={"client_id": payload.target_client_id, "partner_id": cloned_partner.partner_id},
        )
        connection_id_map[row.connection_id] = cloned.connection_id
    copied_items["connections"] = len(connections)

    mapping_profiles = (
        db.query(TradingPartnerMappingProfile)
        .filter(TradingPartnerMappingProfile.partner_id == source_partner.partner_id)
        .all()
    )
    for row in mapping_profiles:
        cloned = _clone_model_row(
            db,
            TradingPartnerMappingProfile,
            row,
            overrides={"client_id": payload.target_client_id, "partner_id": cloned_partner.partner_id},
        )
        mapping_profile_id_map[row.mapping_profile_id] = cloned.mapping_profile_id
    copied_items["mapping_profiles"] = len(mapping_profiles)

    business_rules = (
        db.query(TradingPartnerBusinessRule)
        .filter(TradingPartnerBusinessRule.partner_id == source_partner.partner_id)
        .all()
    )
    for row in business_rules:
        cloned = _clone_model_row(
            db,
            TradingPartnerBusinessRule,
            row,
            overrides={"client_id": payload.target_client_id, "partner_id": cloned_partner.partner_id},
        )
        rule_profile_id_map[row.rule_id] = cloned.rule_id
    copied_items["business_rules"] = len(business_rules)

    uom_rules = (
        db.query(TradingPartnerUomRule)
        .filter(TradingPartnerUomRule.partner_id == source_partner.partner_id)
        .all()
    )
    for row in uom_rules:
        cloned = _clone_model_row(
            db,
            TradingPartnerUomRule,
            row,
            overrides={"client_id": payload.target_client_id, "partner_id": cloned_partner.partner_id},
        )
        uom_rule_id_map[row.uom_rule_id] = cloned.uom_rule_id
    copied_items["uom_rules"] = len(uom_rules)

    address_rows = (
        db.query(models.AddressMaster)
        .filter(models.AddressMaster.partner_id == source_partner.partner_id)
        .all()
    )
    for row in address_rows:
        cloned = _clone_model_row(
            db,
            models.AddressMaster,
            row,
            overrides={"client_id": payload.target_client_id, "partner_id": cloned_partner.partner_id},
        )
        address_id_map[row.address_id] = cloned.address_id
    copied_items["address_master"] = len(address_rows)

    parser_profiles = (
        db.query(models.ParserProfile)
        .filter(models.ParserProfile.partner_id == source_partner.partner_id)
        .all()
    )
    for row in parser_profiles:
        cloned = _clone_model_row(
            db,
            models.ParserProfile,
            row,
            overrides={"client_id": payload.target_client_id, "partner_id": cloned_partner.partner_id},
        )
        parser_profile_id_map[row.parser_profile_id] = cloned.parser_profile_id
    copied_items["parser_profiles"] = len(parser_profiles)

    message_flows = (
        db.query(models.MessageFlow)
        .filter(models.MessageFlow.partner_id == source_partner.partner_id)
        .all()
    )
    for row in message_flows:
        _clone_model_row(
            db,
            models.MessageFlow,
            row,
            overrides={
                "client_id": payload.target_client_id,
                "partner_id": cloned_partner.partner_id,
                "vertical_id": effective_vertical_id,
                "target_connection_id": connection_id_map.get(row.target_connection_id) if row.target_connection_id else None,
                "mapping_profile_id": mapping_profile_id_map.get(row.mapping_profile_id) if row.mapping_profile_id else None,
                "rule_profile_id": rule_profile_id_map.get(row.rule_profile_id) if getattr(row, "rule_profile_id", None) else None,
                "uom_profile_id": uom_rule_id_map.get(row.uom_profile_id) if row.uom_profile_id else None,
                "address_profile_id": address_id_map.get(row.address_profile_id) if row.address_profile_id else None,
                "parser_profile_id": parser_profile_id_map.get(row.parser_profile_id) if row.parser_profile_id else None,
                "validation_profile_id": mapping_profile_id_map.get(row.validation_profile_id) if row.validation_profile_id else None,
            },
        )
    copied_items["message_flows"] = len(message_flows)

    notifications_count = 0
    if payload.copy_notifications:
        notifications = (
            db.query(PartnerNotification)
            .filter(PartnerNotification.partner_id == source_partner.partner_id)
            .all()
        )
        for row in notifications:
            _clone_model_row(
                db,
                PartnerNotification,
                row,
                overrides={"partner_id": cloned_partner.partner_id},
            )
        notifications_count = len(notifications)
    copied_items["notifications"] = notifications_count

    audit_row = TradingPartnerOnboardingAudit(
        client_id=payload.target_client_id,
        partner_id=cloned_partner.partner_id,
        entity_type="PARTNER_TRANSFER",
        entity_id=str(cloned_partner.partner_id),
        action="CREATE",
        before_json={"source_partner_id": str(source_partner.partner_id), "source_client_id": source_partner.client_id},
        after_json={"target_client_id": payload.target_client_id, "target_vertical_id": str(effective_vertical_id) if effective_vertical_id else None},
        remarks="Partner setup transferred for M&A/divestiture or client ownership change.",
    )
    db.add(audit_row)

    db.commit()
    db.refresh(cloned_partner)
    return {"partner": cloned_partner, "copied_items": copied_items}


@router.get("/{partner_id}/profile", response_model=schemas.TradingPartnerProfileRead)
def get_partner_profile(partner_id: UUID, db: Session = Depends(get_db)):
    partner = db.query(models.TradingPartner).filter(models.TradingPartner.partner_id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    row = db.query(models.TradingPartnerProfile).filter(models.TradingPartnerProfile.partner_id == partner_id).first()
    if row:
        return row

    default_profile = models.TradingPartnerProfile(
        client_id=partner.client_id,
        partner_id=partner.partner_id,
        profile_name="Default Profile",
        profile_status="ACTIVE",
        duplicate_check_enabled=True,
        duplicate_check_scope="PO_NUMBER",
        split_rule="NONE",
        split_po_number_strategy="SAME_PO_NUMBER",
        split_po_separator="-",
        delivery_date_source="PO_DELIVERY_DATE",
        delivery_date_offset_type="NONE",
        delivery_date_offset_days=0,
        po_date_source="PO_DATE",
        split_quantity_basis="ORDER_QTY",
        split_rounding_mode="UP",
    )
    db.add(default_profile)
    db.commit()
    db.refresh(default_profile)
    return default_profile


@router.post("/{partner_id}/profile", response_model=schemas.TradingPartnerProfileRead)
def save_partner_profile(partner_id: UUID, payload: schemas.TradingPartnerProfileCreate, db: Session = Depends(get_db)):
    partner = db.query(models.TradingPartner).filter(models.TradingPartner.partner_id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    existing = db.query(models.TradingPartnerProfile).filter(models.TradingPartnerProfile.partner_id == partner_id).first()
    if existing:
        for field, value in payload.model_dump().items():
            setattr(existing, field, value)
        db.commit()
        db.refresh(existing)
        return existing

    row = models.TradingPartnerProfile(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _save_partner_notification_row(db: Session, partner_id: UUID, payload: schemas_partner_patch.PartnerNotificationCreate):
    partner = db.query(models.TradingPartner).filter(models.TradingPartner.partner_id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    row = PartnerNotification(
        partner_id=partner.partner_id,
        email=str(payload.email or "").strip(),
        notification_type=str(payload.notification_type or "FAILED").strip().upper() or "FAILED",
        include_attachment=bool(payload.include_attachment),
        is_active=bool(payload.is_active),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/{partner_id}/notifications", response_model=list[schemas_partner_patch.PartnerNotificationRead])
def get_partner_notifications(partner_id: UUID, db: Session = Depends(get_db)):
    partner = db.query(models.TradingPartner).filter(models.TradingPartner.partner_id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    return (
        db.query(PartnerNotification)
        .filter(PartnerNotification.partner_id == partner_id)
        .order_by(PartnerNotification.created_at.asc())
        .all()
    )


@router.post("/{partner_id}/notifications", response_model=schemas_partner_patch.PartnerNotificationRead)
def create_partner_notification(
    partner_id: UUID,
    payload: schemas_partner_patch.PartnerNotificationCreate,
    db: Session = Depends(get_db),
):
    return _save_partner_notification_row(db, partner_id, payload)


@router.post("/notifications", response_model=schemas_partner_patch.PartnerNotificationRead)
def create_partner_notification_legacy(
    payload: schemas_partner_patch.PartnerNotificationCreate,
    db: Session = Depends(get_db),
):
    return _save_partner_notification_row(db, payload.partner_id, payload)


@router.get("/{partner_id}/connections", response_model=list[schemas.TradingPartnerConnectionRead])
def get_partner_connections(partner_id: UUID, db: Session = Depends(get_db)):
    return (
        db.query(models.TradingPartnerConnection)
        .filter(models.TradingPartnerConnection.partner_id == partner_id)
        .order_by(models.TradingPartnerConnection.created_at.desc())
        .all()
    )


@router.post("/{partner_id}/connections", response_model=schemas.TradingPartnerConnectionRead)
def create_partner_connection(
    partner_id: UUID,
    payload: schemas.TradingPartnerConnectionCreate,
    db: Session = Depends(get_db),
):
    partner = db.query(models.TradingPartner).filter(models.TradingPartner.partner_id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    data = _normalize_connection_payload(payload)
    _validate_connection_payload(data)

    row = models.TradingPartnerConnection(**data)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/connections/{connection_id}", response_model=schemas.TradingPartnerConnectionRead)
def update_partner_connection(
    connection_id: UUID,
    payload: schemas.TradingPartnerConnectionUpdate,
    db: Session = Depends(get_db),
):
    row = db.query(models.TradingPartnerConnection).filter(models.TradingPartnerConnection.connection_id == connection_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Connection not found.")

    data = _normalize_connection_payload(payload, existing_row=row)
    _validate_connection_payload(
        {
            "connection_type": data.get("connection_type", row.connection_type),
            "is_active": data.get("is_active", row.is_active),
            "config_json": data.get("config_json", row.config_json),
        }
    )

    for field, value in data.items():
        setattr(row, field, value)

    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.post("/connections/{connection_id}/test")
def test_partner_connection(connection_id: UUID, db: Session = Depends(get_db)):
    row = (
        db.query(models.TradingPartnerConnection)
        .filter(models.TradingPartnerConnection.connection_id == connection_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Connection not found.")

    if str(row.connection_type or "").upper() != "EMAIL":
        raise HTTPException(status_code=400, detail="Only EMAIL connections support testing right now.")

    normalized = email_polling_service.normalize_email_config(
        client_id=row.client_id,
        connection_key=row.connection_id,
        config=row.config_json,
    )

    try:
        return email_polling_service.validate_connection(normalized)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Email connection test failed: {exc}") from exc


@router.post("/connections/{connection_id}/poll")
def poll_partner_connection(connection_id: UUID, db: Session = Depends(get_db)):
    row = (
        db.query(models.TradingPartnerConnection)
        .filter(models.TradingPartnerConnection.connection_id == connection_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Connection not found.")

    if str(row.connection_type or "").upper() != "EMAIL":
        raise HTTPException(status_code=400, detail="Only EMAIL connections support polling right now.")

    normalized = email_polling_service.normalize_email_config(
        client_id=row.client_id,
        connection_key=row.connection_id,
        config=row.config_json,
    )
    return email_polling_service.poll_connection(db, normalized)


@router.delete("/connections/{connection_id}")
def delete_partner_connection(connection_id: UUID, db: Session = Depends(get_db)):
    row = db.query(models.TradingPartnerConnection).filter(models.TradingPartnerConnection.connection_id == connection_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Connection not found.")
    db.delete(row)
    db.commit()
    return {"status": "deleted"}

@router.get("/{partner_id}/uom/template")
def download_uom_template(partner_id: UUID):
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info["A1"] = "UOM Template Instructions"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "1. Do not change header names"
    ws_info["A4"] = "2. Fill only data rows"
    ws_info["A5"] = "3. Upload back in same format"
    ws_info["A6"] = "4. Use dropdown values where available"

    ws = wb.create_sheet(title="UOM_Data")

    headers = [
        "client_id",
        "partner_id",
        "input_uom",
        "output_uom",
        "factor",
        "divider",
        "material_code",
        "rounding_digits",
        "rounding_mode",
        "is_active",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        cell.protection = Protection(locked=True)

    for row in ws.iter_rows(min_row=2, max_row=500):
        for cell in row:
            cell.protection = Protection(locked=False)

    dv_rounding = DataValidation(type="list", formula1='"HALF_UP,FLOOR,CEILING"', allow_blank=True)
    ws.add_data_validation(dv_rounding)
    dv_rounding.add("I2:I500")

    dv_active = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv_active)
    dv_active.add("J2:J500")

    ws.protection.sheet = True

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=uom_template.xlsx"},
    )


@router.post("/{partner_id}/uom/upload")
async def upload_uom_template(
    partner_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    contents = await file.read()
    result = validate_uom_workbook(contents)

    if not result.is_valid:
        return StreamingResponse(
            io.BytesIO(result.workbook_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=uom_validation_errors.xlsx",
                "X-Upload-Status": "validation_failed",
            },
        )
    wb = load_workbook(filename=io.BytesIO(contents))
    ws = wb["UOM_Data"]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        rows.append(
            {
                "client_id": row[0],
                "partner_id": row[1],
                "input_uom": row[2],
                "output_uom": row[3],
                "factor": row[4],
                "divider": row[5],
                "material_code": row[6],
                "rounding_digits": row[7],
                "rounding_mode": row[8],
                "is_active": row[9],
            }
        )

    return {
        "status": "success",
        "rows_processed": len(result.parsed_rows),
        "rows": result.parsed_rows,
    }

@router.get("/{partner_id}/bulk-onboarding/template")
def download_bulk_onboarding_template(partner_id: UUID):
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info["A1"] = "Bulk Onboarding Template Instructions"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "1. Do not change header names"
    ws_info["A4"] = "2. Fill only data rows"
    ws_info["A5"] = "3. Upload back in same format"
    ws_info["A6"] = "4. Use dropdown values where available"
    ws_info["A7"] = "5. Use req_* columns to choose MANDATORY / OPTIONAL / CONDITIONAL for common PO and invoice fields"
    ws_info["A8"] = "6. Use invoice_profile_type plus invoice source fields when the partner sends AP/AR invoices"
    ws_info["A9"] = "7. JSON validation columns are optional advanced overrides"
    ws_info["A10"] = '8. Example validation JSON: {"field_requirements":{"document_number":"MANDATORY","items.*.quantity":"MANDATORY"}}'

    ws = wb.create_sheet(title="Bulk_Onboarding")

    headers = [
        "client_id",
        "vertical_id",
        "partner_code",
        "partner_name",
        "partner_type",
        "status",
        "connection_method",
        "email",
        "edi_id",
        "sftp_path",
        "as2_id",
        "api_reference",
        *TARGET_PROFILE_TEMPLATE_COLUMNS,
        *TARGET_DEFAULT_TEMPLATE_COLUMNS,
        *FIELD_REQUIREMENT_TEMPLATE_COLUMNS,
        "mapping_validation_json",
        "business_rule_validation_json",
        "message_control_json",
        "notes",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        cell.protection = Protection(locked=True)

    for row in ws.iter_rows(min_row=2, max_row=500):
        for cell in row:
            cell.protection = Protection(locked=False)

    dv_partner_type = DataValidation(type="list", formula1='"CUSTOMER,SUPPLIER,LOGISTICS_PROVIDER"', allow_blank=True)
    ws.add_data_validation(dv_partner_type)
    dv_partner_type.add("E2:E500")

    dv_status = DataValidation(type="list", formula1='"ACTIVE,INACTIVE"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("F2:F500")

    dv_connection = DataValidation(type="list", formula1='"EMAIL,EDI,SFTP,AS2,API"', allow_blank=True)
    ws.add_data_validation(dv_connection)
    dv_connection.add("G2:G500")

    for header_name in ["target_erp", "invoice_profile_type", "customization_required"]:
        col_idx = headers.index(header_name) + 1
        col_letter = get_column_letter(col_idx)
        if header_name == "target_erp":
            formula = '"SAP,ORACLE,D365,JDE,API"'
        elif header_name == "invoice_profile_type":
            formula = '"AP_INVOICE,AR_INVOICE,INVOICE"'
        else:
            formula = '"TRUE,FALSE"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}500")

    requirement_list_formula = '"MANDATORY,OPTIONAL,CONDITIONAL"'
    for header_name in FIELD_REQUIREMENT_TEMPLATE_COLUMNS:
        col_idx = headers.index(header_name) + 1
        col_letter = get_column_letter(col_idx)
        dv_requirement = DataValidation(type="list", formula1=requirement_list_formula, allow_blank=True)
        ws.add_data_validation(dv_requirement)
        dv_requirement.add(f"{col_letter}2:{col_letter}500")

    ws.protection.sheet = True

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_onboarding_template.xlsx"},
    )


@router.post("/{partner_id}/bulk-onboarding/upload")
async def upload_bulk_onboarding(
    partner_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    contents = await file.read()
    result = validate_bulk_onboarding_workbook(contents)

    if not result.is_valid:
        return StreamingResponse(
            io.BytesIO(result.workbook_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=bulk_onboarding_validation_errors.xlsx",
                "X-Upload-Status": "validation_failed",
            },
        )
    wb = load_workbook(filename=io.BytesIO(contents))
    ws = wb["Bulk_Onboarding"]
    workbook_headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        row_map = {
            header_name: row[idx]
            for idx, header_name in enumerate(workbook_headers)
            if header_name
        }

        rows.append(
            {
                **row_map,
            }
        )

    persisted_rows = []
    for parsed_row in result.parsed_rows:
        partner = (
            db.query(models.TradingPartner)
            .filter(
                models.TradingPartner.client_id == parsed_row["client_id"],
                models.TradingPartner.partner_code == parsed_row["partner_code"],
            )
            .first()
        )

        partner_before = _model_to_dict(partner) if partner else None
        partner_action = "UPDATE" if partner else "CREATE"

        if partner:
            partner.vertical_id = parsed_row["vertical_id"]
            partner.partner_name = parsed_row["partner_name"]
            partner.partner_type = parsed_row["partner_type"]
            partner.status = parsed_row["status"]
            partner.connection_method = parsed_row["connection_method"]
            partner.email = parsed_row["email"]
            partner.edi_id = parsed_row["edi_id"]
            partner.sftp_path = parsed_row["sftp_path"]
            partner.as2_id = parsed_row["as2_id"]
            partner.api_reference = parsed_row["api_reference"]
            partner.notes = parsed_row["notes"]
            db.add(partner)
            db.flush()
        else:
            partner = models.TradingPartner(
                client_id=parsed_row["client_id"],
                vertical_id=parsed_row["vertical_id"],
                partner_code=parsed_row["partner_code"],
                partner_name=parsed_row["partner_name"],
                partner_type=parsed_row["partner_type"],
                status=parsed_row["status"],
                connection_method=parsed_row["connection_method"],
                email=parsed_row["email"],
                edi_id=parsed_row["edi_id"],
                sftp_path=parsed_row["sftp_path"],
                as2_id=parsed_row["as2_id"],
                api_reference=parsed_row["api_reference"],
                notes=parsed_row["notes"],
            )
            db.add(partner)
            db.flush()

        _ensure_default_partner_profile(db, partner)

        mapping_profile, mapping_before = _upsert_bulk_mapping_validation(
            db,
            partner=partner,
            validation_json=parsed_row.get("mapping_validation_json"),
            target_defaults=parsed_row.get("target_defaults_json"),
            layout_hints={
                "target_profile": dict(parsed_row.get("target_profile_json") or {}),
                "customization": dict(parsed_row.get("customization_json") or {}),
            },
        )
        business_rule, rule_before = _upsert_bulk_business_validation_rule(
            db,
            partner=partner,
            action_json=parsed_row.get("business_rule_validation_json"),
        )

        write_audit(
            db,
            client_id=partner.client_id,
            partner_id=str(partner.partner_id),
            entity_type="MASTER",
            entity_id=str(partner.partner_id),
            action=partner_action,
            before_json=partner_before,
            after_json=_model_to_dict(partner),
            remarks="Bulk onboarding import",
        )

        if mapping_profile is not None:
            write_audit(
                db,
                client_id=partner.client_id,
                partner_id=str(partner.partner_id),
                entity_type="MAPPING",
                entity_id=str(mapping_profile.mapping_profile_id),
                action="UPDATE" if mapping_before else "CREATE",
                before_json=mapping_before,
                after_json=_model_to_dict(mapping_profile),
                remarks="Bulk onboarding mapping validation import",
            )

        if business_rule is not None:
            write_audit(
                db,
                client_id=partner.client_id,
                partner_id=str(partner.partner_id),
                entity_type="RULE",
                entity_id=str(business_rule.rule_id),
                action="UPDATE" if rule_before else "CREATE",
                before_json=rule_before,
                after_json=_model_to_dict(business_rule),
                remarks="Bulk onboarding business rule validation import",
            )

        persisted_rows.append(
            {
                "partner_id": str(partner.partner_id),
                "partner_code": partner.partner_code,
                "partner_name": partner.partner_name,
                "status": partner.status,
                "mapping_validation_applied": mapping_profile is not None,
                "business_rule_validation_applied": business_rule is not None,
            }
        )

    db.commit()

    return {
        "status": "success",
        "rows_processed": len(result.parsed_rows),
        "rows": persisted_rows,
    }

@router.get("/{partner_id}/mappings")
def get_partner_mappings(partner_id: UUID, db: Session = Depends(get_db)):
    if not hasattr(models, "TradingPartnerMapping"):
        return []

    rows = (
        db.query(models.TradingPartnerMapping)
        .filter(models.TradingPartnerMapping.partner_id == partner_id)
        .order_by(models.TradingPartnerMapping.created_at.desc())
        .all()
    )
    return rows

@router.get("/uom-rules")
def get_uom_rules(partner_id: UUID, db: Session = Depends(get_db)):
    if not hasattr(models, "TradingPartnerUomRule"):
        return []

    rows = (
        db.query(models.TradingPartnerUomRule)
        .filter(models.TradingPartnerUomRule.partner_id == partner_id)
        .order_by(models.TradingPartnerUomRule.created_at.desc())
        .all()
    )
    return rows


@router.get("/{partner_id}/promotion-package")
def download_promotion_package(partner_id: UUID, db: Session = Depends(get_db)):
    if not is_staging():
        raise HTTPException(status_code=403, detail="Promotion package export is available only in staging.")

    partner = db.query(models.TradingPartner).filter(models.TradingPartner.partner_id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Trading partner not found.")

    client = db.query(models.Client).filter(models.Client.client_id == partner.client_id).first()
    business_verticals = (
        db.query(models.BusinessVertical)
        .filter(models.BusinessVertical.client_id == partner.client_id)
        .order_by(models.BusinessVertical.vertical_name.asc())
        .all()
    )
    client_connections = (
        db.query(models.ClientConnection)
        .filter(models.ClientConnection.client_id == partner.client_id)
        .order_by(models.ClientConnection.created_at.asc())
        .all()
    )
    client_erp_configs = (
        db.query(models.ClientERPConfig)
        .filter(models.ClientERPConfig.client_id == partner.client_id)
        .order_by(models.ClientERPConfig.created_at.asc())
        .all()
    )
    client_configs = (
        db.query(models.ClientConfig)
        .filter(models.ClientConfig.client_id == partner.client_id)
        .order_by(models.ClientConfig.created_at.asc())
        .all()
    )
    client_email_config = (
        db.query(models.ClientEmailConfig)
        .filter(models.ClientEmailConfig.client_id == partner.client_id)
        .first()
    )

    profile = db.query(models.TradingPartnerProfile).filter(models.TradingPartnerProfile.partner_id == partner_id).first()
    connections = (
        db.query(models.TradingPartnerConnection)
        .filter(models.TradingPartnerConnection.partner_id == partner_id)
        .order_by(models.TradingPartnerConnection.created_at.asc())
        .all()
    )
    mapping_profiles = (
        db.query(TradingPartnerMappingProfile)
        .filter(TradingPartnerMappingProfile.partner_id == partner_id)
        .order_by(TradingPartnerMappingProfile.created_at.asc())
        .all()
    )
    business_rules = (
        db.query(TradingPartnerBusinessRule)
        .filter(TradingPartnerBusinessRule.partner_id == partner_id)
        .order_by(TradingPartnerBusinessRule.created_at.asc())
        .all()
    )
    uom_rules = (
        db.query(TradingPartnerUomRule)
        .filter(TradingPartnerUomRule.partner_id == partner_id)
        .order_by(TradingPartnerUomRule.created_at.asc())
        .all()
    )
    address_master = (
        db.query(models.AddressMaster)
        .filter(models.AddressMaster.partner_id == partner_id)
        .order_by(models.AddressMaster.created_at.asc())
        .all()
    )
    message_flows = (
        db.query(models.MessageFlow)
        .filter(models.MessageFlow.partner_id == partner_id)
        .order_by(models.MessageFlow.priority.asc(), models.MessageFlow.flow_name.asc())
        .all()
    )
    parser_profiles = (
        db.query(models.ParserProfile)
        .filter(models.ParserProfile.partner_id == partner_id)
        .order_by(models.ParserProfile.priority.asc(), models.ParserProfile.created_at.asc())
        .all()
    )
    notifications = (
        db.query(PartnerNotification)
        .filter(PartnerNotification.partner_id == partner_id)
        .order_by(PartnerNotification.created_at.asc())
        .all()
    )

    package = _promotion_package(
        client,
        business_verticals,
        client_connections,
        client_erp_configs,
        client_configs,
        client_email_config,
        partner,
        profile,
        connections,
        mapping_profiles,
        business_rules,
        uom_rules,
        address_master,
        message_flows,
        parser_profiles,
        notifications,
    )
    summary = dict(package.get("meta", {}).get("domain_counts") or {})
    summary.update(
        {
            "package_version": package.get("meta", {}).get("package_version"),
            "source_environment": current_environment(),
            "target_environment": PROMOTION_TARGET_ENVIRONMENT,
            "config_domains": package.get("meta", {}).get("config_domains") or [],
        }
    )
    write_audit(
        db,
        client_id=getattr(client, "client_id", partner.client_id),
        partner_id=str(partner.partner_id),
        entity_type="PROMOTION",
        entity_id=f"{partner.partner_code}:staging_to_production",
        action="EXPORT",
        before_json=None,
        after_json=summary,
        actor_email="system",
        actor_role="SYSTEM",
        remarks="Promotion package exported from staging for production import.",
    )
    db.commit()

    payload = json.dumps(package, default=str, indent=2).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename={partner.partner_code}_promotion_package.json",
        },
    )


@router.post("/promotion-import")
def import_promotion_package(payload: dict, db: Session = Depends(get_db)):
    if not is_production():
        raise HTTPException(status_code=403, detail="Promotion import is available only in production.")

    meta = dict(payload.get("meta") or {})
    validation_errors = []

    source_environment = str(meta.get("source_environment") or "").lower()
    if source_environment != "staging":
        validation_errors.append("Only staging promotion packages can be imported.")

    if str(meta.get("package_kind") or "") != "TRADING_PARTNER_PROMOTION":
        validation_errors.append("Promotion package kind is invalid or missing.")

    partner_payload = dict(payload.get("partner") or {})
    client_payload = dict(payload.get("client") or {})
    if not partner_payload:
        validation_errors.append("Promotion package is missing trading partner data.")

    client_id = client_payload.get("client_id") or partner_payload.get("client_id")
    partner_code = partner_payload.get("partner_code")
    if not client_id or not partner_code:
        validation_errors.append("Promotion package is missing client_id or partner_code.")

    if validation_errors:
        raise HTTPException(status_code=400, detail=" ".join(validation_errors))

    existing_client = db.query(models.Client).filter(models.Client.client_id == client_id).first()
    client_before = _model_to_dict(existing_client) if existing_client else None
    client = existing_client
    if not client:
        client = models.Client(client_id=client_id)
        db.add(client)
        db.flush()

    partner = (
        db.query(models.TradingPartner)
        .filter(models.TradingPartner.client_id == client_id, models.TradingPartner.partner_code == partner_code)
        .first()
    )
    partner_before = _model_to_dict(partner) if partner else None
    before_counts = {
        "client": 1 if client else 0,
        "business_verticals": db.query(models.BusinessVertical).filter(models.BusinessVertical.client_id == client.client_id).count() if client else 0,
        "client_connections": db.query(models.ClientConnection).filter(models.ClientConnection.client_id == client.client_id).count() if client else 0,
        "client_erp_configs": db.query(models.ClientERPConfig).filter(models.ClientERPConfig.client_id == client.client_id).count() if client else 0,
        "client_configs": db.query(models.ClientConfig).filter(models.ClientConfig.client_id == client.client_id).count() if client else 0,
        "client_email_config": db.query(models.ClientEmailConfig).filter(models.ClientEmailConfig.client_id == client.client_id).count() if client else 0,
        "connections": db.query(models.TradingPartnerConnection).filter(models.TradingPartnerConnection.partner_id == partner.partner_id).count() if partner else 0,
        "mapping_profiles": db.query(TradingPartnerMappingProfile).filter(TradingPartnerMappingProfile.partner_id == partner.partner_id).count() if partner else 0,
        "business_rules": db.query(TradingPartnerBusinessRule).filter(TradingPartnerBusinessRule.partner_id == partner.partner_id).count() if partner else 0,
        "uom_rules": db.query(TradingPartnerUomRule).filter(TradingPartnerUomRule.partner_id == partner.partner_id).count() if partner else 0,
        "address_master": db.query(models.AddressMaster).filter(models.AddressMaster.partner_id == partner.partner_id).count() if partner else 0,
        "message_flows": db.query(models.MessageFlow).filter(models.MessageFlow.partner_id == partner.partner_id).count() if partner else 0,
        "parser_profiles": db.query(models.ParserProfile).filter(models.ParserProfile.partner_id == partner.partner_id).count() if partner else 0,
        "notifications": db.query(PartnerNotification).filter(PartnerNotification.partner_id == partner.partner_id).count() if partner else 0,
    }

    if not partner:
        partner = models.TradingPartner(client_id=client.client_id)
        db.add(partner)
        db.flush()

    client_data = _promotion_importable_payload(
        models.Client,
        client_payload or partner_payload,
        client_id=client.client_id,
        preserve_primary_keys=True,
    ) or {}
    for field, value in client_data.items():
        setattr(client, field, value)
    db.add(client)
    db.flush()

    if "business_verticals" in payload:
        _replace_client_domain_rows(
            db,
            models.BusinessVertical,
            client=client,
            rows=list(payload.get("business_verticals") or []),
        )
    if "client_connections" in payload:
        _replace_client_domain_rows(
            db,
            models.ClientConnection,
            client=client,
            rows=list(payload.get("client_connections") or []),
        )
    if "client_erp_configs" in payload:
        _replace_client_domain_rows(
            db,
            models.ClientERPConfig,
            client=client,
            rows=list(payload.get("client_erp_configs") or []),
        )
    if "client_configs" in payload:
        _replace_client_domain_rows(
            db,
            models.ClientConfig,
            client=client,
            rows=list(payload.get("client_configs") or []),
        )
    if "client_email_config" in payload:
        _replace_client_domain_rows(
            db,
            models.ClientEmailConfig,
            client=client,
            rows=[payload.get("client_email_config")] if payload.get("client_email_config") else [],
        )

    partner_data = _promotion_importable_payload(
        models.TradingPartner,
        partner_payload,
        client_id=client.client_id,
        partner_id=partner.partner_id,
        preserve_primary_keys=True,
    ) or {}
    for field, value in partner_data.items():
        setattr(partner, field, value)
    db.add(partner)
    db.flush()

    profile_payload = dict(payload.get("profile") or {})
    if profile_payload:
        profile = db.query(models.TradingPartnerProfile).filter(models.TradingPartnerProfile.partner_id == partner.partner_id).first()
        if not profile:
            profile = models.TradingPartnerProfile(client_id=partner.client_id, partner_id=partner.partner_id)
            db.add(profile)
            db.flush()
        profile_data = _promotion_importable_payload(
            models.TradingPartnerProfile,
            profile_payload,
            client_id=partner.client_id,
            partner_id=partner.partner_id,
            preserve_primary_keys=True,
        ) or {}
        for field, value in profile_data.items():
            setattr(profile, field, value)
        db.add(profile)
    else:
        profile = None

    if "connections" in payload:
        _replace_partner_domain_rows(db, models.TradingPartnerConnection, partner=partner, rows=list(payload.get("connections") or []))
    if "mapping_profiles" in payload:
        _replace_partner_domain_rows(db, TradingPartnerMappingProfile, partner=partner, rows=list(payload.get("mapping_profiles") or []))
    if "business_rules" in payload:
        _replace_partner_domain_rows(db, TradingPartnerBusinessRule, partner=partner, rows=list(payload.get("business_rules") or []))
    if "uom_rules" in payload:
        _replace_partner_domain_rows(db, TradingPartnerUomRule, partner=partner, rows=list(payload.get("uom_rules") or []))
    if "address_master" in payload:
        _replace_partner_domain_rows(db, models.AddressMaster, partner=partner, rows=list(payload.get("address_master") or []))
    if "message_flows" in payload:
        _replace_partner_domain_rows(db, models.MessageFlow, partner=partner, rows=list(payload.get("message_flows") or []))
    if "parser_profiles" in payload:
        _replace_partner_domain_rows(db, models.ParserProfile, partner=partner, rows=list(payload.get("parser_profiles") or []))
    if "notifications" in payload:
        _replace_partner_domain_rows(db, PartnerNotification, partner=partner, rows=list(payload.get("notifications") or []))

    summary = {
        "package_version": meta.get("package_version") or PROMOTION_PACKAGE_VERSION,
        "source_environment": meta.get("source_environment"),
        "target_environment": current_environment(),
        "config_domains": meta.get("config_domains") or list(PROMOTION_DOMAIN_NAMES),
        "before_counts": before_counts,
        "imported_counts": {
            "client": 1,
            "business_verticals": len(list(payload.get("business_verticals") or [])),
            "client_connections": len(list(payload.get("client_connections") or [])),
            "client_erp_configs": len(list(payload.get("client_erp_configs") or [])),
            "client_configs": len(list(payload.get("client_configs") or [])),
            "client_email_config": 1 if payload.get("client_email_config") else 0,
            "partner": 1,
            "profile": 1 if profile_payload else 0,
            "connections": len(list(payload.get("connections") or [])),
            "mapping_profiles": len(list(payload.get("mapping_profiles") or [])),
            "business_rules": len(list(payload.get("business_rules") or [])),
            "uom_rules": len(list(payload.get("uom_rules") or [])),
            "address_master": len(list(payload.get("address_master") or [])),
            "message_flows": len(list(payload.get("message_flows") or [])),
            "parser_profiles": len(list(payload.get("parser_profiles") or [])),
            "notifications": len(list(payload.get("notifications") or [])),
        },
        "validation": {
            "status": "PASSED",
            "errors": [],
        },
    }
    write_audit(
        db,
        client_id=client.client_id,
        partner_id=str(partner.partner_id),
        entity_type="PROMOTION",
        entity_id=f"{partner.partner_code}:staging_to_production",
        action="IMPORT",
        before_json={"client": client_before, "partner": partner_before, "counts": before_counts},
        after_json=summary,
        actor_email="system",
        actor_role="SYSTEM",
        remarks="Promotion package imported into production.",
    )
    db.commit()

    return {
        "status": "IMPORTED",
        "environment": current_environment(),
        "partner_id": str(partner.partner_id),
        "partner_code": partner.partner_code,
        "summary": summary,
    }


@router.get("/{partner_id}/promotion-history")
def get_promotion_history(partner_id: UUID, db: Session = Depends(get_db)):
    rows = (
        db.query(TradingPartnerOnboardingAudit)
        .filter(
            TradingPartnerOnboardingAudit.partner_id == partner_id,
            TradingPartnerOnboardingAudit.entity_type == "PROMOTION",
        )
        .order_by(TradingPartnerOnboardingAudit.created_at.desc())
        .all()
    )
    return [
        {
            "audit_id": str(row.audit_id),
            "action": row.action,
            "entity_id": row.entity_id,
            "actor_email": row.actor_email,
            "actor_role": row.actor_role,
            "remarks": row.remarks,
            "created_at": row.created_at,
            "before_json": row.before_json,
            "after_json": row.after_json,
        }
        for row in rows
    ]
