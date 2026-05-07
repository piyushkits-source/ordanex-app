import os
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session
from uuid import UUID
from backend.db.database import get_db
from backend.db import schemas_trading_partner_agentic as schemas
from backend.services.agentic_onboarding_service import agentic_onboarding_service

router = APIRouter(prefix="/trading-partners-agentic", tags=["Trading Partners Agentic"])


def _is_staging_environment() -> bool:
    candidates = [
        os.getenv("APP_ENV"),
        os.getenv("ORDANEX_ENV"),
        os.getenv("ENVIRONMENT"),
        os.getenv("FASTAPI_ENV"),
    ]
    return any(str(value or "").strip().lower() == "staging" for value in candidates)


HEADER_FILL = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EEF6FF", end_color="EEF6FF", fill_type="solid")


def _format_sheet(sheet):
    widths = {
        "A": 26,
        "B": 28,
        "C": 28,
        "D": 32,
        "E": 30,
        "F": 26,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def _add_header_row(sheet, row_idx: int, labels: list[str]):
    for idx, label in enumerate(labels, start=1):
        cell = sheet.cell(row=row_idx, column=idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL




def _add_message_control_sheet(wb, title: str = "Message Control") -> None:
    ws = wb.create_sheet(title)
    _add_header_row(ws, 1, ["Control", "Value", "Example / Guidance"])
    rows = [
        ("Horizon Mode", "", "Rolling Days / Relative Months / Absolute Date"),
        ("Horizon Value", "", "Example: 90 or 3"),
        ("Horizon Anchor Field", "", "Example: requested_delivery_date"),
        ("Firm Indicators", "", "Comma-separated values like FIRM, FIXED, CONFIRMED"),
        ("Forecast Indicators", "", "Comma-separated values like FORECAST, FCST, ESTIMATED"),
        ("No Indicator Policy", "", "Use Horizon / Treat as Firm / Treat as Forecast"),
        ("Compare Fields", "", "Comma-separated values like material_code, requested_delivery_date, quantity"),
        ("Zero Quantity Action", "", "CANCEL_ORDER / FLAG_REVIEW"),
        ("Missing Line Action", "", "CANCEL_ORDER / FLAG_REVIEW"),
        ("Outside Horizon Action", "", "NEW_ORDER / FLAG_REVIEW"),
        ("Forecast Action", "", "EMAIL_ONLY / EMAIL_AND_FLAG / FLAG_REVIEW"),
        ("Forecast Email Subject", "", "Example: Forecast received for {{document_number}}"),
        ("Forecast Email Body HTML", "", "Optional HTML template for business-friendly forecast email"),
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws)
def _build_mapping_spec_workbook() -> tuple[bytes, str]:
    wb = Workbook()
    ws_intro = wb.active
    ws_intro.title = "Overview"
    ws_intro["A1"] = "Ordanex Business-Friendly Mapping Workbook"
    ws_intro["A1"].font = Font(bold=True, size=14)
    intro_lines = [
        "Use this workbook if you are a business user onboarding a new trading partner.",
        "Fill the business questions first. You do not need to know EDI segments or canonical field names.",
        "Upload sample files separately in AI Onboarding after filling this workbook.",
        "If you do not know an answer, leave it blank and the AI onboarding flow can still draft the config.",
        "Only use the Advanced Mapping sheet if a technical analyst wants to give exact source-to-target details.",
    ]
    for idx, line in enumerate(intro_lines, start=3):
        ws_intro[f"A{idx}"] = f"{idx - 2}. {line}"
    ws_intro["A10"] = "Expected result"
    ws_intro["A10"].font = Font(bold=True)
    ws_intro["A11"] = "The system will use your answers plus uploaded samples to draft mapping, validation, UOM, address, and target output configuration."
    _format_sheet(ws_intro)

    ws_profile = wb.create_sheet("Business Questionnaire")
    _add_header_row(ws_profile, 1, ["Question", "Answer", "Example / Guidance"])
    business_questions = [
        ("Client ID", "", "Example: DU0001"),
        ("Trading Partner Code", "", "Example: ABC0001"),
        ("Trading Partner Name", "", "Example: ABC Corp"),
        ("What type of partner is this?", "", "CUSTOMER / SUPPLIER / LOGISTICS_PROVIDER"),
        ("How will they send messages?", "", "EMAIL / EDI / SFTP / AS2 / API"),
        ("What kind of document will they send first?", "", "Purchase Order / Order Response / ASN / AP Invoice / AR Invoice / Invoice"),
        ("What format will they send?", "", "PDF / Excel / CSV / IDOC / X12 / EDIFACT / XML / JSON / API"),
        ("Which ERP should the output go to?", "", "SAP / Oracle / D365 / JDE / API"),
        ("Does this partner need custom output behavior?", "", "Yes / No"),
        ("If yes, describe the customization", "", "Example: custom SAP/JDE text IDs or special field logic"),
        ("Who should receive alert emails?", "", "One or more email addresses"),
        ("What is the business priority of this onboarding?", "", "High / Medium / Low"),
    ]
    for row_idx, row in enumerate(business_questions, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_profile.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_profile)
    _add_message_control_sheet(wb)

    ws_fields = wb.create_sheet("Field Rules")
    _add_header_row(ws_fields, 1, ["Business Field", "Required Level", "Business Meaning / Notes", "Example Value"])
    field_rows = [
        ("PO / Document Number", "MANDATORY", "Unique document number from the source order", "4500223977"),
        ("PO / Document Date", "OPTIONAL", "Date shown on the source order", "2026-04-24"),
        ("Customer / Buyer Name", "MANDATORY", "Who is placing the order", "Customer Demo"),
        ("Supplier Name", "MANDATORY", "Who will fulfill the order", "DuPont"),
        ("Currency", "OPTIONAL", "Preferred three-letter currency code", "USD"),
        ("Ship-To Code", "OPTIONAL", "Delivery location code if provided", "HOUSTON_DC"),
        ("Item Product Code", "MANDATORY", "Product or material identifier", "1874015"),
        ("Item Description", "OPTIONAL", "Product description line", "Pad FH2 20"),
        ("Item Quantity", "MANDATORY", "Ordered quantity", "30"),
        ("Item UOM", "OPTIONAL", "Unit of measure", "EA"),
        ("Item Delivery Date", "OPTIONAL", "Requested delivery date", "2026-11-03"),
        ("Item Unit Price", "OPTIONAL", "Price per unit if available", "420"),
    ]
    for row_idx, row in enumerate(field_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_fields.cell(row=row_idx, column=col_idx, value=value)
    req_dv = DataValidation(type="list", formula1='"MANDATORY,OPTIONAL,CONDITIONAL"', allow_blank=True)
    ws_fields.add_data_validation(req_dv)
    req_dv.add("B2:B50")
    _format_sheet(ws_fields)

    ws_target = wb.create_sheet("Target Output")
    _add_header_row(ws_target, 1, ["Question", "Answer", "Example / Guidance"])
    target_rows = [
        ("Target ERP", "", "SAP / Oracle / D365 / JDE / API"),
        ("Target Standard", "", "IDOC / XML / JSON / API / X12 / EDIFACT"),
        ("Target Message Type", "", "ORDERS / ORDRSP / DESADV / INVOIC / 810"),
        ("Target Version", "", "ORDERS05 / D96A / 4010 / 5010"),
        ("Invoice Profile Type", "", "AP_INVOICE / AR_INVOICE / INVOICE"),
        ("Transaction ID Source", "", "PO Number / Delivery Number / Billing Document Number"),
        ("Invoice Number Source", "", "Billing Document Number / Invoice Number"),
        ("Invoice Date Source", "", "Billing Date / Invoice Date"),
        ("Invoice Total Source", "", "Gross Amount / Net Amount / Invoice Total"),
        ("Header Text ID", "", "Example: Z001"),
        ("Line Text ID", "", "Example: ZL01"),
        ("Default Order Type", "", "Example: ZOR"),
        ("Need manual review before outbound delivery?", "", "Yes / No"),
    ]
    for row_idx, row in enumerate(target_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_target.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_target)

    ws_test = wb.create_sheet("Test Scenarios")
    _add_header_row(ws_test, 1, ["Scenario", "What should happen?", "Sample Available?", "Notes"])
    scenarios = [
        ("Happy path", "Document should auto-process with no edits", "Yes", ""),
        ("Missing field", "Document should go Pending and show missing-field reason", "No", ""),
        ("Multiple lines", "Each line should be extracted and mapped correctly", "Yes", ""),
        ("Special formatting", "Dates/UOM/currency should normalize correctly", "No", ""),
    ]
    for row_idx, row in enumerate(scenarios, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_test.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_test)

    ws_advanced = wb.create_sheet("Advanced Mapping")
    _add_header_row(ws_advanced, 1, ["Source Field / Segment", "Business Meaning", "Canonical Field", "Target Field", "Notes"])
    advanced_rows = [
        ("N/A", "Only fill this sheet if a technical analyst wants exact mapping details", "", "", ""),
        ("", "PO Number", "document_number", "", ""),
        ("", "PO Date", "document_date", "", ""),
        ("", "Customer Name", "customer_name", "", ""),
        ("", "Supplier Name", "supplier_name", "", ""),
        ("", "Line Product", "items.material_code", "", ""),
        ("", "Line Quantity", "items.quantity", "", ""),
    ]
    for row_idx, row in enumerate(advanced_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_advanced.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_advanced)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue(), "ordanex_partner_onboarding_questionnaire.xlsx"


def _build_edi_guideline_workbook() -> tuple[bytes, str]:
    wb = Workbook()
    ws_intro = wb.active
    ws_intro.title = "Overview"
    ws_intro["A1"] = "Ordanex EDI / Interface Checklist"
    ws_intro["A1"].font = Font(bold=True, size=14)
    intro_lines = [
        "Use this workbook to capture partner EDI or interface details in business language.",
        "Fill the summary and checklist sheets first.",
        "If you have a formal partner guideline, upload it separately in AI Onboarding.",
        "Only use the Segment Details sheet if someone knows the EDI structure.",
    ]
    for idx, line in enumerate(intro_lines, start=3):
        ws_intro[f"A{idx}"] = f"{idx - 2}. {line}"
    _format_sheet(ws_intro)

    ws_summary = wb.create_sheet("Interface Summary")
    _add_header_row(ws_summary, 1, ["Question", "Answer", "Example / Guidance"])
    summary_rows = [
        ("Client ID", "", "Example: DU0001"),
        ("Trading Partner Code", "", "Example: ABC0001"),
        ("Trading Partner Name", "", "Example: ABC Corp"),
        ("Direction", "", "INBOUND / OUTBOUND"),
        ("Message Standard", "", "X12 / EDIFACT / XML / JSON / API / IDOC"),
        ("Message Type", "", "850 / ORDERS / ORDRSP / DESADV / INVOIC / 810"),
        ("Version", "", "4010 / 5010 / D96A / D01B"),
        ("Connection Type", "", "EMAIL / SFTP / AS2 / API"),
        ("Sender ID", "", "If known"),
        ("Receiver ID", "", "If known"),
    ]
    for row_idx, row in enumerate(summary_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_summary)

    ws_checklist = wb.create_sheet("Business Checklist")
    _add_header_row(ws_checklist, 1, ["Business Question", "Answer", "Notes"])
    checklist_rows = [
        ("What business event does this message represent?", "", "Purchase Order / Order Response / ASN / AP Invoice / AR Invoice / Invoice"),
        ("Should missing PO number block processing?", "", "Yes / No"),
        ("Should missing item quantity block processing?", "", "Yes / No"),
        ("Should missing UOM block processing?", "", "Yes / No"),
        ("Which business number should appear as the transaction ID?", "", "PO Number / Delivery Number / Billing Document Number"),
        ("Do we need invoice number and date in the output?", "", "Yes / No"),
        ("Do we need invoice total in the output?", "", "Yes / No"),
        ("Do we need header text in the output?", "", "Yes / No"),
        ("Do we need line text in the output?", "", "Yes / No"),
        ("Does this partner require custom output behavior?", "", "Yes / No"),
        ("Who gets issue alert emails?", "", "Business users / IT admins / both"),
    ]
    for row_idx, row in enumerate(checklist_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_checklist.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_checklist)
    _add_message_control_sheet(wb)

    ws_segment = wb.create_sheet("Segment Details")
    _add_header_row(ws_segment, 1, ["Segment / Field", "Business Meaning", "Expected Value / Rule", "Mandatory?", "Example"])
    segment_rows = [
        ("BGM / BEG", "Document number", "", "Yes", ""),
        ("DTM", "Document date", "", "Optional", ""),
        ("NAD / N1", "Customer / Buyer", "", "Yes", ""),
        ("NAD / N1", "Supplier / Seller", "", "Yes", ""),
        ("RFF / BGM", "Invoice / billing reference", "", "Optional", ""),
        ("MOA / TDS", "Invoice total", "", "Optional", ""),
        ("LIN / PO1", "Item product code", "", "Yes", ""),
        ("QTY / PO1", "Item quantity", "", "Yes", ""),
        ("PRI / CTP", "Unit price", "", "Optional", ""),
    ]
    for row_idx, row in enumerate(segment_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_segment.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_segment)

    ws_target = wb.create_sheet("Target Output")
    _add_header_row(ws_target, 1, ["Question", "Answer", "Example / Guidance"])
    target_rows = [
        ("Target ERP", "", "SAP / Oracle / D365 / JDE / API"),
        ("Target Standard", "", "IDOC / XML / JSON / API"),
        ("Target Message Type", "", "ORDERS / ORDRSP / DESADV / INVOIC"),
        ("Target Version", "", "ORDERS05 / D96A / 4010"),
        ("Header Text ID", "", "Example: Z001"),
        ("Line Text ID", "", "Example: ZL01"),
        ("Partner-specific customization notes", "", "Business-friendly description is enough"),
    ]
    for row_idx, row in enumerate(target_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_target.cell(row=row_idx, column=col_idx, value=value)
    _format_sheet(ws_target)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue(), "ordanex_edi_partner_checklist.xlsx"


@router.get("/templates/{template_kind}")
def download_onboarding_template(template_kind: str):
    kind = str(template_kind or "").strip().lower()
    templates = {
        "mapping-spec": _build_mapping_spec_workbook,
        "edi-guideline": _build_edi_guideline_workbook,
    }
    builder = templates.get(kind)
    if builder is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found.")

    content, filename = builder()
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )

@router.get("/projects", response_model=list[schemas.AgenticProjectRead])
def get_projects(partner_id: UUID, db: Session = Depends(get_db)):
    return agentic_onboarding_service.list_projects(db, partner_id)

@router.get("/projects/{project_id}", response_model=schemas.AgenticProjectRead)
def get_project(project_id: UUID, db: Session = Depends(get_db)):
    return agentic_onboarding_service.get_project(db, project_id)

@router.post("/projects", response_model=schemas.AgenticProjectRead)
def create_project(payload: schemas.AgenticProjectCreate, db: Session = Depends(get_db)):
    return agentic_onboarding_service.create_project(db, payload)


@router.post("/projects/{project_id}/sample-upload", response_model=schemas.AgenticProjectRead)
async def upload_project_sample(project_id: UUID, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not _is_staging_environment():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Sample upload is enabled only in staging.",
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded sample is empty.")

    return agentic_onboarding_service.attach_sample_file(
        db,
        project_id,
        file_name=file.filename or "sample.bin",
        content=content,
        mime_type=file.content_type,
        uploaded_by="agentic_onboarding",
    )


@router.post("/projects/{project_id}/artifact-upload", response_model=schemas.AgenticProjectRead)
async def upload_project_artifact(
    project_id: UUID,
    artifact_type: str = Form(...),
    scenario_label: str | None = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not _is_staging_environment():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Onboarding artifact upload is enabled only in staging.",
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded artifact is empty.")

    return agentic_onboarding_service.attach_project_artifact(
        db,
        project_id,
        artifact_type=artifact_type,
        scenario_label=scenario_label,
        file_name=file.filename or "artifact.bin",
        content=content,
        mime_type=file.content_type,
        uploaded_by="agentic_onboarding",
    )

@router.patch("/projects/{project_id}", response_model=schemas.AgenticProjectRead)
def update_project(project_id: UUID, payload: schemas.AgenticProjectUpdate, db: Session = Depends(get_db)):
    return agentic_onboarding_service.update_project(db, project_id, payload)

@router.post("/projects/{project_id}/advance", response_model=schemas.AgenticProjectRead)
def advance_project(project_id: UUID, payload: schemas.AgenticProjectAdvance, db: Session = Depends(get_db)):
    return agentic_onboarding_service.advance_project(db, project_id, payload)

@router.post("/discover", response_model=schemas.AgenticDiscoveryResponse)
def discover(payload: schemas.AgenticDiscoveryRequest):
    notes = []
    standard = payload.message_standard
    version = payload.message_version
    if standard == "PAPER_PO":
        notes.append("Paper PO detected. Use hybrid OCR + AI extraction and configurable review.")
    elif standard == "EDIFACT":
        notes.append("Use standard/version registry to resolve parser and validation adapter.")
        if not version:
            notes.append("Version not provided. Infer from interchange headers or implementation guideline.")
    elif standard == "X12":
        notes.append("Use transaction-set + version metadata for validation.")
    else:
        notes.append("Use schema-guided mapping for structured format onboarding.")

    return schemas.AgenticDiscoveryResponse(
        message_standard=standard,
        message_version=version,
        recommended_extraction_mode=payload.extraction_mode,
        suggested_mapping_strategy="STANDARD_MODEL_THEN_TARGET_MAPPING",
        notes=notes,
    )
