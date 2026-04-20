from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


ERROR_FILL = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
ERROR_FONT = Font(color="B91C1C", bold=True)


@dataclass
class ValidationIssue:
    row_number: int
    column_name: str
    message: str


@dataclass
class ValidationResult:
    is_valid: bool
    parsed_rows: list[dict[str, Any]] = field(default_factory=list)
    issues: list[ValidationIssue] = field(default_factory=list)
    workbook_bytes: bytes | None = None


def _normalize_headers(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        value = str(cell.value).strip() if cell.value is not None else ""
        if value:
            headers[value] = idx
    return headers


def _ensure_error_column(ws: Worksheet) -> int:
    header_names = [str(c.value).strip() if c.value else "" for c in ws[1]]
    if "error_message" in header_names:
        return header_names.index("error_message") + 1

    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value="error_message")
    ws.cell(row=1, column=new_col).fill = HEADER_FILL
    ws.cell(row=1, column=new_col).font = ERROR_FONT
    return new_col


def _append_error(ws: Worksheet, row_num: int, error_col: int, message: str) -> None:
    cell = ws.cell(row=row_num, column=error_col)
    existing = str(cell.value).strip() if cell.value else ""
    cell.value = f"{existing}; {message}" if existing else message
    cell.fill = ERROR_FILL
    cell.font = ERROR_FONT


def _mark_cell_error(ws: Worksheet, row_num: int, col_num: int) -> None:
    ws.cell(row=row_num, column=col_num).fill = ERROR_FILL


def _workbook_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def validate_address_master_workbook(contents: bytes) -> ValidationResult:
    wb = load_workbook(filename=io.BytesIO(contents))
    if "Address_Master" not in wb.sheetnames:
        return ValidationResult(
            is_valid=False,
            issues=[ValidationIssue(0, "sheet", "Sheet 'Address_Master' not found.")],
            workbook_bytes=_workbook_to_bytes(wb),
        )

    ws = wb["Address_Master"]
    headers = _normalize_headers(ws)
    error_col = _ensure_error_column(ws)

    required = [
        "client_id",
        "partner_id",
        "direction",
        "partner_type",
        "role_code",
        "address_line1",
        "is_active",
    ]

    missing_headers = [h for h in required if h not in headers]
    if missing_headers:
        for header in missing_headers:
            _append_error(ws, 2, error_col, f"Missing header: {header}")
        return ValidationResult(
            is_valid=False,
            issues=[ValidationIssue(0, h, "Missing required header.") for h in missing_headers],
            workbook_bytes=_workbook_to_bytes(wb),
        )

    valid_direction = {"INBOUND", "OUTBOUND"}
    valid_partner_type = {"CUSTOMER", "SUPPLIER", "LOGISTICS_PROVIDER", "WAREHOUSE"}
    valid_role_code = {
        "SHIP_TO",
        "SOLD_TO",
        "BILL_TO",
        "SUPPLIER",
        "WAREHOUSE",
        "DELIVERY_LOCATION",
    }
    valid_active = {"TRUE", "FALSE", "YES", "NO", "1", "0"}

    parsed_rows: list[dict[str, Any]] = []
    issues: list[ValidationIssue] = []

    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        has_any_value = False

        for header, col_num in headers.items():
          value = ws.cell(row=row_num, column=col_num).value
          row_data[header] = value
          if value not in (None, ""):
              has_any_value = True

        if not has_any_value:
            continue

        row_errors: list[tuple[str, str]] = []

        for field in required:
            if row_data.get(field) in (None, ""):
                row_errors.append((field, f"{field} is required."))

        direction = str(row_data.get("direction") or "").strip().upper()
        partner_type = str(row_data.get("partner_type") or "").strip().upper()
        role_code = str(row_data.get("role_code") or "").strip().upper()
        is_active = str(row_data.get("is_active") or "").strip().upper()

        if direction and direction not in valid_direction:
            row_errors.append(("direction", "Invalid direction."))
        if partner_type and partner_type not in valid_partner_type:
            row_errors.append(("partner_type", "Invalid partner_type."))
        if role_code and role_code not in valid_role_code:
            row_errors.append(("role_code", "Invalid role_code."))
        if is_active and is_active not in valid_active:
            row_errors.append(("is_active", "Invalid is_active value."))

        if row_errors:
            for column_name, message in row_errors:
                issues.append(ValidationIssue(row_num, column_name, message))
                if column_name in headers:
                    _mark_cell_error(ws, row_num, headers[column_name])
                _append_error(ws, row_num, error_col, message)
            continue

        parsed_rows.append(
            {
                "client_id": str(row_data.get("client_id")).strip(),
                "partner_id": str(row_data.get("partner_id")).strip(),
                "direction": direction,
                "partner_type": partner_type,
                "role_code": role_code,
                "address_name": str(row_data.get("address_name")).strip() if row_data.get("address_name") else None,
                "address_line1": str(row_data.get("address_line1")).strip(),
                "address_line2": str(row_data.get("address_line2")).strip() if row_data.get("address_line2") else None,
                "city": str(row_data.get("city")).strip() if row_data.get("city") else None,
                "state": str(row_data.get("state")).strip() if row_data.get("state") else None,
                "postal_code": str(row_data.get("postal_code")).strip() if row_data.get("postal_code") else None,
                "country": str(row_data.get("country")).strip() if row_data.get("country") else None,
                "ship_to_code": str(row_data.get("ship_to_code")).strip() if row_data.get("ship_to_code") else None,
                "sold_to_code": str(row_data.get("sold_to_code")).strip() if row_data.get("sold_to_code") else None,
                "bill_to_code": str(row_data.get("bill_to_code")).strip() if row_data.get("bill_to_code") else None,
                "supplier_code": str(row_data.get("supplier_code")).strip() if row_data.get("supplier_code") else None,
                "warehouse_code": str(row_data.get("warehouse_code")).strip() if row_data.get("warehouse_code") else None,
                "delivery_location_code": str(row_data.get("delivery_location_code")).strip() if row_data.get("delivery_location_code") else None,
                "is_active": is_active in {"TRUE", "YES", "1"},
                "notes": str(row_data.get("notes")).strip() if row_data.get("notes") else None,
            }
        )

    is_valid = len(issues) == 0
    return ValidationResult(
        is_valid=is_valid,
        parsed_rows=parsed_rows,
        issues=issues,
        workbook_bytes=None if is_valid else _workbook_to_bytes(wb),
    )