from __future__ import annotations

import io
from dataclasses import dataclass, field
import json
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


ERROR_FILL = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
ERROR_FONT = Font(color="B91C1C", bold=True)
FIELD_REQUIREMENT_COLUMNS = {
    "req_document_number": "document_number",
    "req_document_date": "document_date",
    "req_document_type": "document_type",
    "req_order_type": "order_type",
    "req_customer_name": "customer_name",
    "req_supplier_name": "supplier_name",
    "req_bill_to_code": "bill_to_code",
    "req_bill_to_name": "bill_to_name",
    "req_currency_code": "currency_code",
    "req_ship_to_code": "ship_to_code",
    "req_ship_to_name": "ship_to_name",
    "req_ship_to_address": "ship_to_address",
    "req_header_details": "header_details",
    "req_invoice_number": "invoice_number",
    "req_invoice_date": "invoice_date",
    "req_invoice_due_date": "invoice_due_date",
    "req_invoice_total": "invoice_total",
    "req_item_material_code": "items.*.material_code",
    "req_item_mapped_product": "items.*.mapped_product",
    "req_item_description": "items.*.description",
    "req_item_line_details": "items.*.line_details",
    "req_item_quantity": "items.*.quantity",
    "req_item_customer_uom": "items.*.customer_uom",
    "req_item_delivery_date": "items.*.delivery_date",
    "req_item_delivery_time": "items.*.delivery_time",
    "req_item_unit_price": "items.*.unit_price",
    "req_item_amount": "items.*.amount",
}
VALID_REQUIREMENT_LEVELS = {"MANDATORY", "OPTIONAL", "CONDITIONAL"}
VALID_TARGET_ERPS = {"SAP", "ORACLE", "D365", "JDE", "API"}
VALID_INVOICE_PROFILE_TYPES = {"AP_INVOICE", "AR_INVOICE", "INVOICE"}
VALID_BOOLEAN_VALUES = {"TRUE", "FALSE", "YES", "NO", "1", "0"}


@dataclass
class ValidationIssue:
    row_number: int
    column_name: str
    message: str


@dataclass
class ValidationResult:
    is_valid: bool
    parsed_rows: list[dict[str, Any]] = field(default_factory=list)
    issues: list[ValidationIssue] = field(default_factory=list)
    workbook_bytes: bytes | None = None


def _normalize_headers(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        value = str(cell.value).strip() if cell.value is not None else ""
        if value:
            headers[value] = idx
    return headers


def _ensure_error_column(ws: Worksheet) -> int:
    max_col = ws.max_column
    header_names = [str(c.value).strip() if c.value else "" for c in ws[1]]
    if "error_message" in header_names:
        return header_names.index("error_message") + 1

    new_col = max_col + 1
    ws.cell(row=1, column=new_col, value="error_message")
    ws.cell(row=1, column=new_col).fill = HEADER_FILL
    ws.cell(row=1, column=new_col).font = ERROR_FONT
    return new_col


def _append_error(ws: Worksheet, row_num: int, error_col: int, message: str) -> None:
    cell = ws.cell(row=row_num, column=error_col)
    existing = str(cell.value).strip() if cell.value else ""
    cell.value = f"{existing}; {message}" if existing else message
    cell.fill = ERROR_FILL
    cell.font = ERROR_FONT


def _mark_cell_error(ws: Worksheet, row_num: int, col_num: int) -> None:
    cell = ws.cell(row=row_num, column=col_num)
    cell.fill = ERROR_FILL


def _workbook_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _parse_json_cell(value: Any, field_name: str) -> tuple[dict[str, Any] | None, str | None]:
    if value in (None, ""):
        return None, None
    if isinstance(value, dict):
        return value, None
    raw = str(value).strip()
    if not raw:
        return None, None
    try:
        parsed = json.loads(raw)
    except Exception:
        return None, f"{field_name} must be valid JSON."
    if not isinstance(parsed, dict):
        return None, f"{field_name} must be a JSON object."
    return parsed, None


def _validate_requirement_json(payload: dict[str, Any], field_name: str) -> str | None:
    requirements = payload.get("field_requirements")
    if requirements is None:
        return None
    if not isinstance(requirements, dict):
        return f"{field_name}.field_requirements must be a JSON object."
    return None


def _build_requirement_json_from_row(row_data: dict[str, Any]) -> tuple[dict[str, Any] | None, list[tuple[str, str]]]:
    field_requirements: dict[str, Any] = {}
    errors: list[tuple[str, str]] = []
    for column_name, field_name in FIELD_REQUIREMENT_COLUMNS.items():
        level = str(row_data.get(column_name) or "").strip().upper()
        if not level:
            continue
        if level not in VALID_REQUIREMENT_LEVELS:
            errors.append((column_name, f"{column_name} must be MANDATORY, OPTIONAL, or CONDITIONAL."))
            continue
        field_requirements[field_name] = level
    if not field_requirements:
        return None, errors
    return {"field_requirements": field_requirements}, errors


def validate_uom_workbook(contents: bytes) -> ValidationResult:
    wb = load_workbook(filename=io.BytesIO(contents))
    if "UOM_Data" not in wb.sheetnames:
        return ValidationResult(
            is_valid=False,
            issues=[ValidationIssue(0, "sheet", "Sheet 'UOM_Data' not found.")],
            workbook_bytes=_workbook_to_bytes(wb),
        )

    ws = wb["UOM_Data"]
    headers = _normalize_headers(ws)
    error_col = _ensure_error_column(ws)

    required = [
        "client_id",
        "input_uom",
        "output_uom",
        "factor",
        "rounding_digits",
        "rounding_mode",
        "is_active",
    ]

    missing_headers = [h for h in required if h not in headers]
    if missing_headers:
        for header in missing_headers:
            _append_error(ws, 2, error_col, f"Missing header: {header}")
        return ValidationResult(
            is_valid=False,
            issues=[ValidationIssue(0, h, "Missing required header.") for h in missing_headers],
            workbook_bytes=_workbook_to_bytes(wb),
        )

    valid_rounding = {"HALF_UP", "FLOOR", "CEILING"}
    valid_active = {"TRUE", "FALSE", "YES", "NO", "1", "0"}

    parsed_rows: list[dict[str, Any]] = []
    issues: list[ValidationIssue] = []

    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        has_any_value = False

        for header, col_num in headers.items():
            value = ws.cell(row=row_num, column=col_num).value
            row_data[header] = value
            if value not in (None, ""):
                has_any_value = True

        if not has_any_value:
            continue

        row_errors: list[tuple[str, str]] = []

        for field in required:
            if row_data.get(field) in (None, ""):
                row_errors.append((field, f"{field} is required."))

        factor = row_data.get("factor")
        if factor not in (None, ""):
            try:
                float(factor)
            except Exception:
                row_errors.append(("factor", "factor must be numeric."))

        divider = row_data.get("divider")
        if divider not in (None, ""):
            try:
                float(divider)
            except Exception:
                row_errors.append(("divider", "divider must be numeric."))

        rounding_digits = row_data.get("rounding_digits")
        if rounding_digits not in (None, ""):
            try:
                int(rounding_digits)
            except Exception:
                row_errors.append(("rounding_digits", "rounding_digits must be an integer."))

        rounding_mode = str(row_data.get("rounding_mode") or "").strip().upper()
        if rounding_mode and rounding_mode not in valid_rounding:
            row_errors.append(("rounding_mode", "Invalid rounding_mode."))

        is_active = str(row_data.get("is_active") or "").strip().upper()
        if is_active and is_active not in valid_active:
            row_errors.append(("is_active", "Invalid is_active value."))

        if row_errors:
            for column_name, message in row_errors:
                issues.append(ValidationIssue(row_num, column_name, message))
                if column_name in headers:
                    _mark_cell_error(ws, row_num, headers[column_name])
                _append_error(ws, row_num, error_col, message)
            continue

        parsed_rows.append(
            {
                "client_id": str(row_data.get("client_id")).strip(),
                "partner_id": str(row_data.get("partner_id")).strip() if row_data.get("partner_id") else None,
                "input_uom": str(row_data.get("input_uom")).strip().upper(),
                "output_uom": str(row_data.get("output_uom")).strip().upper(),
                "factor": float(row_data.get("factor")),
                "divider": float(row_data.get("divider")) if row_data.get("divider") not in (None, "") else 1.0,
                "material_code": str(row_data.get("material_code")).strip() if row_data.get("material_code") else None,
                "rounding_digits": int(row_data.get("rounding_digits")),
                "rounding_mode": rounding_mode,
                "is_active": is_active in {"TRUE", "YES", "1"},
            }
        )

    is_valid = len(issues) == 0
    return ValidationResult(
        is_valid=is_valid,
        parsed_rows=parsed_rows,
        issues=issues,
        workbook_bytes=None if is_valid else _workbook_to_bytes(wb),
    )


def validate_bulk_onboarding_workbook(contents: bytes) -> ValidationResult:
    wb = load_workbook(filename=io.BytesIO(contents))
    if "Bulk_Onboarding" not in wb.sheetnames:
        return ValidationResult(
            is_valid=False,
            issues=[ValidationIssue(0, "sheet", "Sheet 'Bulk_Onboarding' not found.")],
            workbook_bytes=_workbook_to_bytes(wb),
        )

    ws = wb["Bulk_Onboarding"]
    headers = _normalize_headers(ws)
    error_col = _ensure_error_column(ws)

    required = [
        "client_id",
        "vertical_id",
        "partner_code",
        "partner_name",
        "partner_type",
        "status",
        "connection_method",
    ]

    missing_headers = [h for h in required if h not in headers]
    if missing_headers:
        for header in missing_headers:
            _append_error(ws, 2, error_col, f"Missing header: {header}")
        return ValidationResult(
            is_valid=False,
            issues=[ValidationIssue(0, h, "Missing required header.") for h in missing_headers],
            workbook_bytes=_workbook_to_bytes(wb),
        )

    valid_partner_type = {"CUSTOMER", "SUPPLIER", "LOGISTICS_PROVIDER"}
    valid_status = {"ACTIVE", "INACTIVE"}
    valid_connection = {"EMAIL", "EDI", "SFTP", "AS2", "API"}

    parsed_rows: list[dict[str, Any]] = []
    issues: list[ValidationIssue] = []

    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        has_any_value = False

        for header, col_num in headers.items():
            value = ws.cell(row=row_num, column=col_num).value
            row_data[header] = value
            if value not in (None, ""):
                has_any_value = True

        if not has_any_value:
            continue

        row_errors: list[tuple[str, str]] = []

        for field in required:
            if row_data.get(field) in (None, ""):
                row_errors.append((field, f"{field} is required."))

        partner_type = str(row_data.get("partner_type") or "").strip().upper()
        if partner_type and partner_type not in valid_partner_type:
            row_errors.append(("partner_type", "Invalid partner_type."))

        status = str(row_data.get("status") or "").strip().upper()
        if status and status not in valid_status:
            row_errors.append(("status", "Invalid status."))

        connection_method = str(row_data.get("connection_method") or "").strip().upper()
        if connection_method and connection_method not in valid_connection:
            row_errors.append(("connection_method", "Invalid connection_method."))

        email = str(row_data.get("email") or "").strip()
        if connection_method == "EMAIL" and not email:
            row_errors.append(("email", "email is required when connection_method is EMAIL."))

        target_erp = str(row_data.get("target_erp") or "").strip().upper()
        if target_erp and target_erp not in VALID_TARGET_ERPS:
            row_errors.append(("target_erp", "Invalid target_erp."))

        invoice_profile_type = str(row_data.get("invoice_profile_type") or "").strip().upper()
        if invoice_profile_type and invoice_profile_type not in VALID_INVOICE_PROFILE_TYPES:
            row_errors.append(("invoice_profile_type", "invoice_profile_type must be AP_INVOICE, AR_INVOICE, or INVOICE."))

        customization_required_raw = str(row_data.get("customization_required") or "").strip().upper()
        if customization_required_raw and customization_required_raw not in VALID_BOOLEAN_VALUES:
            row_errors.append(("customization_required", "customization_required must be TRUE or FALSE."))

        row_level_validation_json, row_level_validation_errors = _build_requirement_json_from_row(row_data)
        row_errors.extend(row_level_validation_errors)

        mapping_validation_json, mapping_validation_error = _parse_json_cell(
            row_data.get("mapping_validation_json"),
            "mapping_validation_json",
        )
        if mapping_validation_error:
            row_errors.append(("mapping_validation_json", mapping_validation_error))
        elif mapping_validation_json:
            validation_error = _validate_requirement_json(
                mapping_validation_json,
                "mapping_validation_json",
            )
            if validation_error:
                row_errors.append(("mapping_validation_json", validation_error))

        business_rule_validation_json, business_rule_validation_error = _parse_json_cell(
            row_data.get("business_rule_validation_json"),
            "business_rule_validation_json",
        )
        if business_rule_validation_error:
            row_errors.append(("business_rule_validation_json", business_rule_validation_error))
        elif business_rule_validation_json:
            validation_error = _validate_requirement_json(
                business_rule_validation_json,
                "business_rule_validation_json",
            )
            if validation_error:
                row_errors.append(("business_rule_validation_json", validation_error))
        elif row_level_validation_json:
            business_rule_validation_json = row_level_validation_json

        if row_errors:
            for column_name, message in row_errors:
                issues.append(ValidationIssue(row_num, column_name, message))
                if column_name in headers:
                    _mark_cell_error(ws, row_num, headers[column_name])
                _append_error(ws, row_num, error_col, message)
            continue

        parsed_rows.append(
            {
                "client_id": str(row_data.get("client_id")).strip(),
                "vertical_id": str(row_data.get("vertical_id")).strip(),
                "partner_code": str(row_data.get("partner_code")).strip().upper(),
                "partner_name": str(row_data.get("partner_name")).strip(),
                "partner_type": partner_type,
                "status": status,
                "connection_method": connection_method,
                "email": email or None,
                "edi_id": str(row_data.get("edi_id")).strip() if row_data.get("edi_id") else None,
                "sftp_path": str(row_data.get("sftp_path")).strip() if row_data.get("sftp_path") else None,
                "as2_id": str(row_data.get("as2_id")).strip() if row_data.get("as2_id") else None,
                "api_reference": str(row_data.get("api_reference")).strip() if row_data.get("api_reference") else None,
                "target_defaults_json": {
                    key: str(row_data.get(key)).strip()
                    for key in ("header_text_id", "line_text_id")
                    if row_data.get(key) not in (None, "")
                },
                "target_profile_json": {
                    key: str(row_data.get(key)).strip()
                    for key in (
                        "target_message_family",
                        "target_erp",
                        "target_standard",
                        "target_message_type",
                        "target_message_version",
                        "transaction_id_source",
                        "invoice_profile_type",
                        "invoice_number_source",
                        "invoice_date_source",
                        "invoice_total_source",
                    )
                    if row_data.get(key) not in (None, "")
                },
                "customization_json": {
                    "required": customization_required_raw in {"TRUE", "YES", "1"},
                    "notes": str(row_data.get("customization_notes")).strip() if row_data.get("customization_notes") else "",
                }
                if customization_required_raw or row_data.get("customization_notes") not in (None, "")
                else {},
                "mapping_validation_json": mapping_validation_json,
                "business_rule_validation_json": business_rule_validation_json,
                "row_level_field_requirements": row_level_validation_json,
                "notes": str(row_data.get("notes")).strip() if row_data.get("notes") else None,
            }
        )

    is_valid = len(issues) == 0
    return ValidationResult(
        is_valid=is_valid,
        parsed_rows=parsed_rows,
        issues=issues,
        workbook_bytes=None if is_valid else _workbook_to_bytes(wb),
    )
