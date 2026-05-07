from __future__ import annotations

import io
import os
from datetime import datetime
from pathlib import Path
from typing import Any
from uuid import UUID
import uuid

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.db import models_trading_partner_agentic as models
from backend.db import schemas_trading_partner_agentic as schemas
from backend.db import models as core_models
from backend.db.models_rules_uom_mapping import (
    TradingPartnerMappingProfile,
    TradingPartnerBusinessRule,
    TradingPartnerUomRule,
)
from backend.services.email_polling_service import email_polling_service
from backend.services.parser_service import parse_file_smart
from backend.services.onboarding_config_service import write_audit


STAGES = [
    "DISCOVER",
    "COLLECT_REQUIREMENTS",
    "ANALYZE_SAMPLE_MESSAGES",
    "DRAFT_CONFIGURATION",
    "VALIDATE_CONFIGURATION",
    "TEST_CONNECTIVITY",
    "TEST_MESSAGE_PROCESSING",
    "REVIEW_AND_APPROVE",
    "ACTIVATE",
]


def _workflow_defaults(payload: schemas.AgenticProjectCreate) -> dict[str, Any]:
    return {
        "current_stage": "DISCOVER",
        "objective": f"Onboard {payload.message_family} / {payload.message_standard} for partner-driven automation.",
        "approval_status": "PENDING",
        "conversation_summary": "",
        "recommended_actions": [
            "Capture partner requirements and message samples.",
            "Draft extraction, mapping, rule, and connectivity configuration.",
            "Validate with test messages before activation.",
        ],
        "requirements_json": {
            "message_family": payload.message_family,
            "message_standard": payload.message_standard,
            "message_version": payload.message_version,
            "direction": payload.direction,
            "target_message_family": payload.target_message_family,
            "invoice_profile_type": payload.invoice_profile_type,
            "extraction_mode": payload.extraction_mode,
            "sample_reference": payload.sample_reference,
            "artifacts": [],
        },
        "test_plan_json": {
            "connectivity": {"status": "PENDING", "notes": []},
            "sample_message_processing": {"status": "PENDING", "notes": []},
        },
        "test_results_json": {},
        "sample_analysis_json": {},
        "progress_steps": [
            {"stage": stage, "status": "CURRENT" if stage == "DISCOVER" else "PENDING"}
            for stage in STAGES
        ],
    }


def _set_progress_steps(current_stage: str, approval_status: str) -> list[dict[str, str]]:
    steps = []
    current_index = STAGES.index(current_stage) if current_stage in STAGES else 0
    for idx, stage in enumerate(STAGES):
        status_text = "PENDING"
        if idx < current_index:
            status_text = "DONE"
        elif idx == current_index:
            status_text = "CURRENT"
        if approval_status == "APPROVED" and stage == "ACTIVATE":
            status_text = "CURRENT"
        steps.append({"stage": stage, "status": status_text})
    return steps


def _merge_workflow(row: models.AgenticOnboardingProject) -> dict[str, Any]:
    discovery = dict(getattr(row, "discovery_json", None) or {})
    discovery.setdefault("current_stage", "DISCOVER")
    discovery.setdefault("objective", "")
    discovery.setdefault("approval_status", "PENDING")
    discovery.setdefault("conversation_summary", "")
    discovery.setdefault("recommended_actions", [])
    discovery.setdefault("requirements_json", {})
    discovery.setdefault("test_plan_json", {})
    discovery.setdefault("test_results_json", {})
    discovery.setdefault("sample_analysis_json", {})
    discovery["progress_steps"] = _set_progress_steps(
        discovery.get("current_stage") or "DISCOVER",
        discovery.get("approval_status") or "PENDING",
    )
    return discovery


def _safe_label(text: Any) -> str:
    return str(text or "").strip()


def _normalize_requirement_level(value: Any) -> str:
    text = _safe_label(value).upper()
    return text if text in {"MANDATORY", "OPTIONAL", "CONDITIONAL"} else "OPTIONAL"


def _worksheet_pairs(sheet, *, key_col: int = 1, value_col: int = 2) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = _safe_label(row[key_col - 1] if len(row) >= key_col else "")
        value = _safe_label(row[value_col - 1] if len(row) >= value_col else "")
        if key:
            result[key] = value
    return result


def _parse_business_workbook(content: bytes, artifact_type: str, file_name: str) -> dict[str, Any] | None:
    lower_name = file_name.lower()
    if not lower_name.endswith(".xlsx"):
        return None

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None

    parsed: dict[str, Any] = {
        "template_kind": artifact_type,
        "template_file_name": file_name,
    }

    if "Business Questionnaire" in wb.sheetnames:
        questionnaire = _worksheet_pairs(wb["Business Questionnaire"])
        field_rules_sheet = wb["Field Rules"] if "Field Rules" in wb.sheetnames else None
        target_output = _worksheet_pairs(wb["Target Output"]) if "Target Output" in wb.sheetnames else {}
        test_scenarios_sheet = wb["Test Scenarios"] if "Test Scenarios" in wb.sheetnames else None
        advanced_sheet = wb["Advanced Mapping"] if "Advanced Mapping" in wb.sheetnames else None

        field_rules: list[dict[str, str]] = []
        if field_rules_sheet is not None:
            for row in field_rules_sheet.iter_rows(min_row=2, values_only=True):
                business_field = _safe_label(row[0] if len(row) >= 1 else "")
                if not business_field:
                    continue
                field_rules.append(
                    {
                        "business_field": business_field,
                        "required_level": _normalize_requirement_level(row[1] if len(row) >= 2 else ""),
                        "meaning": _safe_label(row[2] if len(row) >= 3 else ""),
                        "example_value": _safe_label(row[3] if len(row) >= 4 else ""),
                    }
                )

        test_scenarios: list[dict[str, str]] = []
        if test_scenarios_sheet is not None:
            for row in test_scenarios_sheet.iter_rows(min_row=2, values_only=True):
                scenario = _safe_label(row[0] if len(row) >= 1 else "")
                if not scenario:
                    continue
                test_scenarios.append(
                    {
                        "scenario": scenario,
                        "expected_outcome": _safe_label(row[1] if len(row) >= 2 else ""),
                        "sample_available": _safe_label(row[2] if len(row) >= 3 else ""),
                        "notes": _safe_label(row[3] if len(row) >= 4 else ""),
                    }
                )

        advanced_mapping_rows: list[dict[str, str]] = []
        if advanced_sheet is not None:
            for row in advanced_sheet.iter_rows(min_row=2, values_only=True):
                source_field = _safe_label(row[0] if len(row) >= 1 else "")
                business_meaning = _safe_label(row[1] if len(row) >= 2 else "")
                canonical_field = _safe_label(row[2] if len(row) >= 3 else "")
                target_field = _safe_label(row[3] if len(row) >= 4 else "")
                notes = _safe_label(row[4] if len(row) >= 5 else "")
                if not any([source_field, business_meaning, canonical_field, target_field, notes]):
                    continue
                if source_field == "N/A":
                    continue
                advanced_mapping_rows.append(
                    {
                        "source_field": source_field,
                        "business_meaning": business_meaning,
                        "canonical_field": canonical_field,
                        "target_field": target_field,
                        "notes": notes,
                    }
                )

        parsed.update(
            {
                "questionnaire": questionnaire,
                "field_rules": field_rules,
                "target_output": target_output,
                "test_scenarios": test_scenarios,
                "advanced_mapping_rows": advanced_mapping_rows,
            }
        )
        return parsed

    if "Interface Summary" in wb.sheetnames:
        interface_summary = _worksheet_pairs(wb["Interface Summary"])
        business_checklist_sheet = wb["Business Checklist"] if "Business Checklist" in wb.sheetnames else None
        segment_sheet = wb["Segment Details"] if "Segment Details" in wb.sheetnames else None
        target_output = _worksheet_pairs(wb["Target Output"]) if "Target Output" in wb.sheetnames else {}

        business_checklist: list[dict[str, str]] = []
        if business_checklist_sheet is not None:
            for row in business_checklist_sheet.iter_rows(min_row=2, values_only=True):
                question = _safe_label(row[0] if len(row) >= 1 else "")
                if not question:
                    continue
                business_checklist.append(
                    {
                        "question": question,
                        "answer": _safe_label(row[1] if len(row) >= 2 else ""),
                        "notes": _safe_label(row[2] if len(row) >= 3 else ""),
                    }
                )

        segment_details: list[dict[str, str]] = []
        if segment_sheet is not None:
            for row in segment_sheet.iter_rows(min_row=2, values_only=True):
                segment_field = _safe_label(row[0] if len(row) >= 1 else "")
                if not segment_field:
                    continue
                segment_details.append(
                    {
                        "segment_field": segment_field,
                        "business_meaning": _safe_label(row[1] if len(row) >= 2 else ""),
                        "expected_rule": _safe_label(row[2] if len(row) >= 3 else ""),
                        "mandatory": _safe_label(row[3] if len(row) >= 4 else ""),
                        "example": _safe_label(row[4] if len(row) >= 5 else ""),
                    }
                )

        parsed.update(
            {
                "interface_summary": interface_summary,
                "business_checklist": business_checklist,
                "segment_details": segment_details,
                "target_output": target_output,
            }
        )
        return parsed

    return None


def _hydrate_read(row: models.AgenticOnboardingProject) -> schemas.AgenticProjectRead:
    workflow = _merge_workflow(row)
    return schemas.AgenticProjectRead.model_validate(
        {
            "project_id": row.project_id,
            "client_id": row.client_id,
            "partner_id": row.partner_id,
            "profile_name": row.profile_name,
            "message_family": row.message_family,
            "message_standard": row.message_standard,
            "message_version": row.message_version,
            "direction": row.direction,
            "target_message_family": row.target_message_family,
            "invoice_profile_type": dict(getattr(row, "discovery_json", None) or {}).get("invoice_profile_type")
            or dict(getattr(row, "discovery_json", None) or {}).get("requirements_json", {}).get("invoice_profile_type"),
            "extraction_mode": row.extraction_mode,
            "sample_reference": row.sample_reference,
            "status": row.status,
            "current_stage": workflow.get("current_stage"),
            "objective": workflow.get("objective"),
            "approval_status": workflow.get("approval_status"),
            "conversation_summary": workflow.get("conversation_summary"),
            "recommended_actions": workflow.get("recommended_actions") or [],
            "requirements_json": workflow.get("requirements_json") or {},
            "test_plan_json": workflow.get("test_plan_json") or {},
            "test_results_json": workflow.get("test_results_json") or {},
            "progress_steps": workflow.get("progress_steps") or [],
            "discovery_json": workflow,
            "extraction_profile_json": row.extraction_profile_json or {},
            "address_match_profile_json": row.address_match_profile_json or {},
            "mapping_profile_json": row.mapping_profile_json or {},
            "rule_profile_json": row.rule_profile_json or {},
            "created_at": row.created_at,
            "updated_at": row.updated_at,
        }
    )


class AgenticOnboardingService:
    @staticmethod
    def _build_scenario_coverage_summary(artifacts: list[dict[str, Any]]) -> dict[str, Any]:
        sample_artifacts = [item for item in artifacts if (item or {}).get("artifact_type") == "paper_po_sample"]
        categories = {"happy_path": 0, "edge_case": 0, "exception": 0, "unclassified": 0}
        scenario_labels: list[str] = []
        for item in sample_artifacts:
            category = str((item or {}).get("scenario_category") or "unclassified").strip().lower()
            if category not in categories:
                category = "unclassified"
            categories[category] += 1
            label = str((item or {}).get("scenario_label") or (item or {}).get("file_name") or "").strip()
            if label:
                scenario_labels.append(label)
        coverage_status = "GOOD" if categories["happy_path"] and (categories["edge_case"] or categories["exception"]) else "PARTIAL"
        if not sample_artifacts:
            coverage_status = "MISSING"
        return {
            "status": coverage_status,
            "sample_count": len(sample_artifacts),
            "categories": categories,
            "scenario_labels": scenario_labels[:12],
        }

    @staticmethod
    def _upsert_artifact(
        workflow: dict[str, Any],
        *,
        artifact: dict[str, Any],
    ) -> list[dict[str, Any]]:
        requirements = dict(workflow.get("requirements_json") or {})
        existing = list(requirements.get("artifacts") or [])
        artifact_id = str(artifact.get("file_id") or "")
        merged = [item for item in existing if str((item or {}).get("file_id") or "") != artifact_id]
        merged.append(artifact)
        requirements["artifacts"] = merged
        requirements["artifact_summary"] = {
            "total": len(merged),
            "mapping_spec_count": len([item for item in merged if (item or {}).get("artifact_type") == "mapping_spec"]),
            "edi_guideline_count": len([item for item in merged if (item or {}).get("artifact_type") == "edi_guideline"]),
            "paper_po_sample_count": len([item for item in merged if (item or {}).get("artifact_type") == "paper_po_sample"]),
        }
        requirements["scenario_coverage_summary"] = AgenticOnboardingService._build_scenario_coverage_summary(merged)
        workflow["requirements_json"] = requirements
        return merged

    @staticmethod
    def _classify_sample_scenario(
        artifact_record: dict[str, Any],
        workflow: dict[str, Any],
    ) -> dict[str, Any]:
        analysis = dict(workflow.get("sample_analysis_json") or {})
        header = dict(analysis.get("header") or {})
        label_source = " ".join(
            [
                str(artifact_record.get("scenario_label") or ""),
                str(artifact_record.get("file_name") or ""),
            ]
        ).lower()
        item_count = int(analysis.get("item_count") or 0)
        category = "happy_path"
        reason = "Representative standard sample."

        exception_keywords = ["exception", "error", "issue", "fail", "rejection", "invalid", "manual"]
        edge_keywords = ["edge", "split", "partial", "rush", "special", "multi", "mixed", "complex", "blank", "missing"]

        missing_core_fields = [
            field_name
            for field_name in ("po_number", "document_number", "customer", "buyer", "supplier", "vendor")
            if not header.get(field_name)
        ]

        if any(keyword in label_source for keyword in exception_keywords) or item_count == 0 or len(missing_core_fields) >= 3:
            category = "exception"
            reason = "Sample appears to represent a failure or incomplete extraction path."
        elif any(keyword in label_source for keyword in edge_keywords) or item_count >= 5:
            category = "edge_case"
            reason = "Sample appears to represent a non-standard or higher-complexity scenario."

        return {
            **artifact_record,
            "scenario_category": category,
            "classification_reason": reason,
            "item_count": item_count,
            "detected_document_number": header.get("po_number") or header.get("document_number"),
            "detected_supplier": header.get("supplier") or header.get("vendor"),
            "detected_customer": header.get("customer") or header.get("buyer"),
        }

    @staticmethod
    def _apply_business_template_to_workflow(
        workflow: dict[str, Any],
        *,
        artifact_type: str,
        template_data: dict[str, Any],
        file_name: str,
    ) -> None:
        requirements = dict(workflow.get("requirements_json") or {})

        if artifact_type == "mapping_spec":
            questionnaire = dict(template_data.get("questionnaire") or {})
            field_rules = list(template_data.get("field_rules") or [])
            target_output = dict(template_data.get("target_output") or {})
            test_scenarios = list(template_data.get("test_scenarios") or [])
            advanced_mapping_rows = list(template_data.get("advanced_mapping_rows") or [])
            message_control = dict(template_data.get("message_control") or {})

            requirements["business_questionnaire"] = questionnaire
            requirements["business_field_rules"] = field_rules
            requirements["business_target_output"] = target_output
            requirements["business_test_scenarios"] = test_scenarios
            requirements["advanced_mapping_rows"] = advanced_mapping_rows
            requirements["message_control"] = message_control
            requirements["message_control"] = dict(template_data.get("message_control") or {})

            if questionnaire.get("What kind of document will they send first?"):
                requirements["business_document_type"] = questionnaire.get("What kind of document will they send first?")
            if questionnaire.get("What format will they send?"):
                requirements["business_source_format"] = questionnaire.get("What format will they send?")
            if questionnaire.get("Which ERP should the output go to?"):
                requirements["business_target_erp"] = questionnaire.get("Which ERP should the output go to?")

            if target_output:
                requirements["target_profile"] = {
                    "target_erp": target_output.get("Target ERP"),
                    "target_standard": target_output.get("Target Standard"),
                    "target_message_type": target_output.get("Target Message Type"),
                    "target_message_version": target_output.get("Target Version"),
                    "transaction_id_source": target_output.get("Transaction ID Source"),
                    "invoice_profile_type": target_output.get("Invoice Profile Type"),
                    "invoice_number_source": target_output.get("Invoice Number Source"),
                    "invoice_date_source": target_output.get("Invoice Date Source"),
                    "invoice_total_source": target_output.get("Invoice Total Source"),
                    "header_text_id": target_output.get("Header Text ID"),
                    "line_text_id": target_output.get("Line Text ID"),
                }

            workflow["recommended_actions"] = [
                "Review the business questionnaire answers for completeness.",
                "Confirm field rules and target output expectations from the uploaded workbook.",
                "Upload representative source documents so the AI can draft the configuration from real samples.",
            ]
            workflow["conversation_summary"] = (
                f"{(workflow.get('conversation_summary') or '').strip()}\n"
                f"Business onboarding workbook parsed from {file_name}."
            ).strip()

        elif artifact_type == "edi_guideline":
            interface_summary = dict(template_data.get("interface_summary") or {})
            business_checklist = list(template_data.get("business_checklist") or [])
            segment_details = list(template_data.get("segment_details") or [])
            target_output = dict(template_data.get("target_output") or {})
            message_control = dict(template_data.get("message_control") or {})

            requirements["edi_interface_summary"] = interface_summary
            requirements["edi_business_checklist"] = business_checklist
            requirements["edi_segment_details"] = segment_details
            requirements["edi_target_output"] = target_output
            requirements["message_control"] = message_control
            requirements["message_control"] = dict(template_data.get("message_control") or {})

            if interface_summary.get("Message Standard"):
                requirements["business_source_standard"] = interface_summary.get("Message Standard")
            if interface_summary.get("Message Type"):
                requirements["business_message_type"] = interface_summary.get("Message Type")
            if interface_summary.get("Version"):
                requirements["business_message_version"] = interface_summary.get("Version")

            if target_output:
                requirements["target_profile"] = {
                    "target_erp": target_output.get("Target ERP"),
                    "target_standard": target_output.get("Target Standard"),
                    "target_message_type": target_output.get("Target Message Type"),
                    "target_message_version": target_output.get("Target Version"),
                    "invoice_profile_type": target_output.get("Invoice Profile Type"),
                    "invoice_number_source": target_output.get("Invoice Number Source"),
                    "invoice_date_source": target_output.get("Invoice Date Source"),
                    "invoice_total_source": target_output.get("Invoice Total Source"),
                    "header_text_id": target_output.get("Header Text ID"),
                    "line_text_id": target_output.get("Line Text ID"),
                }

            workflow["recommended_actions"] = [
                "Review the interface summary and business checklist from the uploaded workbook.",
                "Confirm the transaction ID logic and target ERP expectations.",
                "Upload one or more representative messages so the AI can connect the guideline to real samples.",
            ]
            workflow["conversation_summary"] = (
                f"{(workflow.get('conversation_summary') or '').strip()}\n"
                f"EDI/interface workbook parsed from {file_name}."
            ).strip()

        workflow["requirements_json"] = requirements

    @staticmethod
    def _json_safe(value: Any) -> Any:
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        if isinstance(value, UUID):
            return str(value)
        if isinstance(value, datetime):
            return value.isoformat()
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                pass
        if isinstance(value, dict):
            return {str(k): AgenticOnboardingService._json_safe(v) for k, v in value.items()}
        if isinstance(value, (list, tuple, set)):
            return [AgenticOnboardingService._json_safe(v) for v in value]
        return str(value)

    def _resolve_sample_source(self, db: Session, row: models.AgenticOnboardingProject) -> tuple[str, bytes]:
        reference = str(getattr(row, "sample_reference", "") or "").strip()
        if not reference:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Sample reference is required for sample analysis.",
            )

        if os.path.exists(reference):
            with open(reference, "rb") as fh:
                return os.path.basename(reference), fh.read()

        file_row = None
        try:
            ref_uuid = uuid.UUID(reference)
            file_row = (
                db.query(core_models.FileStore)
                .filter(core_models.FileStore.file_id == ref_uuid)
                .first()
            )
        except Exception:
            file_row = None

        if not file_row:
            file_row = (
                db.query(core_models.FileStore)
                .filter(
                    core_models.FileStore.client_id == row.client_id,
                    core_models.FileStore.original_file_name == reference,
                )
                .order_by(core_models.FileStore.uploaded_at.desc())
                .first()
            )

        if not file_row:
            file_row = (
                db.query(core_models.FileStore)
                .filter(
                    core_models.FileStore.client_id == row.client_id,
                    core_models.FileStore.original_file_name.ilike(f"%{reference}%"),
                )
                .order_by(core_models.FileStore.uploaded_at.desc())
                .first()
            )

        if file_row and file_row.file_path and os.path.exists(file_row.file_path):
            with open(file_row.file_path, "rb") as fh:
                return file_row.original_file_name or os.path.basename(file_row.file_path), fh.read()

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sample reference '{reference}' could not be resolved to a file.",
        )

    def _run_sample_analysis(
        self,
        db: Session,
        row: models.AgenticOnboardingProject,
        workflow: dict[str, Any],
    ) -> None:
        file_name, file_bytes = self._resolve_sample_source(db, row)
        upload_like = io.BytesIO(file_bytes)
        upload_like.name = file_name

        header, items_df, vendor = parse_file_smart(upload_like)
        items = items_df.fillna("").to_dict(orient="records")
        header = self._json_safe(dict(header or {}))
        item_preview = self._json_safe(items[:5])
        parser_meta = self._json_safe((header or {}).get("parser_meta") or {})

        workflow["sample_analysis_json"] = {
            "sample_reference": row.sample_reference,
            "resolved_file_name": file_name,
            "detected_vendor": vendor,
            "header": header,
            "item_preview": item_preview,
            "item_count": len(items),
            "parser_meta": parser_meta,
            "analyzed_at": datetime.utcnow().isoformat(),
        }

        requirements = dict(workflow.get("requirements_json") or {})
        requirements.update(
            {
                "resolved_sample_file_name": file_name,
                "detected_vendor": vendor,
                "detected_document_number": header.get("po_number") or header.get("document_number"),
                "detected_currency": header.get("currency"),
                "detected_supplier": header.get("supplier") or header.get("vendor"),
                "detected_customer": header.get("customer") or header.get("buyer"),
                "detected_item_count": len(items),
                "parser_source_used": parser_meta.get("source_used"),
                "artifact_summary": dict((workflow.get("requirements_json") or {}).get("artifact_summary") or {}),
            }
        )
        workflow["requirements_json"] = requirements
        workflow["recommended_actions"] = [
            "Review detected header fields and party mapping.",
            "Confirm sample line-item structure and quantity/UOM interpretation.",
            "Generate draft connection, mapping, and rules from the analyzed sample.",
        ]

        row.extraction_profile_json = {
            **dict(row.extraction_profile_json or {}),
            "status": "ANALYZED",
            "mode": row.extraction_mode,
            "strategy": "HYBRID",
            "resolved_sample_file_name": file_name,
            "detected_vendor": vendor,
            "parser_meta": parser_meta,
        }

    def _auto_fill_draft_configuration(
        self,
        db: Session,
        row: models.AgenticOnboardingProject,
        workflow: dict[str, Any],
    ) -> None:
        if not workflow.get("sample_analysis_json") and row.sample_reference:
            self._run_sample_analysis(db, row, workflow)

        analysis = dict(workflow.get("sample_analysis_json") or {})
        header = dict(analysis.get("header") or {})
        item_preview = list(analysis.get("item_preview") or [])
        parser_meta = dict(analysis.get("parser_meta") or {})
        artifacts = list((workflow.get("requirements_json") or {}).get("artifacts") or [])
        guidance_artifacts = [
            {
                "artifact_type": item.get("artifact_type"),
                "file_name": item.get("file_name"),
                "scenario_label": item.get("scenario_label"),
                "file_id": item.get("file_id"),
            }
            for item in artifacts
            if item.get("artifact_type") in {"mapping_spec", "edi_guideline"}
        ]
        sample_artifacts = [
            {
                "file_name": item.get("file_name"),
                "scenario_label": item.get("scenario_label"),
                "file_id": item.get("file_id"),
            }
            for item in artifacts
            if item.get("artifact_type") == "paper_po_sample"
        ]

        detected_header_fields = [
            key for key, value in header.items()
            if value not in (None, "", [], {}, False)
        ]
        detected_item_fields = sorted(
            {
                key
                for item in item_preview
                for key, value in dict(item or {}).items()
                if value not in (None, "", [], {}, False)
            }
        )

        active_connections = (
            db.query(core_models.TradingPartnerConnection)
            .filter(
                core_models.TradingPartnerConnection.partner_id == row.partner_id,
                core_models.TradingPartnerConnection.is_active == True,
            )
            .all()
        )

        row.extraction_profile_json = {
            **dict(row.extraction_profile_json or {}),
            "status": "READY_FOR_REVIEW",
            "mode": row.extraction_mode,
            "strategy": "HYBRID",
            "sample_reference": row.sample_reference,
            "detected_header_fields": detected_header_fields,
            "detected_item_fields": detected_item_fields,
            "parser_source_used": parser_meta.get("source_used"),
            "layout_signature": parser_meta.get("layout_signature"),
        }

        row.address_match_profile_json = {
            "status": "DRAFT",
            "strategy": "ADDRESS_MASTER_THEN_FALLBACK",
            "detected_ship_to_name": header.get("ship_to") or header.get("ship_to_name"),
            "detected_ship_to_address": header.get("ship_to_address"),
        }

        row.mapping_profile_json = {
            "status": "DRAFT",
            "strategy": "STANDARD_MODEL_THEN_TARGET_MAPPING",
            "target_message_family": row.target_message_family,
            "detected_header_fields": detected_header_fields,
            "detected_item_fields": detected_item_fields,
            "sample_document_number": header.get("po_number") or header.get("document_number"),
            "sample_currency": header.get("currency"),
            "item_preview": item_preview,
            "source_artifacts": guidance_artifacts,
            "scenario_samples": sample_artifacts,
            "scenario_coverage_summary": dict((workflow.get("requirements_json") or {}).get("scenario_coverage_summary") or {}),
        }

        row.rule_profile_json = {
            "status": "DRAFT",
            "business_rules": [],
            "uom_rules": [],
            "suggested_validations": [
                "Validate party mapping before activation.",
                "Validate quantity, UOM, and requested delivery date extraction.",
                "Confirm duplicate-check and split-rule strategy for this partner.",
            ],
        }

        workflow["recommended_actions"] = [
            "Review the generated draft map against the uploaded guideline/specification artifacts.",
            "Confirm active partner connections for real connectivity testing.",
            "Run sample-message processing validation across the uploaded scenario samples before approval.",
        ]
        workflow["test_plan_json"] = {
            "connectivity": {
                "status": "PENDING",
                "notes": [
                    f"{len(active_connections)} active trading partner connection(s) available for testing."
                ],
            },
            "sample_message_processing": {
                "status": "PENDING",
                "notes": [
                    f"Sample '{analysis.get('resolved_file_name') or row.sample_reference or '-'}' ready for parser validation."
                ],
            },
        }

    def _run_connectivity_tests(
        self,
        db: Session,
        row: models.AgenticOnboardingProject,
        workflow: dict[str, Any],
    ) -> None:
        connections = (
            db.query(core_models.TradingPartnerConnection)
            .filter(
                core_models.TradingPartnerConnection.partner_id == row.partner_id,
                core_models.TradingPartnerConnection.is_active == True,
            )
            .all()
        )

        results: list[dict[str, Any]] = []
        overall_status = "PASSED"

        if not connections:
            overall_status = "BLOCKED"
            results.append(
                {
                    "connection_name": None,
                    "status": "BLOCKED",
                    "message": "No active connections configured for this partner.",
                }
            )

        for connection in connections:
            if str(connection.connection_type or "").upper() == "EMAIL":
                try:
                    normalized = email_polling_service.normalize_email_config(
                        client_id=connection.client_id,
                        connection_key=connection.connection_id,
                        config=connection.config_json,
                    )
                    validation = email_polling_service.validate_connection(normalized)
                    results.append(
                        {
                            "connection_id": str(connection.connection_id),
                            "connection_name": connection.connection_name,
                            "connection_type": connection.connection_type,
                            "status": "PASSED",
                            "message": validation.get("message"),
                        }
                    )
                except Exception as exc:
                    overall_status = "FAILED"
                    results.append(
                        {
                            "connection_id": str(connection.connection_id),
                            "connection_name": connection.connection_name,
                            "connection_type": connection.connection_type,
                            "status": "FAILED",
                            "message": str(exc),
                        }
                    )
            else:
                if overall_status != "FAILED":
                    overall_status = "PARTIAL"
                results.append(
                    {
                        "connection_id": str(connection.connection_id),
                        "connection_name": connection.connection_name,
                        "connection_type": connection.connection_type,
                        "status": "SKIPPED",
                        "message": f"Real connectivity test not implemented yet for {connection.connection_type}.",
                    }
                )

        test_plan = dict(workflow.get("test_plan_json") or {})
        test_plan["connectivity"] = {
            "status": overall_status,
            "notes": [result.get("message", "") for result in results],
        }
        workflow["test_plan_json"] = test_plan

        test_results = dict(workflow.get("test_results_json") or {})
        test_results["connectivity"] = {
            "status": overall_status,
            "results": self._json_safe(results),
            "executed_at": datetime.utcnow().isoformat(),
        }
        workflow["test_results_json"] = test_results

    def _run_message_processing_tests(
        self,
        db: Session,
        row: models.AgenticOnboardingProject,
        workflow: dict[str, Any],
    ) -> None:
        file_name, file_bytes = self._resolve_sample_source(db, row)
        upload_like = io.BytesIO(file_bytes)
        upload_like.name = file_name
        header, items_df, vendor = parse_file_smart(upload_like)
        items = items_df.fillna("").to_dict(orient="records")
        header = dict(header or {})

        status_text = "PASSED" if items or header else "FAILED"

        test_plan = dict(workflow.get("test_plan_json") or {})
        test_plan["sample_message_processing"] = {
            "status": status_text,
            "notes": [
                f"Sample '{file_name}' parsed via {(((header.get('parser_meta') or {}).get('source_used')) or 'UNKNOWN')} source.",
                f"Detected {len(items)} line item(s).",
            ],
        }
        workflow["test_plan_json"] = test_plan

        test_results = dict(workflow.get("test_results_json") or {})
        test_results["sample_message_processing"] = {
            "status": status_text,
            "resolved_file_name": file_name,
            "detected_vendor": vendor,
            "header_preview": self._json_safe(
                {
                    key: value
                    for key, value in header.items()
                    if key in {"po_number", "po_date", "customer", "buyer", "supplier", "currency", "vendor", "source_used", "parser_meta"}
                }
            ),
            "item_preview": self._json_safe(items[:5]),
            "item_count": len(items),
            "executed_at": datetime.utcnow().isoformat(),
        }
        workflow["test_results_json"] = test_results

    @staticmethod
    def _model_to_dict(model_obj: Any) -> dict[str, Any]:
        return {
            column.name: AgenticOnboardingService._json_safe(getattr(model_obj, column.name))
            for column in model_obj.__table__.columns
        }

    def _activate_project(
        self,
        db: Session,
        row: models.AgenticOnboardingProject,
        workflow: dict[str, Any],
    ) -> None:
        if not workflow.get("sample_analysis_json") and row.sample_reference:
            self._run_sample_analysis(db, row, workflow)
        self._auto_fill_draft_configuration(db, row, workflow)

        analysis = dict(workflow.get("sample_analysis_json") or {})
        header = dict(analysis.get("header") or {})
        parser_meta = dict(analysis.get("parser_meta") or {})
        item_preview = list(analysis.get("item_preview") or [])
        actor_email = row.created_by or "agentic_onboarding"

        partner_profile = (
            db.query(core_models.TradingPartnerProfile)
            .filter(core_models.TradingPartnerProfile.partner_id == row.partner_id)
            .first()
        )
        profile_before = self._model_to_dict(partner_profile) if partner_profile else None

        if not partner_profile:
            partner_profile = core_models.TradingPartnerProfile(
                client_id=row.client_id,
                partner_id=row.partner_id,
            )
            db.add(partner_profile)
            db.flush()

        partner_profile.profile_name = row.profile_name or partner_profile.profile_name or "Default Profile"
        partner_profile.profile_status = "ACTIVE"
        partner_profile.duplicate_check_enabled = True
        partner_profile.duplicate_check_scope = "PO_NUMBER"
        partner_profile.split_rule = "NONE"
        partner_profile.split_po_number_strategy = "SAME_PO_NUMBER"
        partner_profile.split_po_separator = "-"
        partner_profile.delivery_date_source = "PO_DELIVERY_DATE"
        partner_profile.delivery_date_offset_type = "NONE"
        partner_profile.delivery_date_offset_days = 0
        partner_profile.po_date_source = "PO_DATE"
        db.add(partner_profile)
        db.flush()

        profile_after = self._model_to_dict(partner_profile)
        write_audit(
            db,
            client_id=row.client_id,
            partner_id=str(row.partner_id),
            entity_type="PROFILE",
            entity_id=str(partner_profile.onboarding_profile_id),
            action="UPDATE" if profile_before else "CREATE",
            before_json=profile_before,
            after_json=profile_after,
            actor_email=actor_email,
            actor_role="AGENTIC_ONBOARDING",
            remarks="Activated onboarding profile from AI onboarding project.",
        )

        input_format = "PDF"
        resolved_file_name = str(analysis.get("resolved_file_name") or row.sample_reference or "")
        lowered = resolved_file_name.lower()
        if lowered.endswith(".xlsx") or lowered.endswith(".xls"):
            input_format = "EXCEL"
        elif lowered.endswith(".csv"):
            input_format = "CSV"
        elif lowered.endswith(".xml"):
            input_format = "XML"
        elif lowered.endswith(".json"):
            input_format = "JSON"

        mapping_profile = (
            db.query(TradingPartnerMappingProfile)
            .filter(
                TradingPartnerMappingProfile.partner_id == row.partner_id,
                TradingPartnerMappingProfile.profile_name == row.profile_name,
            )
            .order_by(TradingPartnerMappingProfile.version_no.desc())
            .first()
        )
        mapping_before = self._model_to_dict(mapping_profile) if mapping_profile else None

        if not mapping_profile:
            mapping_profile = TradingPartnerMappingProfile(
                client_id=row.client_id,
                partner_id=row.partner_id,
                profile_name=row.profile_name,
            )
            db.add(mapping_profile)
            db.flush()

        mapping_profile.document_type = "PO"
        mapping_profile.input_format = input_format
        mapping_profile.source_channel = row.direction
        mapping_profile.sold_to = header.get("customer") or header.get("buyer")
        mapping_profile.ship_to = header.get("ship_to") or header.get("ship_to_name")
        mapping_profile.field_mapping_json = {
            "target_message_family": row.target_message_family,
            "detected_header_fields": list(row.extraction_profile_json.get("detected_header_fields") or []),
            "sample_header_preview": {
                key: value
                for key, value in header.items()
                if key in {"po_number", "po_date", "customer", "buyer", "supplier", "currency", "ship_to", "ship_to_name"}
            },
        }
        mapping_profile.header_defaults_json = {
            "customer_name": header.get("customer") or header.get("buyer"),
            "supplier_name": header.get("supplier") or header.get("vendor"),
            "currency": header.get("currency"),
            "ship_to_name": header.get("ship_to_name") or header.get("ship_to"),
            "ship_to_address": header.get("ship_to_address"),
            "header_text_id": "0001",
            "line_text_id": "0001",
        }
        mapping_profile.line_mapping_json = {
            "detected_item_fields": list(row.extraction_profile_json.get("detected_item_fields") or []),
            "item_preview": item_preview,
        }
        mapping_profile.validation_json = {
            "recommended_validations": list((row.rule_profile_json or {}).get("suggested_validations") or []),
            "sample_item_count": analysis.get("item_count"),
        }
        mapping_profile.layout_hint_json = {
            "source_used": parser_meta.get("source_used"),
            "layout_signature": parser_meta.get("layout_signature"),
            "layout_page": parser_meta.get("layout_page"),
            "target_profile": {
                "target_message_family": row.target_message_family,
                "transaction_id_source": (
                    "delivery_number"
                    if str(row.target_message_family or "").strip().upper() in {"ASN", "DESADV", "DELIVERY", "SHIPMENT"}
                    else "billing_document_number"
                    if str(row.target_message_family or "").strip().upper() in {"INVOICE", "BILLING"}
                    else "document_number"
                ),
            },
            "customization": {
                "required": False,
                "notes": "",
            },
        }
        mapping_profile.is_default = True
        mapping_profile.is_active = True
        mapping_profile.notes = "Generated from agentic onboarding activation."
        mapping_profile.created_by = mapping_profile.created_by or actor_email
        mapping_profile.updated_by = actor_email
        db.add(mapping_profile)
        db.flush()

        mapping_after = self._model_to_dict(mapping_profile)
        write_audit(
            db,
            client_id=row.client_id,
            partner_id=str(row.partner_id),
            entity_type="MAPPING",
            entity_id=str(mapping_profile.mapping_profile_id),
            action="UPDATE" if mapping_before else "CREATE",
            before_json=mapping_before,
            after_json=mapping_after,
            actor_email=actor_email,
            actor_role="AGENTIC_ONBOARDING",
            remarks="Activated mapping profile from AI onboarding project.",
        )

        business_rules = list((row.rule_profile_json or {}).get("business_rules") or [])
        for idx, rule in enumerate(business_rules, start=1):
            rule_name = str((rule or {}).get("rule_name") or f"{row.profile_name} Rule {idx}").strip()
            if not rule_name:
                continue
            existing_rule = (
                db.query(TradingPartnerBusinessRule)
                .filter(
                    TradingPartnerBusinessRule.partner_id == row.partner_id,
                    TradingPartnerBusinessRule.rule_name == rule_name,
                )
                .first()
            )
            rule_before = self._model_to_dict(existing_rule) if existing_rule else None
            if not existing_rule:
                existing_rule = TradingPartnerBusinessRule(
                    client_id=row.client_id,
                    partner_id=row.partner_id,
                    rule_name=rule_name,
                )
                db.add(existing_rule)
                db.flush()
            existing_rule.rule_type = str((rule or {}).get("rule_type") or "TRANSFORMATION")
            existing_rule.document_type = str((rule or {}).get("document_type") or "PO")
            existing_rule.message_direction = str((rule or {}).get("message_direction") or row.direction or "INBOUND")
            existing_rule.condition_json = dict((rule or {}).get("condition_json") or {})
            existing_rule.action_json = dict((rule or {}).get("action_json") or {})
            existing_rule.priority = int((rule or {}).get("priority") or 100)
            existing_rule.stop_on_match = bool((rule or {}).get("stop_on_match") or False)
            existing_rule.is_active = bool((rule or {}).get("is_active", True))
            existing_rule.notes = (rule or {}).get("notes")
            existing_rule.created_by = existing_rule.created_by or actor_email
            existing_rule.updated_by = actor_email
            db.add(existing_rule)
            db.flush()
            write_audit(
                db,
                client_id=row.client_id,
                partner_id=str(row.partner_id),
                entity_type="RULE",
                entity_id=str(existing_rule.rule_id),
                action="UPDATE" if rule_before else "CREATE",
                before_json=rule_before,
                after_json=self._model_to_dict(existing_rule),
                actor_email=actor_email,
                actor_role="AGENTIC_ONBOARDING",
                remarks="Activated business rule from AI onboarding project.",
            )

        uom_rules = list((row.rule_profile_json or {}).get("uom_rules") or [])
        for idx, uom in enumerate(uom_rules, start=1):
            input_uom = str((uom or {}).get("input_uom") or "").strip()
            output_uom = str((uom or {}).get("output_uom") or "").strip()
            if not input_uom or not output_uom:
                continue
            existing_uom = (
                db.query(TradingPartnerUomRule)
                .filter(
                    TradingPartnerUomRule.partner_id == row.partner_id,
                    TradingPartnerUomRule.input_uom == input_uom,
                    TradingPartnerUomRule.output_uom == output_uom,
                )
                .first()
            )
            uom_before = self._model_to_dict(existing_uom) if existing_uom else None
            if not existing_uom:
                existing_uom = TradingPartnerUomRule(
                    client_id=row.client_id,
                    partner_id=row.partner_id,
                    input_uom=input_uom,
                    output_uom=output_uom,
                )
                db.add(existing_uom)
                db.flush()
            existing_uom.conversion_factor = (uom or {}).get("conversion_factor")
            existing_uom.conversion_divider = (uom or {}).get("conversion_divider")
            existing_uom.rounding_digits = int((uom or {}).get("rounding_digits") or 2)
            existing_uom.rounding_mode = str((uom or {}).get("rounding_mode") or "HALF_UP")
            existing_uom.priority = int((uom or {}).get("priority") or 100)
            existing_uom.is_active = bool((uom or {}).get("is_active", True))
            existing_uom.notes = (uom or {}).get("notes")
            existing_uom.created_by = existing_uom.created_by or actor_email
            existing_uom.updated_by = actor_email
            db.add(existing_uom)
            db.flush()
            write_audit(
                db,
                client_id=row.client_id,
                partner_id=str(row.partner_id),
                entity_type="UOM",
                entity_id=str(existing_uom.uom_rule_id),
                action="UPDATE" if uom_before else "CREATE",
                before_json=uom_before,
                after_json=self._model_to_dict(existing_uom),
                actor_email=actor_email,
                actor_role="AGENTIC_ONBOARDING",
                remarks="Activated UOM rule from AI onboarding project.",
            )

        row.status = "ACTIVE"
        workflow["approval_status"] = workflow.get("approval_status") or "APPROVED"
        test_results = dict(workflow.get("test_results_json") or {})
        test_results["activation"] = {
            "status": "COMPLETED",
            "onboarding_profile_id": str(partner_profile.onboarding_profile_id),
            "mapping_profile_id": str(mapping_profile.mapping_profile_id),
            "business_rule_count": len(business_rules),
            "uom_rule_count": len(uom_rules),
            "executed_at": datetime.utcnow().isoformat(),
        }
        workflow["test_results_json"] = test_results
        workflow["recommended_actions"] = [
            "Review the activated profile and mapping profile in the trading partner workspace.",
            "Run one end-to-end partner test message to confirm production readiness.",
        ]

    def list_projects(self, db: Session, partner_id: UUID) -> list[schemas.AgenticProjectRead]:
        rows = (
            db.query(models.AgenticOnboardingProject)
            .filter(models.AgenticOnboardingProject.partner_id == partner_id)
            .order_by(models.AgenticOnboardingProject.updated_at.desc())
            .all()
        )
        return [_hydrate_read(row) for row in rows]

    def get_project(self, db: Session, project_id: UUID) -> schemas.AgenticProjectRead:
        row = (
            db.query(models.AgenticOnboardingProject)
            .filter(models.AgenticOnboardingProject.project_id == project_id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agentic project not found.")
        return _hydrate_read(row)

    def create_project(self, db: Session, payload: schemas.AgenticProjectCreate) -> schemas.AgenticProjectRead:
        payload_data = payload.model_dump()
        payload_data.pop("invoice_profile_type", None)
        row = models.AgenticOnboardingProject(
            **payload_data,
            discovery_json=_workflow_defaults(payload),
            extraction_profile_json={
                "mode": payload.extraction_mode,
                "strategy": "HYBRID",
                "status": "DRAFT",
            },
            address_match_profile_json={
                "status": "DRAFT",
                "strategy": "ADDRESS_MASTER_THEN_FALLBACK",
            },
            mapping_profile_json={
                "status": "DRAFT",
                "strategy": "STANDARD_MODEL_THEN_TARGET_MAPPING",
            },
            rule_profile_json={
                "status": "DRAFT",
                "business_rules": [],
                "uom_rules": [],
            },
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return _hydrate_read(row)

    def attach_sample_file(
        self,
        db: Session,
        project_id: UUID,
        *,
        file_name: str,
        content: bytes,
        mime_type: str | None,
        uploaded_by: str | None = None,
    ) -> schemas.AgenticProjectRead:
        row = (
            db.query(models.AgenticOnboardingProject)
            .filter(models.AgenticOnboardingProject.project_id == project_id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agentic project not found.")

        storage_dir = Path("uploads") / row.client_id / "agentic_samples"
        storage_dir.mkdir(parents=True, exist_ok=True)
        safe_name = Path(file_name).name or f"{project_id}.bin"
        storage_path = storage_dir / f"{project_id}_{safe_name}"
        storage_path.write_bytes(content)

        file_row = core_models.FileStore(
            client_id=row.client_id,
            original_file_name=safe_name,
            mime_type=mime_type,
            source_channel="AI_ONBOARDING",
            file_path=str(storage_path),
            file_size_bytes=len(content),
            uploaded_by=uploaded_by or "agentic_onboarding",
        )
        db.add(file_row)
        db.flush()

        row.sample_reference = str(file_row.file_id)
        workflow = _merge_workflow(row)
        requirements = dict(workflow.get("requirements_json") or {})
        requirements["sample_reference"] = row.sample_reference
        requirements["uploaded_sample_file_name"] = safe_name
        workflow["requirements_json"] = requirements
        workflow["conversation_summary"] = (
            f"{(workflow.get('conversation_summary') or '').strip()}\nSample uploaded: {safe_name}"
        ).strip()
        row.discovery_json = workflow
        db.add(row)
        db.commit()
        db.refresh(row)
        return _hydrate_read(row)

    def attach_project_artifact(
        self,
        db: Session,
        project_id: UUID,
        *,
        artifact_type: str,
        scenario_label: str | None,
        file_name: str,
        content: bytes,
        mime_type: str | None,
        uploaded_by: str | None = None,
    ) -> schemas.AgenticProjectRead:
        row = (
            db.query(models.AgenticOnboardingProject)
            .filter(models.AgenticOnboardingProject.project_id == project_id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agentic project not found.")

        normalized_type = str(artifact_type or "").strip().lower()
        allowed_types = {"mapping_spec", "edi_guideline", "paper_po_sample"}
        if normalized_type not in allowed_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported artifact type '{artifact_type}'.",
            )

        storage_dir = Path("uploads") / row.client_id / "agentic_artifacts" / normalized_type
        storage_dir.mkdir(parents=True, exist_ok=True)
        safe_name = Path(file_name).name or f"{project_id}.bin"
        storage_path = storage_dir / f"{project_id}_{safe_name}"
        storage_path.write_bytes(content)

        file_row = core_models.FileStore(
            client_id=row.client_id,
            original_file_name=safe_name,
            mime_type=mime_type,
            source_channel="AI_ONBOARDING",
            file_path=str(storage_path),
            file_size_bytes=len(content),
            uploaded_by=uploaded_by or "agentic_onboarding",
        )
        db.add(file_row)
        db.flush()

        workflow = _merge_workflow(row)
        artifact_record = {
            "file_id": str(file_row.file_id),
            "file_name": safe_name,
            "mime_type": mime_type,
            "artifact_type": normalized_type,
            "scenario_label": scenario_label,
            "uploaded_at": datetime.utcnow().isoformat(),
        }
        artifacts = self._upsert_artifact(workflow, artifact=artifact_record)
        template_data = None
        if normalized_type in {"mapping_spec", "edi_guideline"}:
            template_data = _parse_business_workbook(content, normalized_type, safe_name)
            if template_data:
                artifact_record["parsed_template"] = True
                artifact_record["template_kind"] = template_data.get("template_kind")
                self._apply_business_template_to_workflow(
                    workflow,
                    artifact_type=normalized_type,
                    template_data=template_data,
                    file_name=safe_name,
                )
                artifacts = self._upsert_artifact(workflow, artifact=artifact_record)

        requirements = dict(workflow.get("requirements_json") or {})
        if normalized_type == "paper_po_sample":
            row.sample_reference = str(file_row.file_id)
            requirements["sample_reference"] = row.sample_reference
            requirements["uploaded_sample_file_name"] = safe_name
            workflow["requirements_json"] = requirements
            self._run_sample_analysis(db, row, workflow)
            artifact_record = self._classify_sample_scenario(artifact_record, workflow)
            artifacts = self._upsert_artifact(workflow, artifact=artifact_record)
            current_actions = list(workflow.get("recommended_actions") or [])
            workflow["recommended_actions"] = [
                f"Review parsed scenario sample '{safe_name}' and confirm the extracted header and line-item structure.",
                f"Scenario classification: {artifact_record.get('scenario_category', 'unclassified').replace('_', ' ')}.",
                *[action for action in current_actions if action],
            ][:6]
        else:
            if not template_data:
                workflow["recommended_actions"] = [
                    f"Use uploaded {normalized_type.replace('_', ' ')} '{safe_name}' as a source document for draft mapping generation.",
                    "Upload one or more representative paper PO samples so the agent can generate and validate the draft map automatically.",
                    "Advance to DRAFT_CONFIGURATION after sample analysis to generate the initial map proposal.",
                ]

        workflow["conversation_summary"] = (
            f"{(workflow.get('conversation_summary') or '').strip()}\n"
            f"Artifact uploaded: {safe_name} ({normalized_type}{f' / {scenario_label}' if scenario_label else ''})"
        ).strip()
        row.discovery_json = workflow
        db.add(row)
        db.commit()
        db.refresh(row)
        return _hydrate_read(row)

    def update_project(self, db: Session, project_id: UUID, payload: schemas.AgenticProjectUpdate) -> schemas.AgenticProjectRead:
        row = (
            db.query(models.AgenticOnboardingProject)
            .filter(models.AgenticOnboardingProject.project_id == project_id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agentic project not found.")

        workflow = _merge_workflow(row)
        updates = payload.model_dump(exclude_unset=True)

        if "objective" in updates:
            workflow["objective"] = updates["objective"]
        if "conversation_summary" in updates:
            workflow["conversation_summary"] = updates["conversation_summary"]
        if "approval_status" in updates:
            workflow["approval_status"] = updates["approval_status"]
        if "recommended_actions" in updates and updates["recommended_actions"] is not None:
            workflow["recommended_actions"] = updates["recommended_actions"]
        if "requirements_json" in updates and updates["requirements_json"] is not None:
            merged = dict(workflow.get("requirements_json") or {})
            merged.update(updates["requirements_json"])
            workflow["requirements_json"] = merged
        if "test_plan_json" in updates and updates["test_plan_json"] is not None:
            workflow["test_plan_json"] = updates["test_plan_json"]
        if "test_results_json" in updates and updates["test_results_json"] is not None:
            workflow["test_results_json"] = updates["test_results_json"]

        if updates.get("extraction_profile_json") is not None:
            row.extraction_profile_json = updates["extraction_profile_json"]
        if updates.get("address_match_profile_json") is not None:
            row.address_match_profile_json = updates["address_match_profile_json"]
        if updates.get("mapping_profile_json") is not None:
            row.mapping_profile_json = updates["mapping_profile_json"]
        if updates.get("rule_profile_json") is not None:
            row.rule_profile_json = updates["rule_profile_json"]

        workflow["progress_steps"] = _set_progress_steps(
            workflow.get("current_stage") or "DISCOVER",
            workflow.get("approval_status") or "PENDING",
        )
        row.discovery_json = workflow
        db.add(row)
        db.commit()
        db.refresh(row)
        return _hydrate_read(row)

    def advance_project(self, db: Session, project_id: UUID, payload: schemas.AgenticProjectAdvance) -> schemas.AgenticProjectRead:
        row = (
            db.query(models.AgenticOnboardingProject)
            .filter(models.AgenticOnboardingProject.project_id == project_id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agentic project not found.")

        workflow = _merge_workflow(row)
        current_stage = workflow.get("current_stage") or "DISCOVER"

        if payload.target_stage:
            if payload.target_stage not in STAGES:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid target stage.")
            next_stage = payload.target_stage
        else:
            current_index = STAGES.index(current_stage) if current_stage in STAGES else 0
            next_stage = STAGES[min(current_index + 1, len(STAGES) - 1)]

        workflow["current_stage"] = next_stage
        if payload.summary_note:
            existing = str(workflow.get("conversation_summary") or "").strip()
            workflow["conversation_summary"] = (
                f"{existing}\n{payload.summary_note}".strip() if existing else payload.summary_note
            )
        if payload.recommended_actions is not None:
            workflow["recommended_actions"] = payload.recommended_actions
        if payload.approval_status:
            workflow["approval_status"] = payload.approval_status

        if next_stage == "ANALYZE_SAMPLE_MESSAGES":
            self._run_sample_analysis(db, row, workflow)
        elif next_stage == "DRAFT_CONFIGURATION":
            self._auto_fill_draft_configuration(db, row, workflow)
        elif next_stage == "TEST_CONNECTIVITY":
            self._run_connectivity_tests(db, row, workflow)
        elif next_stage == "TEST_MESSAGE_PROCESSING":
            self._run_message_processing_tests(db, row, workflow)
        elif next_stage == "ACTIVATE":
            self._activate_project(db, row, workflow)

        workflow["progress_steps"] = _set_progress_steps(
            next_stage,
            workflow.get("approval_status") or "PENDING",
        )
        row.discovery_json = workflow
        row.status = "ACTIVE"
        db.add(row)
        db.commit()
        db.refresh(row)
        return _hydrate_read(row)


agentic_onboarding_service = AgenticOnboardingService()
